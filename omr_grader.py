"""OMR 답안지(OCR100.pdf 양식) 스캔 채점 프로그램."""

import csv
import datetime
import hashlib
import json
import os
import shutil
import sqlite3
import traceback
import uuid
import zipfile

import cv2
import fitz
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill

COLOR_WINDOW = "#F4F7F5"
COLOR_SURFACE = "#FFFFFF"
COLOR_SURFACE_MUTED = "#EAF0ED"
COLOR_INK = "#17332D"
COLOR_INK_MUTED = "#58706A"
COLOR_BORDER = "#D7E2DE"
COLOR_ACCENT = "#1E8E78"
COLOR_ACCENT_DARK = "#146653"
COLOR_ACCENT_PALE = "#DDF3EC"
COLOR_WARNING = "#B46A1B"
COLOR_ERROR = "#B64242"
COLOR_DARK_PANEL = "#17332D"
COLOR_DARK_PANEL_DEEP = "#102A25"
COLOR_DARK_MUTED = "#A9C5BD"
FONT_DISPLAY = ("맑은 고딕", 20, "bold")
FONT_SECTION = ("맑은 고딕", 14, "bold")
FONT_TITLE = ("맑은 고딕", 11, "bold")
FONT_BODY = ("맑은 고딕", 10)
FONT_BODY_BOLD = ("맑은 고딕", 10, "bold")
FONT_SMALL = ("맑은 고딕", 9)
FONT_SMALL_BOLD = ("맑은 고딕", 9, "bold")
FONT_MONO = ("Consolas", 9)
FONT_MONO_BOLD = ("Consolas", 9, "bold")
FONT_BRAND = ("Consolas", 10, "bold")
FONT_RAIL = ("맑은 고딕", 12, "bold")
PATH_WRAP_LENGTH = 520
SPACE_4 = 4
SPACE_2 = 2
SPACE_6 = 6
SPACE_8 = 8
SPACE_10 = 10
SPACE_12 = 12
SPACE_14 = 14
SPACE_16 = 16
SPACE_24 = 24

PAGE_W_PT = 841.0
PAGE_H_PT = 595.0
ZOOM = 3.0

BLOCK_X_STARTS = [239.16, 355.80, 472.44, 589.08, 705.72]
BUBBLE_X_PITCH = 19.44
BUBBLE_SIZE = 11.04
ROW_Y_START = 78.29
ROW_Y_PITCH = 23.7789

ID_COL_X_START = 48.24
ID_COL_PITCH = 19.8
ID_ROW_Y_START = 297.32
ID_ROW_PITCH = 18.7033

TABLE_BORDER_PT = (215.285, 39.579, 797.847, 546.666)

DARK_THRESHOLD = 128
FILL_RATIO_MARK = 0.25
FILL_RATIO_MARGIN = 0.08
DEFAULT_SENSITIVITY = 6


def detection_profile(sensitivity=DEFAULT_SENSITIVITY):
    """민감도(1~10)에 맞는 대비·채움·확신 기준을 반환."""
    level = max(1, min(10, int(sensitivity)))
    progress = (level - 1) / 9
    dark_threshold = round(178 + progress * 42)
    fill_threshold = 0.24 - progress * 0.11
    margin = 0.10 - progress * 0.04
    return dark_threshold, fill_threshold, margin


def preprocess_detection_gray(gray_img):
    """조명 편차와 흐린 마킹을 보정한 8비트 검출용 영상을 반환."""
    denoised = cv2.medianBlur(gray_img, 3)
    background = cv2.GaussianBlur(denoised, (0, 0), 19)
    normalized = cv2.divide(denoised, background, scale=255)
    return cv2.normalize(normalized, None, 0, 255, cv2.NORM_MINMAX)


def answer_bubble_center_pt(qnum, option):
    """qnum: 1~100, option: 1~5. 반환: (x, y) pt 좌표."""
    block = (qnum - 1) // 20
    row = (qnum - 1) % 20
    x = BLOCK_X_STARTS[block] + (option - 1) * BUBBLE_X_PITCH + BUBBLE_SIZE / 2
    y = ROW_Y_START + row * ROW_Y_PITCH + BUBBLE_SIZE / 2
    return x, y


def id_bubble_center_pt(col, digit):
    """col: 0~7 (학번 자리), digit: 0~9. 반환: (x, y) pt 좌표."""
    x = ID_COL_X_START + col * ID_COL_PITCH + BUBBLE_SIZE / 2
    y = ID_ROW_Y_START + digit * ID_ROW_PITCH + BUBBLE_SIZE / 2
    return x, y


def load_answer_key(path):
    """정답표(xlsx/csv, 문항번호+정답 2열)를 읽어 {문항번호: 정답} dict로 반환.
    문항번호는 1부터 연속되어야 하며, 이 dict의 길이가 이번 시험의 문항 수가 된다."""
    ext = os.path.splitext(path)[1].lower()
    rows = []
    if ext == ".csv":
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader, None)  # 헤더 skip
            for row in reader:
                if not row or row[0] == "" or len(row) < 2 or row[1] == "":
                    continue
                rows.append((int(row[0]), int(row[1])))
    else:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None or row[1] is None:
                continue
            rows.append((int(row[0]), int(row[1])))
        wb.close()

    key = dict(rows)
    expected = set(range(1, len(key) + 1))
    if set(key.keys()) != expected:
        raise ValueError(
            f"정답표 문항번호가 1~{len(key)}로 연속되어 있지 않습니다: {sorted(key.keys())}"
        )
    return key


def load_scan_images(paths):
    """이미지 파일(jpg/png) 또는 멀티페이지 PDF 경로 목록을 받아
    (라벨, BGR 이미지) 목록으로 변환. PDF는 페이지마다 한 학생으로 분리."""
    results = []
    for p in paths:
        ext = os.path.splitext(p)[1].lower()
        if ext == ".pdf":
            doc = fitz.open(p)
            base = os.path.splitext(os.path.basename(p))[0]
            for i, page in enumerate(doc):
                pix = page.get_pixmap(matrix=fitz.Matrix(ZOOM, ZOOM), alpha=False)
                img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(
                    pix.height, pix.width, 3
                )
                img_bgr = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
                label = f"{base}_p{i + 1}"
                results.append((label, img_bgr))
        else:
            # cv2.imread는 Windows에서 경로에 한글 등 비ASCII 문자가 있으면
            # 예외 없이 조용히 실패한다. np.fromfile + cv2.imdecode는
            # 파일을 파이썬 자체 API로 열기 때문에 이 문제가 없다.
            data = np.fromfile(p, dtype=np.uint8)
            img_bgr = cv2.imdecode(data, cv2.IMREAD_COLOR)
            if img_bgr is None:
                raise ValueError(f"이미지를 읽을 수 없습니다: {p}")
            label = os.path.basename(p)
            results.append((label, img_bgr))
    return results


def sample_fill_ratio(gray_img, cx_px, cy_px, radius_px, threshold=DARK_THRESHOLD):
    """(cx_px, cy_px) 중심, radius_px 반지름의 원형 영역에서
    DARK_THRESHOLD보다 어두운 픽셀의 비율(0~1)을 반환."""
    x0 = max(int(cx_px - radius_px), 0)
    x1 = min(int(cx_px + radius_px) + 1, gray_img.shape[1])
    y0 = max(int(cy_px - radius_px), 0)
    y1 = min(int(cy_px + radius_px) + 1, gray_img.shape[0])
    patch = gray_img[y0:y1, x0:x1]
    if patch.size == 0:
        return 0.0

    yy, xx = np.ogrid[: patch.shape[0], : patch.shape[1]]
    local_cx = cx_px - x0
    local_cy = cy_px - y0
    mask = (xx - local_cx) ** 2 + (yy - local_cy) ** 2 <= radius_px ** 2
    pixels = patch[mask]
    if pixels.size == 0:
        return 0.0
    dark = np.count_nonzero(pixels < threshold)
    return dark / pixels.size


def detect_marked_index(fill_ratios, threshold=FILL_RATIO_MARK, margin=FILL_RATIO_MARGIN):
    """채움 비율 리스트에서 마킹된 칸의 0-based 인덱스를 판정.
    미응답이면 None, 가장 높은 값이 나머지를 margin 이상 앞서면 확신 있는
    단일 마킹으로 그 인덱스를 반환. 그렇지 않으면(비슷하게 높은 값이
    여럿이면) 애매한 것으로 보고 그 후보들의 인덱스 리스트를 반환한다.
    threshold를 여럿이 넘더라도 그중 하나가 확실히 우세하면 애매하다고
    보지 않는다 — 학번 "0" 자리처럼 인쇄된 글자 자체의 잉크로 threshold를
    넘는 배경 잡음이 있을 수 있기 때문."""
    top = max(fill_ratios)
    if top < threshold:
        return None
    second = sorted(fill_ratios, reverse=True)[1]
    if top - second >= margin:
        return fill_ratios.index(top)
    return sorted(i for i, r in enumerate(fill_ratios) if top - r < margin)


class AlignmentError(Exception):
    pass


def _order_corners(pts):
    s = pts.sum(axis=1)
    diff = np.diff(pts, axis=1).flatten()
    tl = pts[np.argmin(s)]
    br = pts[np.argmax(s)]
    tr = pts[np.argmin(diff)]
    bl = pts[np.argmax(diff)]
    return np.array([tl, tr, br, bl], dtype=np.float32)


def find_table_border(gray_img):
    """스캔 이미지에서 답안표를 감싸는 굵은 외곽 테두리 사각형의
    네 모서리(tl, tr, br, bl)를 검출. 못 찾으면 None."""
    target_ratio = (TABLE_BORDER_PT[2] - TABLE_BORDER_PT[0]) / (
        TABLE_BORDER_PT[3] - TABLE_BORDER_PT[1]
    )
    img_area = gray_img.shape[0] * gray_img.shape[1]

    best = None
    best_area = 0
    blur = cv2.GaussianBlur(gray_img, (5, 5), 0)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8)).apply(blur)
    binaries = [
        cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1],
        cv2.adaptiveThreshold(
            clahe, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 51, 9
        ),
        cv2.Canny(clahe, 40, 140),
    ]
    kernel = np.ones((5, 5), np.uint8)
    for binary in binaries:
        closed = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel, iterations=2)
        dilated = cv2.dilate(closed, kernel, iterations=1)
        contours, _ = cv2.findContours(dilated, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
        for c in contours:
            area = cv2.contourArea(c)
            if area < img_area * 0.08:
                continue
            peri = cv2.arcLength(c, True)
            approx = cv2.approxPolyDP(c, 0.03 * peri, True)
            if len(approx) != 4:
                continue
            pts = approx.reshape(4, 2).astype(np.float32)
            ordered = _order_corners(pts)
            w = np.linalg.norm(ordered[1] - ordered[0])
            h = np.linalg.norm(ordered[3] - ordered[0])
            if h == 0 or w <= h:
                continue
            ratio_error = abs((w / h) - target_ratio) / target_ratio
            if ratio_error > 0.18:
                continue
            if area > best_area:
                best_area = area
                best = ordered
    return best


_ROTATIONS = (None, cv2.ROTATE_90_CLOCKWISE, cv2.ROTATE_180, cv2.ROTATE_90_COUNTERCLOCKWISE)

# 표지 제목("동명대학교...", "20__년 중간·기말고사") 텍스트가 찍히는 영역(pt 좌표).
# 답안표 테두리는 180도 회전해도 사각형 비율이 같아 그것만으로는 상하 반전을 구분할 수
# 없으므로, 진하게 인쇄된 이 표지 텍스트 영역과 그 반대편(대각선 대칭 위치)의 잉크
# 밀도를 비교해 상하가 뒤집혔는지 판별한다.
_HEADER_TEXT_PT = (41.83, 40.06, 200.71, 82.9)


def _is_upside_down(aligned_gray):
    x0, y0, x1, y1 = (int(c * ZOOM) for c in _HEADER_TEXT_PT)
    header_patch = aligned_gray[y0:y1, x0:x1]

    mx0 = PAGE_W_PT - _HEADER_TEXT_PT[2]
    my0 = PAGE_H_PT - _HEADER_TEXT_PT[3]
    mx1 = PAGE_W_PT - _HEADER_TEXT_PT[0]
    my1 = PAGE_H_PT - _HEADER_TEXT_PT[1]
    mx0, my0, mx1, my1 = (int(c * ZOOM) for c in (mx0, my0, mx1, my1))
    mirror_patch = aligned_gray[my0:my1, mx0:mx1]

    if header_patch.size == 0 or mirror_patch.size == 0:
        return False

    header_dark = np.mean(header_patch < DARK_THRESHOLD)
    mirror_dark = np.mean(mirror_patch < DARK_THRESHOLD)
    return mirror_dark > header_dark


def align_sheet(img_bgr):
    """스캔 이미지를 답안표 테두리 기준으로 투시 보정해
    (PAGE_H_PT*ZOOM, PAGE_W_PT*ZOOM) 크기의 정렬된 이미지로 반환.
    스캐너가 답안지를 90/180/270도 회전된 방향으로 스캔한 경우까지 시도하고,
    표지 텍스트 위치를 근거로 상하 반전 여부까지 보정한다."""
    try:
        dst = np.array(
            [
                [TABLE_BORDER_PT[0] * ZOOM, TABLE_BORDER_PT[1] * ZOOM],
                [TABLE_BORDER_PT[2] * ZOOM, TABLE_BORDER_PT[1] * ZOOM],
                [TABLE_BORDER_PT[2] * ZOOM, TABLE_BORDER_PT[3] * ZOOM],
                [TABLE_BORDER_PT[0] * ZOOM, TABLE_BORDER_PT[3] * ZOOM],
            ],
            dtype=np.float32,
        )
        out_size = (int(PAGE_W_PT * ZOOM), int(PAGE_H_PT * ZOOM))

        for rotate_code in _ROTATIONS:
            candidate = img_bgr if rotate_code is None else cv2.rotate(img_bgr, rotate_code)
            gray = cv2.cvtColor(candidate, cv2.COLOR_BGR2GRAY)
            corners = find_table_border(gray)
            if corners is None:
                continue
            M = cv2.getPerspectiveTransform(corners, dst)
            aligned = cv2.warpPerspective(candidate, M, out_size, borderValue=(255, 255, 255))
            aligned_gray = cv2.cvtColor(aligned, cv2.COLOR_BGR2GRAY)
            if _is_upside_down(aligned_gray):
                # 캔버스 전체를 뒤집으면 이미 dst 위치에 맞춰 배치된 표가 반대쪽
                # 구석으로 밀려나므로, 원본 후보를 180도 돌려 테두리 검출/정렬을
                # 처음부터 다시 수행한다.
                candidate180 = cv2.rotate(candidate, cv2.ROTATE_180)
                gray180 = cv2.cvtColor(candidate180, cv2.COLOR_BGR2GRAY)
                corners180 = find_table_border(gray180)
                if corners180 is not None:
                    M180 = cv2.getPerspectiveTransform(corners180, dst)
                    aligned = cv2.warpPerspective(
                        candidate180, M180, out_size, borderValue=(255, 255, 255)
                    )
            return aligned

        raise AlignmentError("답안표 테두리를 찾지 못했습니다")
    except AlignmentError:
        raise
    except Exception as e:
        raise AlignmentError(f"답안지 정렬 중 오류가 발생했습니다: {e}") from e


BLANK_LABEL = "미응답"


def multi_label(options):
    return "중복(" + ",".join(str(o) for o in sorted(options)) + ")"


def recognize_sheet(img_bgr, num_questions, sensitivity=DEFAULT_SENSITIVITY, aligned=None):
    """답안지 스캔 이미지 1장을 인식해 학번과 num_questions개 문항의 답을 반환."""
    if aligned is None:
        aligned = align_sheet(img_bgr)
    gray = cv2.cvtColor(aligned, cv2.COLOR_BGR2GRAY)
    detection_gray = preprocess_detection_gray(gray)
    dark_threshold, fill_threshold, margin = detection_profile(sensitivity)
    radius_px = (BUBBLE_SIZE / 2 * 0.75) * ZOOM

    digits = []
    id_flagged_cols = []
    for col in range(8):
        ratios = []
        for digit in range(10):
            cx, cy = id_bubble_center_pt(col, digit)
            ratios.append(
                sample_fill_ratio(
                    detection_gray, cx * ZOOM, cy * ZOOM, radius_px, dark_threshold
                )
            )
        idx = detect_marked_index(ratios, fill_threshold, margin)
        if idx is None or isinstance(idx, list):
            digits.append("?")
            id_flagged_cols.append(col)
        else:
            digits.append(str(idx))
    student_id = "".join(digits)

    answers = {}
    flagged_questions = []
    for q in range(1, num_questions + 1):
        ratios = []
        for option in range(1, 6):
            cx, cy = answer_bubble_center_pt(q, option)
            ratios.append(
                sample_fill_ratio(
                    detection_gray, cx * ZOOM, cy * ZOOM, radius_px, dark_threshold
                )
            )
        idx = detect_marked_index(ratios, fill_threshold, margin)
        if idx is None:
            answers[q] = BLANK_LABEL
            flagged_questions.append(q)
        elif isinstance(idx, list):
            answers[q] = multi_label([i + 1 for i in idx])
            flagged_questions.append(q)
        else:
            answers[q] = idx + 1

    return {
        "student_id": student_id,
        "id_flagged_cols": id_flagged_cols,
        "answers": answers,
        "flagged_questions": flagged_questions,
    }


def grade_sheet(answers, answer_key):
    """인식된 답안과 정답표를 비교해 (점수, 오답문항번호 리스트)를 반환.
    미응답/중복 등 int가 아닌 답은 항상 오답으로 처리한다."""
    score = 0
    wrong = []
    for q, correct in answer_key.items():
        given = answers.get(q, BLANK_LABEL)
        if isinstance(given, int) and given == correct:
            score += 1
        else:
            wrong.append(q)
    return score, wrong


def _utc_now():
    return datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds")


def _file_hash(path):
    digest = hashlib.sha256()
    with open(path, "rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def default_history_dir():
    """사용자별 로컬 데이터 폴더를 반환하며 관리자 권한을 요구하지 않는다."""
    local_appdata = os.environ.get("LOCALAPPDATA")
    if local_appdata:
        return os.path.join(local_appdata, "OMRGrader")
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), ".omr_grader_data")


class RunStore:
    """채점 실행, 원본 연결, 진단 로그와 수정 결과를 로컬 SQLite에 보관."""

    def __init__(self, root_dir):
        self.root_dir = os.path.abspath(os.fspath(root_dir))
        self.runs_dir = os.path.join(self.root_dir, "runs")
        self.logs_dir = os.path.join(self.root_dir, "logs")
        self.db_path = os.path.join(self.root_dir, "history.sqlite3")
        os.makedirs(self.runs_dir, exist_ok=True)
        os.makedirs(self.logs_dir, exist_ok=True)
        self._init_db()

    def _connect(self):
        connection = sqlite3.connect(self.db_path)
        connection.row_factory = sqlite3.Row
        return connection

    def _init_db(self):
        with self._connect() as connection:
            connection.executescript(
                """
                PRAGMA journal_mode=WAL;
                CREATE TABLE IF NOT EXISTS runs (
                    id TEXT PRIMARY KEY,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    status TEXT NOT NULL,
                    scan_count INTEGER NOT NULL,
                    answer_key_path TEXT NOT NULL,
                    answer_key_hash TEXT NOT NULL,
                    sensitivity INTEGER NOT NULL,
                    output_dir TEXT NOT NULL,
                    result_path TEXT,
                    last_label TEXT,
                    error_message TEXT,
                    failed_labels_json TEXT NOT NULL DEFAULT '[]',
                    source_manifest_json TEXT NOT NULL DEFAULT '{}'
                );
                CREATE TABLE IF NOT EXISTS records (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    run_id TEXT NOT NULL REFERENCES runs(id),
                    label TEXT NOT NULL,
                    source_ref TEXT NOT NULL,
                    student_id TEXT NOT NULL,
                    answers_json TEXT NOT NULL,
                    reanalyzed_answers_json TEXT,
                    corrected_answers_json TEXT,
                    score INTEGER NOT NULL,
                    wrong_json TEXT NOT NULL,
                    flagged_json TEXT NOT NULL,
                    id_flagged_json TEXT NOT NULL,
                    review_status TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS diagnostic_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    run_id TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    level TEXT NOT NULL,
                    event TEXT NOT NULL,
                    message TEXT NOT NULL,
                    details_json TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_runs_created ON runs(created_at DESC);
                CREATE INDEX IF NOT EXISTS idx_records_run ON records(run_id);
                CREATE INDEX IF NOT EXISTS idx_logs_run ON diagnostic_logs(run_id, created_at);
                """
            )
            columns = {row["name"] for row in connection.execute("PRAGMA table_info(records)").fetchall()}
            if "reanalyzed_answers_json" not in columns:
                connection.execute("ALTER TABLE records ADD COLUMN reanalyzed_answers_json TEXT")

    @staticmethod
    def _decode_answers(value):
        return {int(key): answer for key, answer in json.loads(value).items()}

    @staticmethod
    def _decode_run(row):
        if row is None:
            raise ValueError("채점 실행을 찾을 수 없습니다")
        result = dict(row)
        result["failed_labels"] = json.loads(result.pop("failed_labels_json"))
        result["source_manifest"] = json.loads(result.pop("source_manifest_json"))
        return result

    @classmethod
    def _decode_record(cls, row):
        result = dict(row)
        result["answers"] = cls._decode_answers(result.pop("answers_json"))
        reanalyzed = result.pop("reanalyzed_answers_json")
        result["reanalyzed_answers"] = cls._decode_answers(reanalyzed) if reanalyzed else None
        corrected = result.pop("corrected_answers_json")
        result["corrected_answers"] = cls._decode_answers(corrected) if corrected else None
        result["wrong"] = json.loads(result.pop("wrong_json"))
        result["flagged_questions"] = json.loads(result.pop("flagged_json"))
        result["id_flagged_cols"] = json.loads(result.pop("id_flagged_json"))
        return result

    @staticmethod
    def _effective_answers(record):
        return record["corrected_answers"] or record["reanalyzed_answers"] or record["answers"]

    def start_run(self, run_id, scan_count, answer_key_path, answer_key_hash, sensitivity, output_dir):
        now = _utc_now()
        with self._connect() as connection:
            connection.execute(
                """INSERT INTO runs
                (id, created_at, updated_at, status, scan_count, answer_key_path,
                 answer_key_hash, sensitivity, output_dir)
                VALUES (?, ?, ?, 'processing', ?, ?, ?, ?, ?)""",
                (
                    run_id,
                    now,
                    now,
                    scan_count,
                    os.path.abspath(answer_key_path),
                    answer_key_hash,
                    sensitivity,
                    os.path.abspath(output_dir),
                ),
            )
        self.add_log(run_id, "info", "run.started", "채점 실행을 시작했습니다", {"scan_count": scan_count})

    def set_source_manifest(self, run_id, manifest):
        with self._connect() as connection:
            connection.execute(
                "UPDATE runs SET source_manifest_json = ?, updated_at = ? WHERE id = ?",
                (json.dumps(manifest, ensure_ascii=False), _utc_now(), run_id),
            )

    def add_record(self, run_id, record, source_ref):
        review_status = "검토 필요" if record["flagged_questions"] or record["id_flagged_cols"] else "정상"
        with self._connect() as connection:
            connection.execute(
                """INSERT INTO records
                (run_id, label, source_ref, student_id, answers_json, score, wrong_json,
                 flagged_json, id_flagged_json, review_status, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    run_id,
                    record["label"],
                    source_ref,
                    record["student_id"],
                    json.dumps(record["answers"], ensure_ascii=False),
                    record["score"],
                    json.dumps(record["wrong"], ensure_ascii=False),
                    json.dumps(record["flagged_questions"], ensure_ascii=False),
                    json.dumps(record["id_flagged_cols"], ensure_ascii=False),
                    review_status,
                    _utc_now(),
                ),
            )

    def update_progress(self, run_id, label):
        with self._connect() as connection:
            connection.execute(
                "UPDATE runs SET last_label = ?, updated_at = ? WHERE id = ?",
                (label, _utc_now(), run_id),
            )

    def finish_run(self, run_id, result_path, failed_labels=None):
        failed_labels = failed_labels or []
        with self._connect() as connection:
            connection.execute(
                """UPDATE runs SET status = 'completed', result_path = ?, failed_labels_json = ?,
                updated_at = ?, error_message = NULL WHERE id = ?""",
                (result_path, json.dumps(failed_labels, ensure_ascii=False), _utc_now(), run_id),
            )
        self.add_log(run_id, "info", "run.completed", "채점 실행이 완료되었습니다", {"result_path": result_path})

    def fail_run(self, run_id, error_message):
        with self._connect() as connection:
            connection.execute(
                """UPDATE runs SET status = 'failed', error_message = ?, updated_at = ?
                WHERE id = ?""",
                (error_message, _utc_now(), run_id),
            )
        self.add_log(run_id, "error", "run.failed", error_message, {"traceback": traceback.format_exc()})

    def mark_interrupted_runs(self):
        with self._connect() as connection:
            rows = connection.execute("SELECT id FROM runs WHERE status = 'processing'").fetchall()
            connection.execute(
                "UPDATE runs SET status = 'interrupted', updated_at = ? WHERE status = 'processing'",
                (_utc_now(),),
            )
        for row in rows:
            self.add_log(row["id"], "warn", "run.interrupted", "이전 실행이 비정상 종료된 것으로 보입니다", {})
        return len(rows)

    def add_log(self, run_id, level, event, message, details=None):
        details = details or {}
        created_at = _utc_now()
        payload = {
            "created_at": created_at,
            "run_id": run_id,
            "level": level,
            "event": event,
            "message": message,
            "details": details,
        }
        with open(os.path.join(self.logs_dir, f"{run_id}.jsonl"), "a", encoding="utf-8") as log_file:
            log_file.write(json.dumps(payload, ensure_ascii=False) + "\n")
        with self._connect() as connection:
            connection.execute(
                """INSERT INTO diagnostic_logs
                (run_id, created_at, level, event, message, details_json)
                VALUES (?, ?, ?, ?, ?, ?)""",
                (run_id, created_at, level, event, message, json.dumps(details, ensure_ascii=False)),
            )

    def get_run(self, run_id):
        with self._connect() as connection:
            row = connection.execute("SELECT * FROM runs WHERE id = ?", (run_id,)).fetchone()
        return self._decode_run(row)

    def list_runs(self, query="", page=1, page_size=50):
        query = f"%{query.strip()}%"
        offset = max(0, page - 1) * page_size
        with self._connect() as connection:
            rows = connection.execute(
                """SELECT * FROM runs
                WHERE id LIKE ? OR answer_key_path LIKE ? OR status LIKE ?
                ORDER BY created_at DESC LIMIT ? OFFSET ?""",
                (query, query, query, page_size, offset),
            ).fetchall()
        return [self._decode_run(row) for row in rows]

    def get_records(self, run_id):
        with self._connect() as connection:
            rows = connection.execute(
                "SELECT * FROM records WHERE run_id = ? ORDER BY id", (run_id,)
            ).fetchall()
        return [self._decode_record(row) for row in rows]

    def get_logs(self, run_id):
        with self._connect() as connection:
            rows = connection.execute(
                "SELECT * FROM diagnostic_logs WHERE run_id = ? ORDER BY id", (run_id,)
            ).fetchall()
        return [dict(row, details=json.loads(row["details_json"])) for row in rows]

    def update_manual_answer(self, run_id, record_id, qnum, answer, reason):
        if not isinstance(answer, int) or answer not in range(1, 6):
            raise ValueError("수동 답안은 1~5 중 하나여야 합니다")
        run = self.get_run(run_id)
        answer_key = load_answer_key(run["answer_key_path"])
        if qnum not in answer_key:
            raise ValueError(f"정답표에 문항 {qnum}이 없습니다")
        with self._connect() as connection:
            row = connection.execute(
                "SELECT * FROM records WHERE id = ? AND run_id = ?", (record_id, run_id)
            ).fetchone()
            record = self._decode_record(row)
            corrected = record["corrected_answers"] or record["reanalyzed_answers"] or dict(record["answers"])
            corrected[qnum] = answer
            score, wrong = grade_sheet(corrected, answer_key)
            connection.execute(
                """UPDATE records SET corrected_answers_json = ?, score = ?, wrong_json = ?,
                review_status = '수정됨', updated_at = ? WHERE id = ? AND run_id = ?""",
                (
                    json.dumps(corrected, ensure_ascii=False),
                    score,
                    json.dumps(wrong, ensure_ascii=False),
                    _utc_now(),
                    record_id,
                    run_id,
                ),
            )
        self.add_log(
            run_id,
            "info",
            "record.manual_correction",
            "수동 답안을 적용했습니다",
            {"record_id": record_id, "question": qnum, "answer": answer, "reason": reason},
        )
        return next(record for record in self.get_records(run_id) if record["id"] == record_id)

    def regrade(self, run_id, answer_key_path, apply=False):
        run = self.get_run(run_id)
        answer_key = load_answer_key(answer_key_path)
        changes = []
        records = self.get_records(run_id)
        for record in records:
            answers = self._effective_answers(record)
            score_after, wrong_after = grade_sheet(answers, answer_key)
            if score_after != record["score"] or wrong_after != record["wrong"]:
                changes.append(
                    {
                        "record_id": record["id"],
                        "student_id": record["student_id"],
                        "score_before": record["score"],
                        "score_after": score_after,
                        "wrong_before": record["wrong"],
                        "wrong_after": wrong_after,
                    }
                )
        if not apply:
            return changes
        with self._connect() as connection:
            for record in records:
                answers = self._effective_answers(record)
                score, wrong = grade_sheet(answers, answer_key)
                connection.execute(
                    "UPDATE records SET score = ?, wrong_json = ?, updated_at = ? WHERE id = ?",
                    (score, json.dumps(wrong), _utc_now(), record["id"]),
                )
            connection.execute(
                """UPDATE runs SET answer_key_path = ?, answer_key_hash = ?, updated_at = ?
                WHERE id = ?""",
                (os.path.abspath(answer_key_path), _file_hash(answer_key_path), _utc_now(), run_id),
            )
        self.add_log(
            run_id,
            "info",
            "run.regraded",
            "새 정답표로 재채점을 적용했습니다",
            {"change_count": len(changes), "answer_key_path": answer_key_path},
        )
        return changes

    def export_run(self, run_id, output_path):
        run = self.get_run(run_id)
        answer_key = load_answer_key(run["answer_key_path"])
        records = []
        for record in self.get_records(run_id):
            answers = self._effective_answers(record)
            records.append(
                {
                    "label": record["label"],
                    "student_id": record["student_id"],
                    "answers": answers,
                    "score": record["score"],
                    "wrong": record["wrong"],
                    "flagged_questions": record["flagged_questions"],
                    "id_flagged_cols": record["id_flagged_cols"],
                }
            )
        write_result_excel(output_path, records, answer_key, run["failed_labels"])
        self.add_log(run_id, "info", "run.exported", "결과 파일을 다시 내보냈습니다", {"path": output_path})
        return output_path

    def reanalyze(self, run_id, sensitivity, apply=False):
        run = self.get_run(run_id)
        answer_key = load_answer_key(run["answer_key_path"])
        records = {record["label"]: record for record in self.get_records(run_id)}
        changes = []
        pending = []
        for relative_path in run["source_manifest"].values():
            source_path = os.path.join(self.root_dir, relative_path)
            if not os.path.isfile(source_path):
                self.add_log(run_id, "warn", "reanalyze.source_missing", "재분석 원본을 찾지 못했습니다", {"path": relative_path})
                continue
            try:
                scans = load_scan_images([source_path])
            except (OSError, ValueError, RuntimeError) as error:
                self.add_log(run_id, "warn", "reanalyze.load_failed", str(error), {"path": relative_path})
                continue
            for label, image in scans:
                record = records.get(label)
                if record is None:
                    continue
                try:
                    aligned = align_sheet(image)
                    recognition = recognize_sheet(image, len(answer_key), sensitivity=sensitivity, aligned=aligned)
                except AlignmentError as error:
                    self.add_log(run_id, "warn", "reanalyze.alignment_failed", str(error), {"label": label})
                    continue
                new_answers = recognition["answers"]
                old_answers = self._effective_answers(record)
                score_after, wrong_after = grade_sheet(new_answers, answer_key)
                if new_answers != old_answers or recognition["flagged_questions"] != record["flagged_questions"]:
                    changes.append(
                        {
                            "record_id": record["id"],
                            "student_id": record["student_id"],
                            "label": label,
                            "score_before": record["score"],
                            "score_after": score_after,
                            "answers_before": old_answers,
                            "answers_after": new_answers,
                        }
                    )
                pending.append((record, recognition, score_after, wrong_after, aligned))
        if not apply:
            return changes
        reanalysis_dir = os.path.join(self.runs_dir, run_id, "debug_reanalysis")
        os.makedirs(reanalysis_dir, exist_ok=True)
        with self._connect() as connection:
            for record, recognition, score, wrong, aligned in pending:
                effective_score, effective_wrong = score, wrong
                if record["corrected_answers"]:
                    effective_score, effective_wrong = grade_sheet(record["corrected_answers"], answer_key)
                connection.execute(
                    """UPDATE records SET reanalyzed_answers_json = ?, score = ?, wrong_json = ?,
                    flagged_json = ?, id_flagged_json = ?, review_status = ?, updated_at = ? WHERE id = ?""",
                    (
                        json.dumps(recognition["answers"], ensure_ascii=False),
                        effective_score,
                        json.dumps(effective_wrong, ensure_ascii=False),
                        json.dumps(recognition["flagged_questions"], ensure_ascii=False),
                        json.dumps(recognition["id_flagged_cols"], ensure_ascii=False),
                        "수정됨" if record["corrected_answers"] else "재분석됨",
                        _utc_now(),
                        record["id"],
                    ),
                )
                save_debug_overlay(
                    aligned,
                    recognition,
                    os.path.join(reanalysis_dir, f"{record['id']:04d}_{record['label']}.png"),
                )
            connection.execute(
                "UPDATE runs SET sensitivity = ?, updated_at = ? WHERE id = ?",
                (sensitivity, _utc_now(), run_id),
            )
        self.add_log(
            run_id,
            "info",
            "run.reanalyzed",
            "원본을 새 민감도로 재분석했습니다",
            {"change_count": len(changes), "sensitivity": sensitivity},
        )
        return changes

    def backup_run(self, run_id, destination):
        run = self.get_run(run_id)
        run_dir = os.path.join(self.runs_dir, run_id)
        with zipfile.ZipFile(destination, "w", zipfile.ZIP_DEFLATED) as archive:
            if os.path.isdir(run_dir):
                for root, _dirs, files in os.walk(run_dir):
                    for name in files:
                        path = os.path.join(root, name)
                        archive.write(path, os.path.relpath(path, self.root_dir))
            log_path = os.path.join(self.logs_dir, f"{run_id}.jsonl")
            if os.path.isfile(log_path):
                archive.write(log_path, os.path.relpath(log_path, self.root_dir))
            if run["result_path"] and os.path.isfile(run["result_path"]):
                archive.write(run["result_path"], os.path.join("results", os.path.basename(run["result_path"])))
            archive.writestr("run_metadata.json", json.dumps(run, ensure_ascii=False, indent=2))
        self.add_log(run_id, "info", "run.backup", "채점 실행을 백업했습니다", {"path": os.fspath(destination)})
        return destination

    def delete_run(self, run_id):
        self.get_run(run_id)
        trash_root = os.path.join(self.root_dir, "trash")
        os.makedirs(trash_root, exist_ok=True)
        trash_path = os.path.join(trash_root, f"{run_id}_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}")
        run_dir = os.path.join(self.runs_dir, run_id)
        if os.path.isdir(run_dir):
            shutil.move(run_dir, trash_path)
        log_path = os.path.join(self.logs_dir, f"{run_id}.jsonl")
        if os.path.isfile(log_path):
            shutil.move(log_path, os.path.join(trash_root, f"{run_id}.jsonl"))
        with self._connect() as connection:
            connection.execute("DELETE FROM records WHERE run_id = ?", (run_id,))
            connection.execute("DELETE FROM diagnostic_logs WHERE run_id = ?", (run_id,))
            connection.execute("DELETE FROM runs WHERE id = ?", (run_id,))
        return trash_path

    def analyze(self, run_id):
        run = self.get_run(run_id)
        answer_key = load_answer_key(run["answer_key_path"])
        records = self.get_records(run_id)
        question_stats = {}
        for q, correct in answer_key.items():
            option_counts = {str(option): 0 for option in range(1, 6)}
            correct_count = 0
            for record in records:
                answer = self._effective_answers(record).get(q, BLANK_LABEL)
                if isinstance(answer, int):
                    option_counts[str(answer)] += 1
                    correct_count += answer == correct
            question_stats[q] = {
                "correct": correct_count,
                "total": len(records),
                "correct_rate": correct_count / len(records) if records else 0,
                "option_counts": option_counts,
            }
        scores = [record["score"] for record in records]
        return {
            "run_id": run_id,
            "student_count": len(records),
            "average_score": sum(scores) / len(scores) if scores else 0,
            "max_score": max(scores, default=0),
            "min_score": min(scores, default=0),
            "question_stats": question_stats,
        }


WRONG_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def write_result_excel(path, student_records, answer_key, failed_labels):
    """채점 결과를 엑셀로 저장. 시트 '채점결과'(학생별 1행) + '정렬실패'(실패 파일 목록)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "채점결과"

    qnums = sorted(answer_key.keys())
    header = ["학번"] + [f"Q{q}" for q in qnums] + ["점수", "오답문항", "확인필요"]
    ws.append(header)

    for rec in student_records:
        wrong_set = set(rec["wrong"])
        row = [rec["student_id"]]
        for q in qnums:
            row.append(rec["answers"].get(q, BLANK_LABEL))
        row.append(rec["score"])
        row.append(",".join(str(q) for q in rec["wrong"]))
        flagged = sorted(set(rec["flagged_questions"]))
        row.append(",".join(f"Q{q}" for q in flagged))
        ws.append(row)

        r = ws.max_row
        for i, q in enumerate(qnums):
            if q in wrong_set:
                ws.cell(row=r, column=2 + i).fill = WRONG_FILL

    fail_ws = wb.create_sheet("정렬실패")
    fail_ws.append(["파일명"])
    for label in failed_labels:
        fail_ws.append([label])

    wb.save(path)


def save_debug_overlay(aligned_img_bgr, recognition, out_path):
    """정렬된 이미지 위에 인식된 마킹 위치를 원으로 표시해 저장.
    확인필요로 플래그된 문항/학번 자리는 다른 색으로 강조."""
    overlay = aligned_img_bgr.copy()
    radius_px = int(BUBBLE_SIZE / 2 * ZOOM)

    for col in range(8):
        digit_char = recognition["student_id"][col]
        color = (0, 0, 255) if col in recognition["id_flagged_cols"] else (0, 200, 0)
        if digit_char == "?":
            continue
        cx, cy = id_bubble_center_pt(col, int(digit_char))
        cv2.circle(overlay, (int(cx * ZOOM), int(cy * ZOOM)), radius_px, color, 2)

    for q, value in recognition["answers"].items():
        flagged = q in recognition["flagged_questions"]
        color = (0, 0, 255) if flagged else (0, 200, 0)
        if isinstance(value, int):
            cx, cy = answer_bubble_center_pt(q, value)
            cv2.circle(overlay, (int(cx * ZOOM), int(cy * ZOOM)), radius_px, color, 2)
        # 미응답/중복은 원으로 표시할 단일 위치가 없으므로 그리지 않음
        # (엑셀의 확인필요 열 + 이 오버레이의 빨간 원 부재 자체가 "이상함" 신호)

    # cv2.imwrite는 Windows에서 경로에 한글 등 비ASCII 문자가 있으면 예외 없이
    # 조용히 실패한다(파일이 안 만들어지는데 에러도 없음). cv2.imencode로 메모리
    # 상에서 인코딩한 뒤 파이썬 표준 파일 API로 쓰면 이 문제가 없다.
    ok, buf = cv2.imencode(".png", overlay)
    if ok:
        with open(out_path, "wb") as f:
            f.write(buf.tobytes())


def _snapshot_sources(store, run_id, scan_paths):
    destination = os.path.join(store.runs_dir, run_id, "originals")
    os.makedirs(destination, exist_ok=True)
    manifest = {}
    for index, path in enumerate(scan_paths, start=1):
        absolute_path = os.path.abspath(path)
        if not os.path.isfile(absolute_path):
            continue
        name = f"{index:04d}_{os.path.basename(absolute_path)}"
        copied_path = os.path.join(destination, name)
        shutil.copy2(absolute_path, copied_path)
        manifest[absolute_path] = os.path.relpath(copied_path, store.root_dir)
    return manifest


def _source_ref_for_label(label, scan_paths, manifest):
    for path in scan_paths:
        base = os.path.splitext(os.path.basename(path))[0]
        if label == os.path.basename(path) or label.startswith(f"{base}_p"):
            return manifest.get(os.path.abspath(path), "")
    return ""


def run_pipeline(scan_paths, answer_key_path, output_dir, sensitivity=DEFAULT_SENSITIVITY):
    """스캔 파일 목록 + 정답표 경로를 받아 결과와 영구 실행 기록을 생성."""
    os.makedirs(output_dir, exist_ok=True)
    debug_dir = os.path.join(output_dir, "debug")
    os.makedirs(debug_dir, exist_ok=True)
    store = RunStore(default_history_dir())
    store.mark_interrupted_runs()
    run_id = uuid.uuid4().hex
    answer_key_hash = _file_hash(answer_key_path) if os.path.isfile(answer_key_path) else ""
    store.start_run(run_id, len(scan_paths), answer_key_path, answer_key_hash, sensitivity, output_dir)

    try:
        answer_key = load_answer_key(answer_key_path)
        num_questions = len(answer_key)
        manifest = _snapshot_sources(store, run_id, scan_paths)
        store.set_source_manifest(run_id, manifest)
        scans = load_scan_images(scan_paths)

        records = []
        failed_labels = []
        for i, (label, img_bgr) in enumerate(scans):
            store.update_progress(run_id, label)
            try:
                aligned = align_sheet(img_bgr)
            except AlignmentError as error:
                failed_labels.append(label)
                store.add_log(run_id, "warn", "scan.alignment_failed", str(error), {"label": label})
                continue

            recognition = recognize_sheet(
                img_bgr, num_questions, sensitivity=sensitivity, aligned=aligned
            )
            score, wrong = grade_sheet(recognition["answers"], answer_key)
            record = {
                "label": label,
                "student_id": recognition["student_id"],
                "answers": recognition["answers"],
                "score": score,
                "wrong": wrong,
                "flagged_questions": recognition["flagged_questions"],
                "id_flagged_cols": recognition["id_flagged_cols"],
            }
            records.append(record)
            store.add_record(run_id, record, _source_ref_for_label(label, scan_paths, manifest))

            safe_name = f"{i:04d}_{label.replace('.', '_')}.png"
            save_debug_overlay(aligned, recognition, os.path.join(debug_dir, safe_name))

        result_path = os.path.join(output_dir, "채점결과.xlsx")
        write_result_excel(result_path, records, answer_key, failed_labels)
        with open(os.path.join(output_dir, "omr_run_id.txt"), "w", encoding="utf-8") as run_file:
            run_file.write(run_id)
        store.finish_run(run_id, result_path, failed_labels)
        return result_path
    except Exception as error:  # noqa: BROAD_EXCEPT_OK - pipeline boundary records and re-raises
        store.fail_run(run_id, str(error))
        raise


import queue
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk


def _run_pipeline_worker(
    scan_paths, key_path, output_dir, sensitivity, log_queue, done_queue
):
    try:
        result_path = run_pipeline(scan_paths, key_path, output_dir, sensitivity)
        log_queue.put(f"완료: {result_path}")
        done_queue.put(("ok", result_path))
    except Exception as e:  # noqa: BLE001 - GUI 최상위 경계, 사용자에게 그대로 보여줌
        log_queue.put(f"오류: {e}")
        done_queue.put(("error", str(e)))


class LegacyApp:
    def __init__(self, root):
        self.root = root
        self.root.title("OMR 답안지 채점")
        self.root.geometry("920x800")
        self.root.minsize(820, 720)
        self.root.configure(bg=COLOR_WINDOW)
        self.scan_paths = []
        self.key_path = None
        self.log_queue = queue.Queue()
        self.done_queue = queue.Queue()
        self.store = RunStore(default_history_dir())
        self.interrupted_count = self.store.mark_interrupted_runs()
        self.sensitivity_var = tk.DoubleVar(value=DEFAULT_SENSITIVITY)

        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "Action.TButton",
            background=COLOR_ACCENT_DARK,
            foreground=COLOR_SURFACE,
            borderwidth=0,
            padding=(SPACE_14, SPACE_8),
            font=FONT_BODY_BOLD,
        )
        style.map(
            "Action.TButton",
            background=[
                ("disabled", COLOR_SURFACE_MUTED),
                ("active", COLOR_INK),
                ("pressed", COLOR_INK),
            ],
            foreground=[("disabled", COLOR_INK_MUTED)],
        )
        style.configure(
            "Secondary.TButton",
            background=COLOR_SURFACE_MUTED,
            foreground=COLOR_INK,
            borderwidth=0,
            padding=(SPACE_12, SPACE_8),
            font=FONT_SMALL_BOLD,
        )
        style.map("Secondary.TButton", background=[("active", COLOR_BORDER)])
        style.configure(
            "Modern.Horizontal.TScale",
            troughcolor=COLOR_SURFACE_MUTED,
            background=COLOR_ACCENT,
        )

        header = tk.Frame(root, bg=COLOR_DARK_PANEL, padx=SPACE_24, pady=SPACE_12)
        header.pack(fill="x")
        self.history_button = tk.Button(
            header,
            text="채점 이력",
            command=self.open_history,
            bg=COLOR_ACCENT_DARK,
            fg=COLOR_SURFACE,
            activebackground=COLOR_ACCENT,
            activeforeground=COLOR_SURFACE,
            relief="flat",
            padx=SPACE_12,
            pady=SPACE_6,
            font=FONT_SMALL_BOLD,
        )
        self.history_button.pack(side="right", anchor="n")
        tk.Label(
            header,
            text="OMR GRADER",
            bg=COLOR_DARK_PANEL,
            fg=COLOR_ACCENT_PALE,
            font=FONT_BRAND,
        ).pack(anchor="w")
        tk.Label(
            header,
            text="답안지를 읽고, 결과를 정리합니다",
            bg=COLOR_DARK_PANEL,
            fg=COLOR_SURFACE,
            font=FONT_DISPLAY,
        ).pack(anchor="w", pady=(SPACE_4, SPACE_2))
        tk.Label(
            header,
            text="흐린 스캔도 자동 보정하고, 애매한 표시는 확인 대상으로 남깁니다.",
            bg=COLOR_DARK_PANEL,
            fg=COLOR_DARK_MUTED,
            font=FONT_BODY,
        ).pack(anchor="w")

        shell = tk.Frame(root, bg=COLOR_WINDOW, padx=SPACE_16, pady=SPACE_8)
        shell.pack(fill="both", expand=True)
        shell.grid_columnconfigure(0, weight=0, minsize=230)
        shell.grid_columnconfigure(1, weight=1)
        shell.grid_rowconfigure(0, weight=1)

        rail = tk.Frame(shell, bg=COLOR_DARK_PANEL, padx=SPACE_16, pady=SPACE_16)
        rail.grid(row=0, column=0, sticky="nsew", padx=(0, SPACE_16))
        tk.Label(
            rail,
            text="작업 순서",
            bg=COLOR_DARK_PANEL,
            fg=COLOR_SURFACE,
            font=FONT_RAIL,
        ).pack(anchor="w")
        tk.Label(
            rail,
            text="세 단계만 준비하면\n채점 결과가 만들어집니다.",
            bg=COLOR_DARK_PANEL,
            fg=COLOR_DARK_MUTED,
            justify="left",
            font=FONT_SMALL,
            ).pack(anchor="w", pady=(SPACE_8, SPACE_24))
        for number, text in (("01", "답안지 선택"), ("02", "정답표 연결"), ("03", "결과 저장")):
            row = tk.Frame(rail, bg=COLOR_DARK_PANEL)
            row.pack(fill="x", pady=SPACE_8)
            tk.Label(
                row,
                text=number,
                bg=COLOR_ACCENT_DARK,
                fg=COLOR_SURFACE,
                width=3,
                font=FONT_MONO_BOLD,
            ).pack(side="left")
            tk.Label(
                row,
                text=text,
                bg=COLOR_DARK_PANEL,
                fg=COLOR_SURFACE,
                font=FONT_SMALL,
            ).pack(side="left", padx=SPACE_8)
        tk.Label(
            rail,
            text="제작: 조승현",
            bg=COLOR_DARK_PANEL,
            fg=COLOR_DARK_MUTED,
            font=FONT_SMALL,
        ).pack(side="bottom", anchor="w")

        workspace = tk.Frame(shell, bg=COLOR_WINDOW)
        workspace.grid(row=0, column=1, sticky="nsew")
        workspace.grid_columnconfigure(0, weight=1)
        workspace.grid_rowconfigure(5, weight=1)
        tk.Label(
            workspace,
            text="채점 준비",
            bg=COLOR_WINDOW,
            fg=COLOR_INK,
            font=FONT_SECTION,
        ).grid(row=0, column=0, sticky="w")
        tk.Label(
            workspace,
            text="파일을 고르고 인식 민감도를 조절하세요.",
            bg=COLOR_WINDOW,
            fg=COLOR_INK_MUTED,
            font=FONT_BODY,
        ).grid(row=1, column=0, sticky="w", pady=(SPACE_4, SPACE_12))

        self.scan_card = self._card(workspace, "01", "답안지", "PDF, JPG, PNG 파일 또는 폴더")
        self.scan_card.grid(row=2, column=0, sticky="ew", pady=(0, SPACE_8))
        self.scan_label = tk.Label(
            self.scan_card, text="선택된 답안지 없음", bg=COLOR_SURFACE, fg=COLOR_INK_MUTED,
            font=FONT_SMALL, anchor="w", justify="left", wraplength=PATH_WRAP_LENGTH
        )
        self.scan_label.pack(fill="x", padx=SPACE_16, pady=(0, SPACE_10))
        scan_actions = tk.Frame(self.scan_card, bg=COLOR_SURFACE)
        scan_actions.pack(fill="x", padx=SPACE_16, pady=(0, SPACE_16))
        self.scan_folder_button = ttk.Button(scan_actions, text="폴더 선택", style="Secondary.TButton", command=self.pick_folder)
        self.scan_folder_button.pack(side="left")
        self.scan_file_button = ttk.Button(scan_actions, text="파일 선택", style="Secondary.TButton", command=self.pick_files)
        self.scan_file_button.pack(side="left", padx=(SPACE_8, 0))

        self.key_card = self._card(workspace, "02", "정답표", "문항번호와 정답이 담긴 XLSX 또는 CSV")
        self.key_card.grid(row=3, column=0, sticky="ew", pady=(0, SPACE_8))
        self.key_label = tk.Label(
            self.key_card, text="선택된 정답표 없음", bg=COLOR_SURFACE, fg=COLOR_INK_MUTED,
            font=FONT_SMALL, anchor="w", justify="left", wraplength=PATH_WRAP_LENGTH
        )
        self.key_label.pack(fill="x", padx=SPACE_16, pady=(0, SPACE_10))
        self.key_button = ttk.Button(self.key_card, text="정답표 선택", style="Secondary.TButton", command=self.pick_key)
        self.key_button.pack(anchor="w", padx=SPACE_16, pady=(0, SPACE_16))

        settings = self._card(workspace, "설정", "인식 민감도", "흐린 마킹은 높이고, 인쇄 잡음은 낮춰 보세요.")
        settings.grid(row=4, column=0, sticky="new", pady=(0, SPACE_8))
        scale_row = tk.Frame(settings, bg=COLOR_SURFACE)
        scale_row.pack(fill="x", padx=SPACE_16, pady=(0, SPACE_16))
        self.sensitivity_scale = ttk.Scale(
            scale_row, from_=1, to=10, variable=self.sensitivity_var,
            command=self._on_sensitivity, style="Modern.Horizontal.TScale"
        )
        self.sensitivity_scale.pack(side="left", fill="x", expand=True, padx=(0, SPACE_12))
        self.sensitivity_value = tk.Label(
            scale_row, text="6 / 10", bg=COLOR_ACCENT_PALE, fg=COLOR_ACCENT_DARK,
            font=FONT_MONO_BOLD, padx=SPACE_8, pady=SPACE_4
        )
        self.sensitivity_value.pack(side="right")

        activity = tk.Frame(workspace, bg=COLOR_DARK_PANEL, padx=SPACE_16, pady=SPACE_10)
        activity.grid(row=5, column=0, sticky="nsew", pady=(0, SPACE_8))
        activity.grid_columnconfigure(0, weight=1)
        activity.grid_columnconfigure(1, weight=0)
        activity.grid_rowconfigure(1, weight=1)
        status_row = tk.Frame(activity, bg=COLOR_DARK_PANEL)
        status_row.grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, SPACE_8))
        self.status_dot = tk.Canvas(
            status_row, width=10, height=10, bg=COLOR_DARK_PANEL, highlightthickness=0
        )
        self.status_dot_item = self.status_dot.create_oval(
            2, 2, 8, 8, fill=COLOR_ACCENT, outline=COLOR_ACCENT
        )
        self.status_dot.pack(side="left", padx=(0, SPACE_8))
        self.status_label = tk.Label(
            status_row, text="준비됨", bg=COLOR_DARK_PANEL, fg=COLOR_ACCENT_PALE,
            font=FONT_BODY_BOLD, anchor="w"
        )
        self.status_label.pack(side="left")
        self.log_text = tk.Text(
            activity, height=3, bg=COLOR_DARK_PANEL_DEEP, fg=COLOR_SURFACE, insertbackground=COLOR_SURFACE,
            relief="flat", borderwidth=0, font=FONT_MONO, padx=SPACE_10, pady=SPACE_8
        )
        self.log_text.grid(row=1, column=0, sticky="nsew")
        log_scroll = ttk.Scrollbar(activity, orient="vertical", command=self.log_text.yview)
        log_scroll.grid(row=1, column=1, sticky="ns", padx=(SPACE_8, 0))
        self.log_text.configure(yscrollcommand=log_scroll.set)
        self.start_button = ttk.Button(
            workspace, text="채점 시작 · 결과 폴더 선택", style="Action.TButton", command=self.start
        )
        self.start_button.grid(row=6, column=0, sticky="ew")

        self.root.after(200, self._poll_log_queue)

    def _card(self, parent, number, title, subtitle):
        card = tk.Frame(
            parent, bg=COLOR_SURFACE, highlightbackground=COLOR_BORDER,
            highlightthickness=1, padx=SPACE_16, pady=SPACE_6
        )
        heading = tk.Frame(card, bg=COLOR_SURFACE)
        heading.pack(fill="x", pady=(0, SPACE_8))
        tk.Label(
            heading, text=number, bg=COLOR_ACCENT_PALE, fg=COLOR_ACCENT_DARK,
            font=FONT_MONO_BOLD, padx=SPACE_8, pady=SPACE_4
        ).pack(side="left")
        title_box = tk.Frame(heading, bg=COLOR_SURFACE)
        title_box.pack(side="left", padx=SPACE_10)
        tk.Label(title_box, text=title, bg=COLOR_SURFACE, fg=COLOR_INK, font=FONT_TITLE).pack(anchor="w")
        tk.Label(title_box, text=subtitle, bg=COLOR_SURFACE, fg=COLOR_INK_MUTED, font=FONT_SMALL).pack(anchor="w")
        return card

    def _on_sensitivity(self, value):
        self.sensitivity_value.config(text=f"{round(float(value))} / 10")

    def _set_status(self, text, color):
        self.status_label.config(text=text)
        self.status_dot.itemconfig(self.status_dot_item, fill=color, outline=color)

    def _mark_selected(self, card, label):
        card.config(highlightbackground=COLOR_ACCENT, highlightthickness=2)
        label.config(fg=COLOR_ACCENT_DARK)

    def _set_input_state(self, state):
        for control in (
            self.scan_folder_button,
            self.scan_file_button,
            self.key_button,
            self.sensitivity_scale,
        ):
            control.config(state=state)

    def pick_folder(self):
        folder = filedialog.askdirectory(title="답안지 폴더 선택")
        if folder:
            self.scan_paths = [
                os.path.join(folder, f)
                for f in sorted(os.listdir(folder))
                if f.lower().endswith((".png", ".jpg", ".jpeg", ".pdf"))
            ]
            self.scan_label.config(text=f"{len(self.scan_paths)}개 선택 · {folder}")
            self._mark_selected(self.scan_card, self.scan_label)

    def pick_files(self):
        files = filedialog.askopenfilenames(
            title="답안지 파일 선택",
            filetypes=[("답안지 파일", "*.png *.jpg *.jpeg *.pdf")],
        )
        if files:
            self.scan_paths = list(files)
            self.scan_label.config(text=f"{len(self.scan_paths)}개 파일 선택 · {os.path.dirname(self.scan_paths[0])}")
            self._mark_selected(self.scan_card, self.scan_label)

    def pick_key(self):
        path = filedialog.askopenfilename(
            title="정답표 파일 선택", filetypes=[("정답표", "*.xlsx *.csv")]
        )
        if path:
            self.key_path = path
            self.key_label.config(text=f"선택됨 · {path}")
            self._mark_selected(self.key_card, self.key_label)

    def start(self):
        if not self.scan_paths:
            messagebox.showerror("오류", "답안지를 먼저 선택하세요")
            return
        if not self.key_path:
            messagebox.showerror("오류", "정답표를 먼저 선택하세요")
            return
        output_dir = filedialog.askdirectory(title="결과를 저장할 폴더 선택")
        if not output_dir:
            return
        sensitivity = round(self.sensitivity_var.get())
        self._set_input_state("disabled")
        self.start_button.config(state="disabled")
        self._set_status(f"채점 중 · 민감도 {sensitivity}/10", COLOR_WARNING)
        self.log_text.insert("end", f"채점 시작 · 민감도 {sensitivity}/10\n")
        threading.Thread(
            target=_run_pipeline_worker,
            args=(self.scan_paths, self.key_path, output_dir, sensitivity, self.log_queue, self.done_queue),
            daemon=True,
        ).start()

    def _poll_log_queue(self):
        while not self.log_queue.empty():
            self.log_text.insert("end", self.log_queue.get() + "\n")
            self.log_text.see("end")
        while not self.done_queue.empty():
            status, msg = self.done_queue.get()
            if status == "ok":
                self._set_input_state("normal")
                self.start_button.config(state="normal")
                self._set_status("완료 · 결과 파일이 준비되었습니다", COLOR_ACCENT)
                messagebox.showinfo("완료", f"채점 결과: {msg}")
            else:
                self._set_input_state("normal")
                self.start_button.config(state="normal")
                self._set_status("오류 · 입력을 확인하세요", COLOR_ERROR)
                messagebox.showerror("오류", msg)
        self.root.after(200, self._poll_log_queue)

    def open_history(self):
        HistoryWindow(self.root, self.store)


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("OMR 답안지 채점")
        self.root.geometry("1120x780")
        self.root.minsize(960, 680)
        self.root.configure(bg=COLOR_WINDOW)
        self.scan_paths = []
        self.key_path = None
        self.log_queue = queue.Queue()
        self.done_queue = queue.Queue()
        self.store = RunStore(default_history_dir())
        self.interrupted_count = self.store.mark_interrupted_runs()
        self.sensitivity_var = tk.DoubleVar(value=DEFAULT_SENSITIVITY)
        self._configure_style()
        self._build_shell()
        self.show_page("prepare")
        self.root.after(200, self._poll_log_queue)

    def _configure_style(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "Action.TButton",
            background=COLOR_ACCENT_DARK,
            foreground=COLOR_SURFACE,
            borderwidth=0,
            padding=(SPACE_14, SPACE_8),
            font=FONT_BODY_BOLD,
        )
        style.map("Action.TButton", background=[("active", COLOR_INK), ("pressed", COLOR_INK)])
        style.configure(
            "Secondary.TButton",
            background=COLOR_SURFACE_MUTED,
            foreground=COLOR_INK,
            borderwidth=0,
            padding=(SPACE_12, SPACE_8),
            font=FONT_SMALL_BOLD,
        )
        style.map("Secondary.TButton", background=[("active", COLOR_BORDER)])
        style.configure("Modern.Horizontal.TScale", troughcolor=COLOR_SURFACE_MUTED, background=COLOR_ACCENT)

    def _build_shell(self):
        shell = tk.Frame(self.root, bg=COLOR_WINDOW)
        shell.pack(fill="both", expand=True)
        rail = tk.Frame(shell, bg=COLOR_DARK_PANEL, width=208, padx=SPACE_14, pady=SPACE_16)
        rail.pack(side="left", fill="y")
        rail.pack_propagate(False)
        tk.Label(rail, text="OMR GRADER", bg=COLOR_DARK_PANEL, fg=COLOR_ACCENT_PALE, font=FONT_BRAND).pack(anchor="w")
        tk.Label(rail, text="채점 워크스페이스", bg=COLOR_DARK_PANEL, fg=COLOR_SURFACE, font=FONT_SECTION).pack(anchor="w", pady=(SPACE_8, SPACE_2))
        tk.Label(rail, text="원본·결과·검토 이력을\n한 곳에서 관리합니다.", bg=COLOR_DARK_PANEL, fg=COLOR_DARK_MUTED, justify="left", font=FONT_SMALL).pack(anchor="w", pady=(0, SPACE_24))
        self.nav_buttons = {}
        for key, label in (
            ("prepare", "채점 준비"),
            ("history", "최근 채점"),
            ("students", "학생 결과"),
            ("analysis", "문항 분석"),
            ("settings", "설정·진단"),
        ):
            button = tk.Button(
                rail,
                text=label,
                command=lambda page=key: self.show_page(page),
                anchor="w",
                relief="flat",
                borderwidth=0,
                padx=SPACE_12,
                pady=SPACE_10,
                font=FONT_BODY_BOLD,
                bg=COLOR_DARK_PANEL,
                fg=COLOR_SURFACE,
                activebackground=COLOR_ACCENT_DARK,
                activeforeground=COLOR_SURFACE,
            )
            button.pack(fill="x", pady=(0, SPACE_4))
            self.nav_buttons[key] = button
        tk.Label(rail, text="오프라인 모드\n서버 연결 없음", bg=COLOR_DARK_PANEL, fg=COLOR_DARK_MUTED, justify="left", font=FONT_SMALL).pack(side="bottom", anchor="w")
        self.content = tk.Frame(shell, bg=COLOR_WINDOW, padx=SPACE_24, pady=SPACE_16)
        self.content.pack(side="left", fill="both", expand=True)

    def show_page(self, page):
        for child in self.content.winfo_children():
            child.destroy()
        for key, button in self.nav_buttons.items():
            button.config(bg=COLOR_ACCENT_DARK if key == page else COLOR_DARK_PANEL)
        match page:
            case "prepare":
                self._build_prepare_page()
            case "history":
                self._build_history_page()
            case "students":
                self._build_students_page()
            case "analysis":
                self._build_analysis_page()
            case "settings":
                self._build_settings_page()

    def _page_title(self, title, subtitle):
        tk.Label(self.content, text=title, bg=COLOR_WINDOW, fg=COLOR_INK, font=FONT_DISPLAY).pack(anchor="w")
        tk.Label(self.content, text=subtitle, bg=COLOR_WINDOW, fg=COLOR_INK_MUTED, font=FONT_BODY).pack(anchor="w", pady=(SPACE_4, SPACE_16))

    def _panel(self, parent):
        return tk.Frame(parent, bg=COLOR_SURFACE, highlightbackground=COLOR_BORDER, highlightthickness=1, padx=SPACE_16, pady=SPACE_14)

    def _build_prepare_page(self):
        self._page_title("채점 준비", "시험 정보와 파일을 확인한 뒤 채점을 시작합니다.")
        summary = tk.Frame(self.content, bg=COLOR_WINDOW)
        summary.pack(fill="x", pady=(0, SPACE_12))
        for title, value, key in (
            ("선택 파일", f"{len(self.scan_paths)}개", "files"),
            ("정답표", "연결됨" if self.key_path else "미선택", "key"),
            ("검토 필요", "자동 분류", "review"),
        ):
            card = self._panel(summary)
            card.pack(side="left", fill="x", expand=True, padx=(0, SPACE_8) if key != "review" else (0, 0))
            tk.Label(card, text=title, bg=COLOR_SURFACE, fg=COLOR_INK_MUTED, font=FONT_SMALL).pack(anchor="w")
            tk.Label(card, text=value, bg=COLOR_SURFACE, fg=COLOR_INK, font=FONT_SECTION).pack(anchor="w", pady=(SPACE_4, 0))

        file_panel = self._panel(self.content)
        file_panel.pack(fill="x", pady=(0, SPACE_10))
        tk.Label(file_panel, text="1. 답안지 파일", bg=COLOR_SURFACE, fg=COLOR_INK, font=FONT_TITLE).pack(anchor="w")
        self.scan_label = tk.Label(file_panel, text=self._scan_text(), bg=COLOR_SURFACE, fg=COLOR_INK_MUTED, font=FONT_SMALL, anchor="w", justify="left", wraplength=760)
        self.scan_label.pack(fill="x", pady=(SPACE_6, SPACE_10))
        file_actions = tk.Frame(file_panel, bg=COLOR_SURFACE)
        file_actions.pack(fill="x")
        self.scan_folder_button = ttk.Button(file_actions, text="폴더 선택", style="Secondary.TButton", command=self.pick_folder)
        self.scan_folder_button.pack(side="left")
        self.scan_file_button = ttk.Button(file_actions, text="파일 선택", style="Secondary.TButton", command=self.pick_files)
        self.scan_file_button.pack(side="left", padx=SPACE_8)

        key_panel = self._panel(self.content)
        key_panel.pack(fill="x", pady=(0, SPACE_10))
        tk.Label(key_panel, text="2. 정답표와 인식 설정", bg=COLOR_SURFACE, fg=COLOR_INK, font=FONT_TITLE).pack(anchor="w")
        self.key_label = tk.Label(key_panel, text=self._key_text(), bg=COLOR_SURFACE, fg=COLOR_INK_MUTED, font=FONT_SMALL, anchor="w", justify="left", wraplength=760)
        self.key_label.pack(fill="x", pady=(SPACE_6, SPACE_8))
        key_actions = tk.Frame(key_panel, bg=COLOR_SURFACE)
        key_actions.pack(fill="x")
        self.key_button = ttk.Button(key_actions, text="정답표 선택", style="Secondary.TButton", command=self.pick_key)
        self.key_button.pack(side="left")
        ttk.Label(key_actions, text="민감도").pack(side="left", padx=(SPACE_24, SPACE_8))
        self.sensitivity_scale = ttk.Scale(key_actions, from_=1, to=10, variable=self.sensitivity_var, command=self._on_sensitivity, style="Modern.Horizontal.TScale")
        self.sensitivity_scale.pack(side="left", fill="x", expand=True)
        self.sensitivity_value = tk.Label(key_actions, text=f"{round(self.sensitivity_var.get())} / 10", bg=COLOR_ACCENT_PALE, fg=COLOR_ACCENT_DARK, font=FONT_MONO_BOLD, padx=SPACE_8, pady=SPACE_4)
        self.sensitivity_value.pack(side="left", padx=(SPACE_8, 0))

        activity = tk.Frame(self.content, bg=COLOR_DARK_PANEL, padx=SPACE_14, pady=SPACE_10)
        activity.pack(fill="both", expand=True, pady=(0, SPACE_10))
        self.status_label = tk.Label(activity, text="준비됨", bg=COLOR_DARK_PANEL, fg=COLOR_ACCENT_PALE, font=FONT_BODY_BOLD)
        self.status_label.pack(anchor="w")
        self.log_text = tk.Text(activity, height=5, bg=COLOR_DARK_PANEL_DEEP, fg=COLOR_SURFACE, relief="flat", font=FONT_MONO, padx=SPACE_8, pady=SPACE_6)
        self.log_text.pack(fill="both", expand=True, pady=(SPACE_8, 0))
        self.start_button = ttk.Button(self.content, text="채점 시작", style="Action.TButton", command=self.start)
        self.start_button.pack(fill="x")

    def _build_history_page(self):
        self._page_title("최근 채점", "원본과 채점 기준을 다시 열어 검토·재분석·재채점할 수 있습니다.")
        actions = tk.Frame(self.content, bg=COLOR_WINDOW)
        actions.pack(fill="x", pady=(0, SPACE_10))
        self.history_query = tk.StringVar()
        search = ttk.Entry(actions, textvariable=self.history_query, width=34)
        search.pack(side="left")
        search.bind("<Return>", lambda _event: self._refresh_history())
        ttk.Button(actions, text="검색", style="Secondary.TButton", command=self._refresh_history).pack(side="left", padx=SPACE_8)
        ttk.Button(actions, text="상세 열기", style="Action.TButton", command=self._open_selected_history).pack(side="right")
        table = self._panel(self.content)
        table.pack(fill="both", expand=True)
        columns = ("created", "status", "count", "key", "review")
        self.history_tree = ttk.Treeview(table, columns=columns, show="headings", selectmode="browse")
        for column, heading, width in (
            ("created", "생성 시각", 180),
            ("status", "상태", 100),
            ("count", "스캔 수", 80),
            ("key", "정답표", 280),
            ("review", "마지막 처리", 220),
        ):
            self.history_tree.heading(column, text=heading)
            self.history_tree.column(column, width=width, anchor="w")
        self.history_tree.pack(fill="both", expand=True)
        self.history_tree.bind("<Double-1>", lambda _event: self._open_selected_history())
        self._refresh_history()

    def _refresh_history(self):
        if not hasattr(self, "history_tree"):
            return
        for item in self.history_tree.get_children():
            self.history_tree.delete(item)
        query = self.history_query.get() if hasattr(self, "history_query") else ""
        for run in self.store.list_runs(query):
            self.history_tree.insert("", "end", iid=run["id"], values=(run["created_at"].replace("T", " "), run["status"], run["scan_count"], os.path.basename(run["answer_key_path"]), run["last_label"] or "-"))

    def _open_selected_history(self):
        selected = self.history_tree.selection() if hasattr(self, "history_tree") else ()
        if not selected:
            messagebox.showinfo("최근 채점", "채점 실행을 선택하세요", parent=self.root)
            return
        RunDetailWindow(self.root, self.store, selected[0])

    def _build_students_page(self):
        self._page_title("학생 결과", "학생별 원본·인식 결과·수정 이력을 확인합니다.")
        runs = self.store.list_runs(page_size=1)
        if not runs:
            tk.Label(self.content, text="저장된 채점 이력이 없습니다.", bg=COLOR_WINDOW, fg=COLOR_INK_MUTED, font=FONT_BODY).pack(anchor="w")
            return
        run = runs[0]
        tk.Label(self.content, text=f"최근 실행 · {run['created_at'].replace('T', ' ')}", bg=COLOR_WINDOW, fg=COLOR_INK_MUTED, font=FONT_SMALL).pack(anchor="w", pady=(0, SPACE_8))
        panel = self._panel(self.content)
        panel.pack(fill="both", expand=True)
        tree = ttk.Treeview(panel, columns=("student", "score", "review"), show="headings")
        for column, heading, width in (("student", "학번", 220), ("score", "점수", 140), ("review", "검토 상태", 180)):
            tree.heading(column, text=heading)
            tree.column(column, width=width, anchor="w")
        tree.pack(fill="both", expand=True)
        for record in self.store.get_records(run["id"]):
            tree.insert("", "end", iid=str(record["id"]), values=(record["student_id"], record["score"], record["review_status"]))
        tree.bind("<Double-1>", lambda _event: RunDetailWindow(self.root, self.store, run["id"]))

    def _build_analysis_page(self):
        self._page_title("문항 분석", "문항별 정답률과 선택지 분포를 수업·문항 개선에 활용합니다.")
        runs = self.store.list_runs(page_size=1)
        if not runs:
            tk.Label(self.content, text="분석할 채점 이력이 없습니다.", bg=COLOR_WINDOW, fg=COLOR_INK_MUTED, font=FONT_BODY).pack(anchor="w")
            return
        analysis = self.store.analyze(runs[0]["id"])
        summary = tk.Frame(self.content, bg=COLOR_WINDOW)
        summary.pack(fill="x", pady=(0, SPACE_12))
        for title, value in (("응시 학생", f"{analysis['student_count']}명"), ("평균 점수", f"{analysis['average_score']:.1f}"), ("최고 점수", str(analysis["max_score"]))):
            card = self._panel(summary)
            card.pack(side="left", fill="x", expand=True, padx=(0, SPACE_8))
            tk.Label(card, text=title, bg=COLOR_SURFACE, fg=COLOR_INK_MUTED, font=FONT_SMALL).pack(anchor="w")
            tk.Label(card, text=value, bg=COLOR_SURFACE, fg=COLOR_INK, font=FONT_SECTION).pack(anchor="w", pady=(SPACE_4, 0))
        panel = self._panel(self.content)
        panel.pack(fill="both", expand=True)
        tree = ttk.Treeview(panel, columns=("question", "rate", "options"), show="headings")
        for column, heading, width in (("question", "문항", 120), ("rate", "정답률", 160), ("options", "선택지 분포 (1~5)", 420)):
            tree.heading(column, text=heading)
            tree.column(column, width=width, anchor="w")
        tree.pack(fill="both", expand=True)
        for question, stats in analysis["question_stats"].items():
            options = " · ".join(f"{option}:{count}" for option, count in stats["option_counts"].items())
            tree.insert("", "end", values=(f"Q{question}", f"{stats['correct_rate'] * 100:.1f}%", options))

    def _build_settings_page(self):
        self._page_title("설정·진단", "민감도, 보존 위치, 중단 작업과 진단 로그를 확인합니다.")
        panel = self._panel(self.content)
        panel.pack(fill="x")
        tk.Label(panel, text="로컬 저장 위치", bg=COLOR_SURFACE, fg=COLOR_INK, font=FONT_TITLE).pack(anchor="w")
        tk.Label(panel, text=default_history_dir(), bg=COLOR_SURFACE, fg=COLOR_INK_MUTED, font=FONT_SMALL, wraplength=760, justify="left").pack(anchor="w", pady=(SPACE_6, SPACE_12))
        interrupted = f"중단 작업 {self.interrupted_count}건을 감지했습니다." if self.interrupted_count else "중단된 작업이 없습니다."
        tk.Label(panel, text=interrupted, bg=COLOR_SURFACE, fg=COLOR_WARNING if self.interrupted_count else COLOR_ACCENT_DARK, font=FONT_BODY_BOLD).pack(anchor="w")
        ttk.Button(panel, text="채점 이력과 진단 로그 열기", style="Action.TButton", command=lambda: HistoryWindow(self.root, self.store)).pack(anchor="w", pady=(SPACE_16, 0))

    def _scan_text(self):
        if not self.scan_paths:
            return "선택된 답안지 없음"
        return f"{len(self.scan_paths)}개 파일 · {os.path.dirname(self.scan_paths[0])}"

    def _key_text(self):
        return f"선택됨 · {self.key_path}" if self.key_path else "선택된 정답표 없음"

    def _on_sensitivity(self, value):
        if hasattr(self, "sensitivity_value"):
            self.sensitivity_value.config(text=f"{round(float(value))} / 10")

    def _set_status(self, text, color):
        if hasattr(self, "status_label"):
            self.status_label.config(text=text, fg=color)

    def _set_input_state(self, state):
        for name in ("scan_folder_button", "scan_file_button", "key_button", "sensitivity_scale"):
            control = getattr(self, name, None)
            if control is not None:
                control.config(state=state)

    def pick_folder(self):
        folder = filedialog.askdirectory(title="답안지 폴더 선택")
        if folder:
            self.scan_paths = [os.path.join(folder, name) for name in sorted(os.listdir(folder)) if name.lower().endswith((".png", ".jpg", ".jpeg", ".pdf"))]
            if hasattr(self, "scan_label"):
                self.scan_label.config(text=self._scan_text(), fg=COLOR_ACCENT_DARK)

    def pick_files(self):
        files = filedialog.askopenfilenames(title="답안지 파일 선택", filetypes=[("답안지 파일", "*.png *.jpg *.jpeg *.pdf")])
        if files:
            self.scan_paths = list(files)
            if hasattr(self, "scan_label"):
                self.scan_label.config(text=self._scan_text(), fg=COLOR_ACCENT_DARK)

    def pick_key(self):
        path = filedialog.askopenfilename(title="정답표 파일 선택", filetypes=[("정답표", "*.xlsx *.csv")])
        if path:
            self.key_path = path
            if hasattr(self, "key_label"):
                self.key_label.config(text=self._key_text(), fg=COLOR_ACCENT_DARK)

    def start(self):
        if not self.scan_paths:
            messagebox.showerror("오류", "답안지를 먼저 선택하세요", parent=self.root)
            return
        if not self.key_path:
            messagebox.showerror("오류", "정답표를 먼저 선택하세요", parent=self.root)
            return
        output_dir = filedialog.askdirectory(title="결과를 저장할 폴더 선택")
        if not output_dir:
            return
        sensitivity = round(self.sensitivity_var.get())
        self._set_input_state("disabled")
        if hasattr(self, "start_button"):
            self.start_button.config(state="disabled")
        self._set_status(f"채점 중 · 민감도 {sensitivity}/10", COLOR_WARNING)
        if hasattr(self, "log_text"):
            self.log_text.insert("end", f"채점 시작 · 민감도 {sensitivity}/10\n")
        threading.Thread(target=_run_pipeline_worker, args=(self.scan_paths, self.key_path, output_dir, sensitivity, self.log_queue, self.done_queue), daemon=True).start()

    def _poll_log_queue(self):
        while not self.log_queue.empty():
            message = self.log_queue.get()
            if hasattr(self, "log_text"):
                self.log_text.insert("end", message + "\n")
                self.log_text.see("end")
        while not self.done_queue.empty():
            status, message = self.done_queue.get()
            self._set_input_state("normal")
            if hasattr(self, "start_button"):
                self.start_button.config(state="normal")
            if status == "ok":
                self._set_status("완료 · 결과와 이력이 저장되었습니다", COLOR_ACCENT)
                messagebox.showinfo("완료", f"채점 결과: {message}", parent=self.root)
            else:
                self._set_status("오류 · 진단 로그를 확인하세요", COLOR_ERROR)
                messagebox.showerror("오류", message, parent=self.root)
        self.root.after(200, self._poll_log_queue)


class HistoryWindow:
    def __init__(self, parent, store):
        self.store = store
        self.window = tk.Toplevel(parent)
        self.window.title("채점 이력")
        self.window.geometry("900x560")
        self.window.configure(bg=COLOR_WINDOW)
        self.query_var = tk.StringVar()

        header = tk.Frame(self.window, bg=COLOR_WINDOW, padx=SPACE_16, pady=SPACE_12)
        header.pack(fill="x")
        tk.Label(header, text="채점 이력", bg=COLOR_WINDOW, fg=COLOR_INK, font=FONT_SECTION).pack(side="left")
        search = ttk.Entry(header, textvariable=self.query_var, width=36)
        search.pack(side="right", padx=(SPACE_8, 0))
        search.bind("<Return>", lambda _event: self.refresh())
        ttk.Button(header, text="검색", style="Secondary.TButton", command=self.refresh).pack(side="right")

        body = tk.Frame(self.window, bg=COLOR_WINDOW, padx=SPACE_16)
        body.pack(fill="both", expand=True)
        body.grid_columnconfigure(0, weight=1)
        body.grid_rowconfigure(0, weight=1)
        columns = ("created", "status", "count", "key", "last")
        self.tree = ttk.Treeview(body, columns=columns, show="headings", selectmode="browse")
        headings = {
            "created": "생성 시각",
            "status": "상태",
            "count": "스캔 수",
            "key": "정답키",
            "last": "마지막 처리",
        }
        widths = {"created": 170, "status": 90, "count": 80, "key": 330, "last": 180}
        for column in columns:
            self.tree.heading(column, text=headings[column])
            self.tree.column(column, width=widths[column], anchor="w")
        self.tree.grid(row=0, column=0, sticky="nsew")
        scrollbar = ttk.Scrollbar(body, orient="vertical", command=self.tree.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.bind("<Double-1>", lambda _event: self.open_detail())

        actions = tk.Frame(self.window, bg=COLOR_WINDOW, padx=SPACE_16, pady=SPACE_12)
        actions.pack(fill="x")
        ttk.Button(actions, text="상세 열기", style="Action.TButton", command=self.open_detail).pack(side="left")
        ttk.Button(actions, text="진단 로그", style="Secondary.TButton", command=self.open_logs).pack(side="left", padx=SPACE_8)
        ttk.Button(actions, text="백업", style="Secondary.TButton", command=self.backup).pack(side="left")
        ttk.Button(actions, text="삭제", style="Secondary.TButton", command=self.delete).pack(side="left", padx=SPACE_8)
        ttk.Button(actions, text="새로 고침", style="Secondary.TButton", command=self.refresh).pack(side="right")
        self.refresh()

    def refresh(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for run in self.store.list_runs(self.query_var.get()):
            self.tree.insert(
                "",
                "end",
                iid=run["id"],
                values=(
                    run["created_at"].replace("T", " "),
                    run["status"],
                    run["scan_count"],
                    os.path.basename(run["answer_key_path"]),
                    run["last_label"] or "-",
                ),
            )

    def selected_run_id(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showinfo("채점 이력", "먼저 채점 실행을 선택하세요", parent=self.window)
            return None
        return selected[0]

    def open_detail(self):
        run_id = self.selected_run_id()
        if run_id:
            RunDetailWindow(self.window, self.store, run_id)

    def open_logs(self):
        run_id = self.selected_run_id()
        if not run_id:
            return
        logs = self.store.get_logs(run_id)
        text = "\n".join(
            f"{log['created_at']} [{log['level']}] {log['event']} · {log['message']}"
            for log in logs
        ) or "로그가 없습니다."
        _show_text_window(self.window, "진단 로그", text)

    def backup(self):
        run_id = self.selected_run_id()
        if not run_id:
            return
        path = filedialog.asksaveasfilename(
            parent=self.window,
            title="채점 실행 백업",
            defaultextension=".zip",
            filetypes=[("ZIP 백업", "*.zip")],
        )
        if path:
            self.store.backup_run(run_id, path)
            messagebox.showinfo("백업", "원본·결과·진단 로그를 백업했습니다", parent=self.window)

    def delete(self):
        run_id = self.selected_run_id()
        if not run_id or not messagebox.askyesno(
            "채점 이력 삭제",
            "이력과 연결된 원본을 복구 가능한 보관 폴더로 이동하시겠습니까?",
            parent=self.window,
        ):
            return
        self.store.delete_run(run_id)
        self.refresh()


class RunDetailWindow:
    def __init__(self, parent, store, run_id):
        self.store = store
        self.run_id = run_id
        self.window = tk.Toplevel(parent)
        self.window.title("결과 상세")
        self.window.geometry("1100x700")
        self.window.configure(bg=COLOR_WINDOW)
        self.current_record_id = None

        run = store.get_run(run_id)
        top = tk.Frame(self.window, bg=COLOR_WINDOW, padx=SPACE_16, pady=SPACE_12)
        top.pack(fill="x")
        tk.Label(top, text="결과 상세", bg=COLOR_WINDOW, fg=COLOR_INK, font=FONT_SECTION).pack(side="left")
        tk.Label(
            top,
            text=f"{run['status']} · {run['created_at'].replace('T', ' ')} · {run['answer_key_path']}",
            bg=COLOR_WINDOW,
            fg=COLOR_INK_MUTED,
            font=FONT_SMALL,
        ).pack(side="left", padx=SPACE_16)
        ttk.Button(top, text="재채점", style="Action.TButton", command=self.regrade).pack(side="right")
        ttk.Button(top, text="재분석", style="Secondary.TButton", command=self.reanalyze).pack(side="right", padx=SPACE_8)
        ttk.Button(top, text="결과 내보내기", style="Secondary.TButton", command=self.export).pack(side="right", padx=SPACE_8)
        ttk.Button(top, text="문항 분석", style="Secondary.TButton", command=self.analysis).pack(side="right", padx=SPACE_8)

        body = tk.PanedWindow(self.window, orient="horizontal", sashrelief="flat", bg=COLOR_WINDOW)
        body.pack(fill="both", expand=True, padx=SPACE_16, pady=(0, SPACE_12))
        left = tk.Frame(body, bg=COLOR_SURFACE, highlightbackground=COLOR_BORDER, highlightthickness=1)
        right = tk.Frame(body, bg=COLOR_SURFACE, highlightbackground=COLOR_BORDER, highlightthickness=1, padx=SPACE_12, pady=SPACE_12)
        body.add(left, minsize=480)
        body.add(right, minsize=360)

        columns = ("student", "score", "status", "review")
        self.tree = ttk.Treeview(left, columns=columns, show="headings", selectmode="browse")
        for column, heading, width in (
            ("student", "학번", 150),
            ("score", "점수", 90),
            ("status", "상태", 100),
            ("review", "검토", 100),
        ):
            self.tree.heading(column, text=heading)
            self.tree.column(column, width=width, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=SPACE_8, pady=SPACE_8)
        self.tree.bind("<<TreeviewSelect>>", self.show_record)

        self.detail = tk.Text(right, height=20, width=46, relief="flat", bg=COLOR_SURFACE_MUTED, fg=COLOR_INK, font=FONT_MONO)
        self.detail.pack(fill="both", expand=True)
        controls = tk.Frame(right, bg=COLOR_SURFACE)
        controls.pack(fill="x", pady=(SPACE_10, 0))
        ttk.Button(controls, text="수동 채점 적용", style="Action.TButton", command=self.manual_correct).pack(side="left")
        ttk.Button(controls, text="원본 열기", style="Secondary.TButton", command=self.open_original).pack(side="left", padx=SPACE_8)
        self.refresh()

    def refresh(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for record in self.store.get_records(self.run_id):
            self.tree.insert(
                "",
                "end",
                iid=str(record["id"]),
                values=(
                    record["student_id"],
                    record["score"],
                    record["review_status"],
                    ",".join(f"Q{q}" for q in record["flagged_questions"]) or "-",
                ),
            )
        if self.current_record_id is not None and self.tree.exists(str(self.current_record_id)):
            self.tree.selection_set(str(self.current_record_id))
            self.show_record()

    def selected_record(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showinfo("결과 상세", "먼저 학생을 선택하세요", parent=self.window)
            return None
        self.current_record_id = int(selected[0])
        return next(
            record for record in self.store.get_records(self.run_id) if record["id"] == self.current_record_id
        )

    def show_record(self, _event=None):
        record = self.selected_record() if _event is None else None
        if record is None and _event is not None:
            selected = self.tree.selection()
            if not selected:
                return
            self.current_record_id = int(selected[0])
            record = next(
                item for item in self.store.get_records(self.run_id) if item["id"] == self.current_record_id
            )
        answers = self.store._effective_answers(record)
        lines = [
            f"학번: {record['student_id']}",
            f"점수: {record['score']}",
            f"상태: {record['review_status']}",
            f"원본: {record['source_ref'] or '연결된 원본 없음'}",
            "",
            "문항별 응답",
        ]
        lines.extend(f"Q{question}: {answers.get(question, BLANK_LABEL)}" for question in sorted(answers))
        self.detail.delete("1.0", "end")
        self.detail.insert("end", "\n".join(lines))

    def manual_correct(self):
        record = self.selected_record()
        if record is None:
            return
        question = simpledialog.askinteger("수동 채점", "수정할 문항 번호", parent=self.window, minvalue=1)
        if question is None:
            return
        answer = simpledialog.askinteger("수동 채점", "적용할 답안(1~5)", parent=self.window, minvalue=1, maxvalue=5)
        if answer is None:
            return
        reason = simpledialog.askstring("수동 채점", "수정 사유", parent=self.window) or "수동 검토"
        try:
            self.store.update_manual_answer(self.run_id, record["id"], question, answer, reason)
        except (OSError, ValueError, sqlite3.Error) as error:
            messagebox.showerror("수동 채점", str(error), parent=self.window)
            return
        self.refresh()

    def open_original(self):
        record = self.selected_record()
        if record is None or not record["source_ref"]:
            return
        path = os.path.join(self.store.root_dir, record["source_ref"])
        if not os.path.exists(path):
            messagebox.showerror("원본 열기", "보존된 원본 파일을 찾을 수 없습니다", parent=self.window)
            return
        os.startfile(path)

    def regrade(self):
        path = filedialog.askopenfilename(
            parent=self.window,
            title="새 정답표 선택",
            filetypes=[("정답표", "*.xlsx *.csv")],
        )
        if not path:
            return
        changes = self.store.regrade(self.run_id, path, apply=False)
        summary = "변경되는 학생이 없습니다." if not changes else "\n".join(
            f"{change['student_id']}: {change['score_before']} → {change['score_after']}"
            for change in changes[:20]
        )
        prompt = f"{summary}\n\n적용하시겠습니까?" if changes else "변경되는 점수는 없습니다. 새 정답키 버전을 저장하시겠습니까?"
        if not messagebox.askyesno("재채점 미리보기", prompt, parent=self.window):
            return
        self.store.regrade(self.run_id, path, apply=True)
        self.refresh()
        messagebox.showinfo("재채점", f"변경 학생 수: {len(changes)}", parent=self.window)

    def reanalyze(self):
        run = self.store.get_run(self.run_id)
        sensitivity = simpledialog.askinteger(
            "재분석", "새 인식 민감도(1~10)", parent=self.window,
            initialvalue=run["sensitivity"], minvalue=1, maxvalue=10,
        )
        if sensitivity is None:
            return
        changes = self.store.reanalyze(self.run_id, sensitivity, apply=False)
        summary = "변경되는 인식 결과가 없습니다." if not changes else "\n".join(
            f"{change['student_id']} {change['label']}: {change['score_before']} → {change['score_after']}"
            for change in changes[:20]
        )
        if not messagebox.askyesno("재분석 미리보기", f"{summary}\n\n적용하시겠습니까?", parent=self.window):
            return
        self.store.reanalyze(self.run_id, sensitivity, apply=True)
        self.refresh()
        messagebox.showinfo("재분석", f"변경 학생 수: {len(changes)}", parent=self.window)

    def export(self):
        path = filedialog.asksaveasfilename(
            parent=self.window,
            title="결과 내보내기",
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
        )
        if path:
            self.store.export_run(self.run_id, path)
            messagebox.showinfo("내보내기", f"저장했습니다.\n{path}", parent=self.window)

    def analysis(self):
        analysis = self.store.analyze(self.run_id)
        difficult = sorted(analysis["question_stats"].items(), key=lambda item: item[1]["correct_rate"])[:5]
        lines = [
            f"학생 수: {analysis['student_count']}",
            f"평균 점수: {analysis['average_score']:.2f}",
            f"최고/최저: {analysis['max_score']} / {analysis['min_score']}",
            "",
            "오답률이 높은 문항",
        ]
        lines.extend(f"Q{q}: {(1 - stats['correct_rate']) * 100:.1f}%" for q, stats in difficult)
        _show_text_window(self.window, "문항 분석", "\n".join(lines))


def _show_text_window(parent, title, content):
    window = tk.Toplevel(parent)
    window.title(title)
    window.geometry("720x480")
    text = tk.Text(window, wrap="word", font=FONT_MONO, bg=COLOR_SURFACE, fg=COLOR_INK)
    text.pack(fill="both", expand=True, padx=SPACE_12, pady=SPACE_12)
    text.insert("end", content)
    text.configure(state="disabled")


def main():
    root = tk.Tk()
    App(root)
    root.update()
    try:
        import pyi_splash

        pyi_splash.close()
    except ImportError:
        pass
    root.mainloop()


if __name__ == "__main__":
    main()
