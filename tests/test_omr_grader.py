import sys
import os
import csv
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import omr_grader as og
import fitz
import numpy as np
import cv2
import openpyxl


def _render_template_page(zoom=og.ZOOM):
    doc = fitz.open("OCR100.pdf")
    page = doc[0]
    pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), alpha=False)
    img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, 3)
    return img.copy()


def test_answer_bubble_center_q1_option1():
    x, y = og.answer_bubble_center_pt(1, 1)
    assert abs(x - 244.68) < 0.1, f"x={x}"
    assert abs(y - 83.81) < 0.1, f"y={y}"


def test_answer_bubble_center_q21_option1():
    # 21번은 두 번째 블록의 1행(행 인덱스 0)이어야 함
    x, y = og.answer_bubble_center_pt(21, 1)
    assert abs(x - 361.32) < 0.1, f"x={x}"
    assert abs(y - 83.81) < 0.1, f"y={y}"


def test_answer_bubble_center_q100_option5():
    # 100번은 다섯 번째 블록의 마지막 행(행 인덱스 19)이어야 함
    x, y = og.answer_bubble_center_pt(100, 5)
    assert abs(x - (705.72 + 4 * 19.44 + 11.04 / 2)) < 0.1, f"x={x}"
    assert abs(y - (78.29 + 19 * 23.7789 + 11.04 / 2)) < 0.1, f"y={y}"


def test_id_bubble_center_col0_digit0():
    x, y = og.id_bubble_center_pt(0, 0)
    assert abs(x - (48.24 + 11.04 / 2)) < 0.1, f"x={x}"
    assert abs(y - (297.32 + 11.04 / 2)) < 0.1, f"y={y}"


def test_id_bubble_center_col7_digit9():
    x, y = og.id_bubble_center_pt(7, 9)
    assert abs(x - (48.24 + 7 * 19.8 + 11.04 / 2)) < 0.1, f"x={x}"
    assert abs(y - (297.32 + 9 * 18.7033 + 11.04 / 2)) < 0.1, f"y={y}"


def test_load_answer_key_csv():
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "key.csv")
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["문항번호", "정답"])
            for q in range(1, 51):
                w.writerow([q, (q % 5) + 1])
        key = og.load_answer_key(path)
        assert len(key) == 50
        assert key[1] == 2
        assert key[50] == 1


def test_load_answer_key_missing_question_raises():
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "key.csv")
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["문항번호", "정답"])
            w.writerow([1, 1])
            w.writerow([3, 2])  # 2번이 빠짐 -> 연속되지 않음
        try:
            og.load_answer_key(path)
            assert False, "ValueError가 발생해야 함"
        except ValueError:
            pass


def test_load_scan_images_single_page_pdf():
    with tempfile.TemporaryDirectory() as d:
        doc = fitz.open("OCR100.pdf")
        out_path = os.path.join(d, "scan.pdf")
        doc.save(out_path)
        results = og.load_scan_images([out_path])
        assert len(results) == 1
        label, img = results[0]
        assert img.ndim == 3 and img.shape[2] == 3


def test_load_scan_images_image_file():
    with tempfile.TemporaryDirectory() as d:
        img = _render_template_page()
        path = os.path.join(d, "student1.png")
        cv2.imwrite(path, cv2.cvtColor(img, cv2.COLOR_RGB2BGR))
        results = og.load_scan_images([path])
        assert len(results) == 1
        label, loaded = results[0]
        assert label == "student1.png"
        assert loaded.shape[:2] == img.shape[:2]


def test_sample_fill_ratio_filled_circle():
    img = np.full((60, 60), 255, dtype=np.uint8)
    cv2.circle(img, (30, 30), 10, 0, -1)  # 채워진 검은 원
    ratio = og.sample_fill_ratio(img, 30, 30, 8)
    assert ratio > 0.9, f"ratio={ratio}"


def test_sample_fill_ratio_blank():
    img = np.full((60, 60), 255, dtype=np.uint8)
    ratio = og.sample_fill_ratio(img, 30, 30, 8)
    assert ratio < 0.05, f"ratio={ratio}"


def test_detect_marked_index_blank():
    assert og.detect_marked_index([0.02, 0.03, 0.01, 0.02, 0.02]) is None


def test_detect_marked_index_single():
    assert og.detect_marked_index([0.02, 0.03, 0.85, 0.02, 0.02]) == 2


def test_detect_marked_index_multi():
    result = og.detect_marked_index([0.02, 0.80, 0.03, 0.78, 0.02])
    assert result == [1, 3], f"result={result}"


def _render_template_page_gray(zoom=og.ZOOM):
    img = _render_template_page(zoom)
    return cv2.cvtColor(img, cv2.COLOR_RGB2BGR)


def test_find_table_border_on_clean_render():
    img_bgr = _render_template_page_gray()
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    corners = og.find_table_border(gray)
    assert corners is not None
    expected_tl = (og.TABLE_BORDER_PT[0] * og.ZOOM, og.TABLE_BORDER_PT[1] * og.ZOOM)
    assert abs(corners[0][0] - expected_tl[0]) < 15
    assert abs(corners[0][1] - expected_tl[1]) < 15


def test_align_sheet_recovers_rotated_scan():
    img_bgr = _render_template_page_gray()
    h, w = img_bgr.shape[:2]
    # 스캐너에 종이가 살짝 삐뚤게 놓인 상황을 흉내: 2도 회전 + 여백 추가
    canvas = cv2.copyMakeBorder(img_bgr, 80, 80, 80, 80, cv2.BORDER_CONSTANT, value=(255, 255, 255))
    ch, cw = canvas.shape[:2]
    M = cv2.getRotationMatrix2D((cw / 2, ch / 2), 2.0, 1.0)
    rotated = cv2.warpAffine(canvas, M, (cw, ch), borderValue=(255, 255, 255))

    aligned = og.align_sheet(rotated)
    assert aligned.shape[0] == int(og.PAGE_H_PT * og.ZOOM)
    assert aligned.shape[1] == int(og.PAGE_W_PT * og.ZOOM)

    # 정렬된 이미지에서 다시 테두리를 검출하면 캔버스 기준 좌표와 거의 일치해야 함
    aligned_gray = cv2.cvtColor(aligned, cv2.COLOR_BGR2GRAY)
    corners = og.find_table_border(aligned_gray)
    assert corners is not None
    expected_tl = (og.TABLE_BORDER_PT[0] * og.ZOOM, og.TABLE_BORDER_PT[1] * og.ZOOM)
    assert abs(corners[0][0] - expected_tl[0]) < 20
    assert abs(corners[0][1] - expected_tl[1]) < 20


def test_align_sheet_raises_on_blank_image():
    blank = np.full((500, 700, 3), 255, dtype=np.uint8)
    try:
        og.align_sheet(blank)
        assert False, "AlignmentError가 발생해야 함"
    except og.AlignmentError:
        pass


def _draw_filled_bubble(img_bgr, cx_pt, cy_pt, zoom=og.ZOOM, radius_pt=og.BUBBLE_SIZE / 2 * 0.6):
    cx_px = int(cx_pt * zoom)
    cy_px = int(cy_pt * zoom)
    r_px = int(radius_pt * zoom)
    cv2.circle(img_bgr, (cx_px, cy_px), r_px, (0, 0, 0), -1)


def test_recognize_sheet_end_to_end():
    canonical = _render_template_page()
    img_bgr = cv2.cvtColor(canonical, cv2.COLOR_RGB2BGR)

    # 학번 12345678 마킹
    student_id = "12345678"
    for col, ch in enumerate(student_id):
        cx, cy = og.id_bubble_center_pt(col, int(ch))
        _draw_filled_bubble(img_bgr, cx, cy)

    # Q1=3, Q2=미응답, Q3=1과4 중복 마킹
    cx, cy = og.answer_bubble_center_pt(1, 3)
    _draw_filled_bubble(img_bgr, cx, cy)
    for opt in (1, 4):
        cx, cy = og.answer_bubble_center_pt(3, opt)
        _draw_filled_bubble(img_bgr, cx, cy)

    result = og.recognize_sheet(img_bgr, num_questions=3)

    assert result["student_id"] == "12345678", result["student_id"]
    assert result["id_flagged_cols"] == []
    assert result["answers"][1] == 3
    assert result["answers"][2] == og.BLANK_LABEL
    assert result["answers"][3] == "중복(1,4)"
    assert set(result["flagged_questions"]) == {2, 3}


def test_recognize_sheet_raises_alignment_error_on_blank():
    blank = np.full((500, 700, 3), 255, dtype=np.uint8)
    try:
        og.recognize_sheet(blank, num_questions=10)
        assert False, "AlignmentError가 발생해야 함"
    except og.AlignmentError:
        pass


def test_grade_sheet_basic():
    answers = {1: 3, 2: og.BLANK_LABEL, 3: 5, 4: "중복(1,4)"}
    key = {1: 3, 2: 1, 3: 2, 4: 1}
    score, wrong = og.grade_sheet(answers, key)
    assert score == 1
    assert wrong == [2, 3, 4]


def test_write_result_excel():
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "result.xlsx")
        key = {1: 3, 2: 1}
        records = [
            {
                "label": "s1.png",
                "student_id": "12345678",
                "answers": {1: 3, 2: og.BLANK_LABEL},
                "score": 1,
                "wrong": [2],
                "flagged_questions": [2],
                "id_flagged_cols": [],
            }
        ]
        og.write_result_excel(path, records, key, failed_labels=["broken.png"])

        wb = openpyxl.load_workbook(path)
        ws = wb["채점결과"]
        header = [c.value for c in ws[1]]
        assert header == ["학번", "Q1", "Q2", "점수", "오답문항", "확인필요"]
        row2 = [c.value for c in ws[2]]
        assert row2[0] == "12345678"
        assert row2[1] == 3
        assert row2[2] == og.BLANK_LABEL
        assert row2[3] == 1
        assert row2[4] == "2"
        assert row2[5] == "Q2"

        fail_ws = wb["정렬실패"]
        assert fail_ws["A2"].value == "broken.png"


def test_save_debug_overlay():
    canonical = _render_template_page()
    img_bgr = cv2.cvtColor(canonical, cv2.COLOR_RGB2BGR)
    recognition = {
        "student_id": "12345678",
        "id_flagged_cols": [],
        "answers": {1: 3, 2: og.BLANK_LABEL},
        "flagged_questions": [2],
    }
    with tempfile.TemporaryDirectory() as d:
        out_path = os.path.join(d, "s1_debug.png")
        og.save_debug_overlay(img_bgr, recognition, out_path)
        assert os.path.exists(out_path)
        saved = cv2.imread(out_path)
        assert saved.shape == img_bgr.shape


ALL_TESTS = [
    test_answer_bubble_center_q1_option1,
    test_answer_bubble_center_q21_option1,
    test_answer_bubble_center_q100_option5,
    test_id_bubble_center_col0_digit0,
    test_id_bubble_center_col7_digit9,
    test_load_answer_key_csv,
    test_load_answer_key_missing_question_raises,
    test_load_scan_images_single_page_pdf,
    test_load_scan_images_image_file,
    test_sample_fill_ratio_filled_circle,
    test_sample_fill_ratio_blank,
    test_detect_marked_index_blank,
    test_detect_marked_index_single,
    test_detect_marked_index_multi,
    test_find_table_border_on_clean_render,
    test_align_sheet_recovers_rotated_scan,
    test_align_sheet_raises_on_blank_image,
    test_recognize_sheet_end_to_end,
    test_recognize_sheet_raises_alignment_error_on_blank,
    test_grade_sheet_basic,
    test_write_result_excel,
    test_save_debug_overlay,
]

if __name__ == "__main__":
    for t in ALL_TESTS:
        t()
        print(f"OK  {t.__name__}")
    print(f"\n{len(ALL_TESTS)} tests passed")
