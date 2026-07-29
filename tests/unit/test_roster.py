from __future__ import annotations

from datetime import date
from zipfile import ZIP_BZIP2, ZipFile

import pytest
from openpyxl import Workbook, load_workbook

from omr_grader.domain.enums import RosterRowStatus
from omr_grader.domain.errors import Err, Ok
from omr_grader.ingestion import roster


def _book(path, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "명단"
    sheet.append(roster.ROSTER_HEADERS)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


def _append_member(path, name, content=b"hostile"):
    with ZipFile(path, "a") as package:
        package.writestr(name, content)


def _mark_member_encrypted(path, name):
    with ZipFile(path) as package:
        member = package.getinfo(name)
    payload = bytearray(path.read_bytes())
    payload[member.header_offset + 6] |= 0x1
    offset = 0
    while True:
        offset = payload.find(b"PK\x01\x02", offset)
        assert offset >= 0
        filename_length = int.from_bytes(payload[offset + 28 : offset + 30], "little")
        filename = bytes(payload[offset + 46 : offset + 46 + filename_length]).decode()
        if filename == name:
            payload[offset + 8] |= 0x1
            path.write_bytes(payload)
            return
        offset += 46 + filename_length


def test_roster_keeps_text_ids_and_marks_duplicate_name_conflicts(tmp_path):
    path = tmp_path / "roster.xlsx"
    _book(path, [("1", "00123456", "가"), ("2", "00123456", "나"), ("3", "", "")])

    result = roster.import_roster(str(path), "명단")

    assert isinstance(result, Ok)
    first, second, invalid = result.value.rows
    assert first.student_id == "00123456"
    assert first.status is RosterRowStatus.DUPLICATE_ID
    assert first.issues == ("duplicate_id", "name_conflict")
    assert first.roster_row_id != second.roster_row_id
    assert invalid.student_id is None
    assert invalid.status is RosterRowStatus.INVALID_ID


def test_roster_ignores_physical_blank_rows_and_marks_blank_ids_invalid(tmp_path):
    path = tmp_path / "roster.xlsx"
    _book(path, [(None, None, None), ("1", None, None)])

    result = roster.import_roster(str(path), "명단")

    assert isinstance(result, Ok)
    assert len(result.value.rows) == 1
    entry = result.value.rows[0]
    assert entry.student_id is None
    assert entry.name == ""
    assert entry.status is RosterRowStatus.INVALID_ID


def test_roster_rejects_formula_and_n_plus_one_rows(tmp_path, monkeypatch):
    path = tmp_path / "roster.xlsx"
    _book(path, [("1", "00123456", "가"), ("2", "00123457", "나")])
    monkeypatch.setattr(roster, "MAX_ROWS", 2)

    result = roster.import_roster(str(path), "명단")

    assert isinstance(result, Err)
    assert result.errors[0].code == "XLSX_DIMENSION_QUOTA"


@pytest.mark.parametrize(
    ("row", "code"),
    [
        (("=1", "00123456", "가"), "XLSX_FORMULA_FORBIDDEN"),
        ((1, "00123456", "가"), "XLSX_CELL_TYPE"),
        (("1", 12345678, "가"), "XLSX_CELL_TYPE"),
        (("1", True, "가"), "XLSX_CELL_TYPE"),
        (("1", "00123456", date(2026, 1, 1)), "XLSX_CELL_TYPE"),
        (("1", "00123456", "#DIV/0!"), "XLSX_CELL_TYPE"),
    ],
)
def test_roster_rejects_formula_and_non_text_physical_cells(tmp_path, row, code):
    path = tmp_path / "roster.xlsx"
    _book(path, [row])
    if row[2] == "#DIV/0!":
        workbook = load_workbook(path)
        workbook["명단"]["C2"].data_type = "e"
        workbook.save(path)

    result = roster.import_roster(str(path), "명단")

    assert isinstance(result, Err)
    assert result.errors[0].code == code


@pytest.mark.parametrize(
    ("name", "content"),
    [
        ("xl/vbaProject.bin", b"macro"),
        ("EncryptionInfo", b"encrypted"),
        ("EncryptedPackage", b"encrypted"),
        ("xl/externalLinks/externalLink1.xml", b"link"),
        ("xl/connections2.xml", b"connection"),
        ("xl/embeddings/oleObject1.bin", b"ole"),
        ("xl/ddeLink.xml", b"dde"),
        ("xl/activeX/activeX1.bin", b"activeX"),
        (
            "xl/_rels/workbook-hostile.rels",
            (
                b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
                b'relationships"><Relationship Id="rId1" Type="hostile" '
                b'TargetMode="ExTeRnAl" Target="https://hostile.invalid"/></Relationships>'
            ),
        ),
    ],
)
def test_roster_rejects_hostile_package_features(tmp_path, name, content):
    path = tmp_path / "roster.xlsx"
    _book(path, [("1", "00123456", "가")])
    _append_member(path, name, content)

    result = roster.import_roster(str(path), "명단")

    assert isinstance(result, Err)
    assert result.errors[0].code == "XLSX_FORBIDDEN_FEATURE"


@pytest.mark.parametrize(
    ("selected_rows", "accepted_rows"),
    [(0, 0), (1, 1), (2, 2), (3, None), (4, None)],
)
def test_roster_n_three_row_quota_counts_header_separately(
    tmp_path, monkeypatch, selected_rows, accepted_rows
):
    path = tmp_path / "roster.xlsx"
    _book(
        path,
        [
            (str(index), f"001234{index:02d}", f"학생{index}")
            for index in range(1, selected_rows + 1)
        ],
    )
    monkeypatch.setattr(roster, "MAX_ROWS", 3)

    result = roster.import_roster(str(path), "명단")

    if accepted_rows is None:
        assert isinstance(result, Err)
        assert result.errors[0].code == "XLSX_DIMENSION_QUOTA"
    else:
        assert isinstance(result, Ok)
        assert len(result.value.rows) == accepted_rows


def test_roster_rejects_malformed_relationship_xml(tmp_path):
    path = tmp_path / "roster.xlsx"
    _book(path, [("1", "00123456", "가")])
    _append_member(path, "xl/_rels/hostile.rels", b"<Relationships>")

    result = roster.import_roster(str(path), "명단")

    assert isinstance(result, Err)
    assert result.errors[0].code == "XLSX_INVALID_PACKAGE"


@pytest.mark.parametrize("name", ("xl/worksheets/encrypted.xml", "irrelevant.bin"))
def test_roster_rejects_encrypted_members_before_zip_reads(tmp_path, name):
    path = tmp_path / "roster.xlsx"
    _book(path, [("1", "00123456", "가")])
    _append_member(path, name)
    _mark_member_encrypted(path, name)

    result = roster.import_roster(str(path), "명단")

    assert isinstance(result, Err)
    assert result.errors[0].code == "XLSX_INVALID_PACKAGE"


def test_roster_rejects_unsupported_member_compression_before_zip_reads(tmp_path):
    path = tmp_path / "roster.xlsx"
    _book(path, [("1", "00123456", "가")])
    with ZipFile(path, "a", compression=ZIP_BZIP2) as package:
        package.writestr("irrelevant.bin", b"hostile")

    result = roster.import_roster(str(path), "명단")

    assert isinstance(result, Err)
    assert result.errors[0].code == "XLSX_INVALID_PACKAGE"


@pytest.mark.parametrize(
    ("content", "code"),
    [
        (b"<Relationships/>", "XLSX_INVALID_PACKAGE"),
        (
            b'<Relationships xmlns="urn:wrong"><Relationship Id="rId1" Type="type" '
            b'Target="target"/></Relationships>',
            "XLSX_INVALID_PACKAGE",
        ),
        (
            b'<relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            b'relationships"/>',
            "XLSX_INVALID_PACKAGE",
        ),
        (
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            b'relationships"><Relationship xmlns="urn:wrong" Id="rId1" Type="type" '
            b'Target="target"/></Relationships>',
            "XLSX_INVALID_PACKAGE",
        ),
        (
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            b'relationships"><relationship Id="rId1" Type="type" Target="target"/>'
            b"</Relationships>",
            "XLSX_INVALID_PACKAGE",
        ),
        (
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            b'relationships"><Relationship Id="rId1" Type="type" Target="target">'
            b"<nested/></Relationship></Relationships>",
            "XLSX_INVALID_PACKAGE",
        ),
        (
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            b'relationships" Extra="value"/>',
            "XLSX_INVALID_PACKAGE",
        ),
        (
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            b'relationships">text</Relationships>',
            "XLSX_INVALID_PACKAGE",
        ),
        (
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            b'relationships"><Relationship Id="rId1" Type="type" Target="target">'
            b"text</Relationship></Relationships>",
            "XLSX_INVALID_PACKAGE",
        ),
        (
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            b'relationships"><Relationship Id="rId1" Type="type" Target="target"/>'
            b"tail</Relationships>",
            "XLSX_INVALID_PACKAGE",
        ),
        (
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            b'relationships"><Relationship Id="rId1" Type="type" Target="target" '
            b'Unknown="value"/></Relationships>',
            "XLSX_INVALID_PACKAGE",
        ),
        (
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            b'relationships"><Relationship iD="rId1" Type="type" Target="target"/>'
            b"</Relationships>",
            "XLSX_INVALID_PACKAGE",
        ),
        (
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            b'relationships" xmlns:bad="urn:bad"><Relationship Id="rId1" Type="type" '
            b'Target="target" bad:TargetMode="Internal"/></Relationships>',
            "XLSX_INVALID_PACKAGE",
        ),
        (
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            b'relationships" xmlns:r="http://schemas.openxmlformats.org/package/2006/'
            b'relationships"><Relationship Id="rId1" r:Id="rId2" Type="type" '
            b'Target="target"/></Relationships>',
            "XLSX_INVALID_PACKAGE",
        ),
        (
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            b'relationships"><Relationship Id="rId1" Type="type"/></Relationships>',
            "XLSX_INVALID_PACKAGE",
        ),
        (
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            b'relationships"><Relationship Id=" " Type="type" Target="target"/>'
            b"</Relationships>",
            "XLSX_INVALID_PACKAGE",
        ),
        (
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            b'relationships"><Relationship Id="rId1" Type=" " Target="target"/>'
            b"</Relationships>",
            "XLSX_INVALID_PACKAGE",
        ),
        (
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            b'relationships"><Relationship Id="rId1" Type="type" Target=" "/>'
            b"</Relationships>",
            "XLSX_INVALID_PACKAGE",
        ),
        (
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            b'relationships"><Relationship Id="rId1" Type="type" Target="target"/>'
            b'<Relationship Id="rId1" Type="type" Target="target"/></Relationships>',
            "XLSX_INVALID_PACKAGE",
        ),
        (
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            b'relationships"><Relationship Id="rId1" Type="type" Target="target" '
            b'TargetMode="remote"/></Relationships>',
            "XLSX_INVALID_PACKAGE",
        ),
        (
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            b'relationships"><Relationship Id="rId1" Type="type" Target="target" '
            b'TargetMode=" \tExTeRnAl\n"/></Relationships>',
            "XLSX_FORBIDDEN_FEATURE",
        ),
    ],
)
def test_roster_rejects_nonconforming_opc_relationships(tmp_path, content, code):
    path = tmp_path / "roster.xlsx"
    _book(path, [("1", "00123456", "가")])
    _append_member(path, "xl/_rels/hostile.rels", content)

    result = roster.import_roster(str(path), "명단")

    assert isinstance(result, Err)
    assert result.errors[0].code == code


def test_roster_accepts_exact_opc_internal_relationship(tmp_path):
    path = tmp_path / "roster.xlsx"
    _book(path, [("1", "00123456", "가")])
    _append_member(
        path,
        "xl/_rels/internal.rels",
        (
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            b'relationships"><Relationship Id="rId1" Type="type" Target="target" '
            b'TargetMode=" InTeRnAl "/></Relationships>'
        ),
    )

    result = roster.import_roster(str(path), "명단")

    assert isinstance(result, Ok)


def test_roster_row_identity_is_stable_and_has_source_provenance(tmp_path):
    path = tmp_path / "roster.xlsx"
    _book(path, [("1", "00123456", "가"), (None, None, None), ("2", "00123457", "나")])

    first = roster.import_roster(str(path), "명단")
    second = roster.import_roster(str(path), "명단")

    assert isinstance(first, Ok)
    assert isinstance(second, Ok)
    assert first.value.source_name == path.name
    assert first.value.source_sha256 == second.value.source_sha256
    assert [row.roster_row_id for row in first.value.rows] == [
        row.roster_row_id for row in second.value.rows
    ]
    assert [row.source_row_number for row in first.value.rows] == [2, 4]
    assert [row.input_ordinal for row in first.value.rows] == [0, 2]


def test_roster_package_and_sheet_boundaries(tmp_path, monkeypatch):
    path = tmp_path / "roster.xlsx"
    _book(path, [("1", "00123456", "가")])
    data = path.read_bytes()
    for size, code in ((0, "XLSX_PACKAGE_QUOTA"), (1, "XLSX_INVALID_PACKAGE")):
        malformed = tmp_path / f"{size}-byte.xlsx"
        malformed.write_bytes(b"x" * size)
        result = roster.import_roster(str(malformed), "명단")
        assert isinstance(result, Err)
        assert result.errors[0].code == code

    monkeypatch.setattr(roster, "MAX_PACKAGE_BYTES", len(data))
    assert isinstance(roster.import_roster(str(path), "명단"), Ok)
    monkeypatch.setattr(roster, "MAX_PACKAGE_BYTES", len(data) - 1)
    too_large = roster.import_roster(str(path), "명단")
    assert isinstance(too_large, Err)
    assert too_large.errors[0].code == "XLSX_PACKAGE_QUOTA"

    monkeypatch.setattr(roster, "MAX_PACKAGE_BYTES", len(data))
    missing_sheet = roster.import_roster(str(path), "없음")
    assert isinstance(missing_sheet, Err)
    assert missing_sheet.errors[0].code == "XLSX_SHEET_NOT_FOUND"

    monkeypatch.setattr(roster, "MAX_ROWS", 1)
    at_limit = roster.import_roster(str(path), "명단")
    assert isinstance(at_limit, Err)
    assert at_limit.errors[0].code == "XLSX_DIMENSION_QUOTA"
