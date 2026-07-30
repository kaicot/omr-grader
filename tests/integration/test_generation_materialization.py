from __future__ import annotations

import json
from pathlib import Path
from types import SimpleNamespace

import cv2
import numpy as np
import pytest
from openpyxl import Workbook, load_workbook

from omr_grader.application.dto import (
    AnswerKeyValidation,
    ImportResponseCommand,
    RegradeCommand,
    ResponseBookRequest,
)
from omr_grader.application.grading_use_case import GradingUseCase
from omr_grader.application.response_import_use_case import ResponseImportUseCase
from omr_grader.domain.enums import (
    AnswerKeySnapshotKind,
    AnswerStatus,
    ExamTerm,
    KeyQuestionStatus,
    OperationKind,
    RosterSnapshotKind,
    SourceKind,
    StudentIdStatus,
)
from omr_grader.domain.errors import Ok
from omr_grader.domain.models import (
    AnswerKeyEntry,
    AnswerKeySnapshot,
    AnswerValue,
    EffectiveResponse,
    RosterSnapshot,
)
from omr_grader.infrastructure.generation_materializer import (
    GenerationMaterializer,
    StagingToken,
    _read_combined,
)
from omr_grader.infrastructure.grading_runtime import (
    CommittedGradingSnapshotReader,
    ResponseImportCommitCoordinator,
)
from omr_grader.infrastructure.session_store import (
    SessionStore,
    _preserved_artifact,
    _refresh_result_view,
)
from omr_grader.workbooks.answer_key import import_answer_key
from omr_grader.workbooks.schemas import RESPONSE_HEADERS, RESPONSE_SHEET_NAME


def test_store_owned_staging_token_rejects_traversal_operation_paths(tmp_path: Path) -> None:
    token = StagingToken(tmp_path / "staging")
    token._root.mkdir()

    with pytest.raises(ValueError, match="escapes staging"):
        token.path("../published")


def test_store_owned_staging_token_writes_only_beneath_its_staging_tree(tmp_path: Path) -> None:
    token = StagingToken(tmp_path / "staging")
    token._root.mkdir()
    token.write_json("sidecars/provenance.json", {"schema_version": 1})

    assert json.loads((token._root / "sidecars" / "provenance.json").read_text()) == {
        "schema_version": 1
    }
    assert not (tmp_path / "sidecars" / "provenance.json").exists()


def test_malformed_parent_semantic_inputs_fail_closed(tmp_path: Path) -> None:
    generation = tmp_path / "generation"
    generation.mkdir()
    (generation / "semantic_inputs.json").write_text('{"combined":[]}', encoding="utf-8")

    with pytest.raises(ValueError, match="committed semantic inputs are corrupt"):
        _read_combined(generation)


@pytest.mark.parametrize(
    "path, preserved",
    (
        ("recognition/page-1.json", False),
        ("evidence/page-1.png", False),
        ("images/page-1.png", True),
        ("correction_history.json", False),
        ("artifacts/02_score_old.xlsx", False),
        ("artifacts/03_final_old.xlsx", False),
        ("corrected_overlay.json", False),
        ("detail_index.json", False),
        ("detail/work-item-1.json", False),
        ("semantic_inputs.json", False),
    ),
)
def test_lifecycle_materialization_copies_only_immutable_or_audit_artifacts(
    path: str, preserved: bool
) -> None:
    assert _preserved_artifact(path, OperationKind.CORRECT) is preserved


def test_finalization_preserves_the_last_score_workbook() -> None:
    assert _preserved_artifact(
        "02_score_시험_260730_120000_채점결과.xlsx", OperationKind.FINALIZE
    )
    assert not _preserved_artifact(
        "artifacts/02_score_시험_260730_120000_채점결과.xlsx",
        OperationKind.FINALIZE,
    )


def test_regrade_compacts_to_one_minimal_current_generation() -> None:
    retained = (
        "images/page-1.png",
        "01_ocr_시험_260730_120000_응답결과.xlsx",
    )
    discarded = (
        "recognition/page-1.json",
        "evidence/page-1.json",
        "automatic/page-1.json",
        "좌표데이터/page-1.json",
        "01_인식결과_이미지/page-1.png",
        "correction_events.json",
        "correction_history.json",
        "details/page-1.json",
        "detail_index.json",
        "projection_request.json",
    )

    assert all(_preserved_artifact(path, OperationKind.REGRADE) for path in retained)
    assert not any(_preserved_artifact(path, OperationKind.REGRADE) for path in discarded)


def test_compact_result_view_exposes_retained_source_scans_without_copying(
    tmp_path: Path,
) -> None:
    session = tmp_path / "session"
    generation = session / "generations" / "g00000002_generation"
    source = generation / "images" / "work-item-1.png"
    source.parent.mkdir(parents=True)
    source.write_bytes(b"normalized scan")

    _refresh_result_view(session, generation)

    visible = session / "00_원본스캔" / "work-item-1.png"
    assert visible.read_bytes() == b"normalized scan"
    assert visible.stat().st_ino == source.stat().st_ino


def test_grading_review_images_are_jpeg_quality_80_and_bounded_to_2000px(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    root = tmp_path / "generation"
    source = root / "images" / "work-item-1.png"
    source.parent.mkdir(parents=True)
    assert cv2.imwrite(str(source), np.zeros((3000, 2500, 3), dtype=np.uint8))
    unanswered = AnswerValue((), AnswerStatus.UNASKED)
    response = EffectiveResponse(
        "work-item-1",
        SourceKind.IMAGE,
        "scan.png",
        None,
        StudentIdStatus.UNREADABLE,
        (unanswered,) * 100,
    )
    answer_key = AnswerKeySnapshot(
        1,
        AnswerKeySnapshotKind.WORKBOOK,
        "answer-key.xlsx",
        "a" * 64,
        "정답표",
        "v1",
        tuple(
            AnswerKeyEntry(question, unanswered, "0", KeyQuestionStatus.UNASKED)
            for question in range(1, 101)
        ),
        (),
    )
    combined = {
        "responses": [response.to_dict()],
        "roster": RosterSnapshot(
            1,
            RosterSnapshotKind.NONE,
            None,
            None,
            None,
            "v1",
            (),
            (),
        ).to_dict(),
        "scores": [],
        "answer_key": answer_key.to_dict(),
    }
    page = SimpleNamespace(
        work_item_id="work-item-1",
        page_ref=SimpleNamespace(work_item_id="work-item-1"),
        evidence=(),
        answers=(),
        to_dict=lambda: {"work_item_id": "work-item-1"},
    )
    request = SimpleNamespace(
        token=StagingToken(root),
        record=SimpleNamespace(
            session_id="session-1",
            revision=2,
        ),
        generation_id="generation-2",
        recognition_artifacts=None,
    )
    monkeypatch.setattr(
        "omr_grader.infrastructure.generation_materializer.render_scored_overlay",
        lambda *_: Ok(np.zeros((3000, 2500, 3), dtype=np.uint8)),
    )
    encoded: list[tuple[str, tuple[int, ...]]] = []
    encode = cv2.imencode

    def capture_encode(
        extension: str, image: np.ndarray, parameters: list[int] | None = None
    ) -> tuple[bool, np.ndarray]:
        encoded.append((extension, tuple(parameters or ())))
        return encode(extension, image, parameters or ())

    monkeypatch.setattr(
        "omr_grader.infrastructure.generation_materializer.cv2.imencode",
        capture_encode,
    )

    GenerationMaterializer()._write_details(
        request,  # type: ignore[arg-type]
        combined,
        (page,),  # type: ignore[arg-type]
        persist_recognition=False,
    )

    assert encoded == [
        (
            ".jpg",
            (
                cv2.IMWRITE_JPEG_QUALITY,
                80,
                cv2.IMWRITE_JPEG_OPTIMIZE,
                1,
            ),
        )
    ]
    review = root / "02_채점결과_이미지" / "work-item-1.jpg"
    payload = review.read_bytes()
    assert payload.startswith(b"\xff\xd8") and payload.endswith(b"\xff\xd9")
    decoded = cv2.imdecode(np.frombuffer(payload, dtype=np.uint8), cv2.IMREAD_COLOR)
    assert decoded is not None
    assert max(decoded.shape[:2]) <= 2000
    assert not (root / "02_채점결과_이미지" / "work-item-1.png").exists()


def test_regrade_after_compaction_uses_finalized_responses_and_new_answer_key(
    tmp_path: Path,
) -> None:
    source = tmp_path / "responses.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = RESPONSE_SHEET_NAME
    sheet.append(list(RESPONSE_HEADERS))
    sheet.append([1, "scan.png", "00123456", "홍길동", "1", *("" for _ in range(99)), ""])
    workbook.save(source)
    store = SessionStore(tmp_path / "sessions")
    importer = ResponseImportUseCase(ResponseImportCommitCoordinator(store))
    validation = importer.validate_response_book(
        ResponseBookRequest(
            str(source),
            RESPONSE_SHEET_NAME,
            "Math",
            2026,
            ExamTerm.FIRST,
        )
    )
    assert isinstance(validation, Ok)
    imported = importer.import_response_book(
        ImportResponseCommand(
            validation.value.validation_token,
            "imported-session",
            "import-operation",
            0,
        )
    )
    assert isinstance(imported, Ok)
    unanswered = AnswerValue((), AnswerStatus.UNASKED)
    key = AnswerKeySnapshot(
        1,
        AnswerKeySnapshotKind.WORKBOOK,
        "answer-key.xlsx",
        "b" * 64,
        "정답표",
        "v1",
        (
            AnswerKeyEntry(
                1,
                AnswerValue((1,), AnswerStatus.NORMAL),
                "1",
                KeyQuestionStatus.ANSWER,
            ),
        )
        + tuple(
            AnswerKeyEntry(question, unanswered, "0", KeyQuestionStatus.UNASKED)
            for question in range(2, 101)
        ),
        (),
    )

    class AnswerKeys:
        def validate_answer_key(self, _request):
            return Ok(AnswerKeyValidation(key))

    use_case = GradingUseCase(
        CommittedGradingSnapshotReader(store),
        AnswerKeys(),  # type: ignore[arg-type]
        store,
    )
    first = use_case.regrade(
        RegradeCommand(
            imported.value.session_id,
            1,
            "answer-key.xlsx",
            "정답표",
            "regrade-operation-1",
        )
    )
    assert isinstance(first, Ok)
    session = next(path for path in store.root.iterdir() if (path / "CURRENT.json").exists())
    pointer = json.loads((session / "CURRENT.json").read_text(encoding="utf-8"))
    compacted = session / pointer["generation_relpath"]
    assert not (compacted / "projection_request.json").exists()
    response_workbooks = tuple(compacted.glob("01_ocr_*_응답결과.xlsx"))
    assert len(response_workbooks) == 1
    response_workbook = load_workbook(response_workbooks[0], read_only=True)
    try:
        response_properties = {
            item.name: item.value for item in response_workbook.custom_doc_props
        }
    finally:
        response_workbook.close()
    assert response_properties["revision"] == "2"
    answer_key_workbooks = tuple(compacted.glob("정답표_*.xlsx"))
    assert len(answer_key_workbooks) == 1
    retained_key = import_answer_key(str(answer_key_workbooks[0]), "정답표")
    assert isinstance(retained_key, Ok)
    assert retained_key.value.entries == key.entries
    assert retained_key.value.sheet_name == key.sheet_name

    second = use_case.regrade(
        RegradeCommand(
            imported.value.session_id,
            2,
            "answer-key.xlsx",
            "정답표",
            "regrade-operation-2",
        )
    )

    assert isinstance(second, Ok)
    assert second.value.revision == 3
