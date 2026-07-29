from __future__ import annotations

from datetime import date
from hashlib import sha256
from io import BytesIO
from zipfile import ZipFile

import pytest
from openpyxl import Workbook, load_workbook

from omr_grader.domain.enums import AnswerStatus
from omr_grader.domain.errors import Err, Ok
from omr_grader.workbooks import response_import
from omr_grader.workbooks.response_import import parse_response_book
from omr_grader.workbooks.schemas import RESPONSE_HEADERS, RESPONSE_SHEET_NAME


def _workbook(path, rows: list[list[object]]) -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = RESPONSE_SHEET_NAME
    sheet.append(list(RESPONSE_HEADERS))
    for row in rows:
        sheet.append(row)
    book.save(path)


def _row(serial: int, student_id: str = "00123456", answer: object = "13") -> list[object]:
    return [serial, "scan-01.jpg", student_id, "홍길동", *([answer] + [""] * 99), ""]


def test_response_workbook_roundtrips_canonical_rows_and_text_id(tmp_path) -> None:
    path = tmp_path / "responses.xlsx"
    _workbook(path, [_row(7), _row(8, "00123456", "2")])
    with path.open("rb") as source:
        parsed = parse_response_book(
            source,
            sheet_name=RESPONSE_SHEET_NAME,
            session_id="import-session",
            source_sha256=sha256(path.read_bytes()).hexdigest(),
        )
    assert isinstance(parsed, Ok)
    assert [row.serial for row in parsed.value] == [7, 8]
    assert parsed.value[0].raw_student_id == "00123456"
    assert parsed.value[0].answers[0].choices == (1, 3)
    assert parsed.value[0].answers[0].status is AnswerStatus.MULTIPLE
    assert parsed.value[0].work_item_id != parsed.value[1].work_item_id


def test_response_workbook_rejects_numeric_id_without_normalizing(tmp_path) -> None:
    path = tmp_path / "numeric-id.xlsx"
    _workbook(path, [_row(1, 12345678)])
    with path.open("rb") as source:
        parsed = parse_response_book(
            source,
            sheet_name=RESPONSE_SHEET_NAME,
            session_id="import-session",
            source_sha256=sha256(path.read_bytes()).hexdigest(),
        )
    assert isinstance(parsed, Err)
    assert parsed.errors[0].code == "XLSX_ID_TEXT_REQUIRED"


def test_response_workbook_rejects_duplicate_serial_and_formula(tmp_path) -> None:
    path = tmp_path / "hostile.xlsx"
    _workbook(path, [_row(1), _row(1)])
    book = load_workbook(path)
    book[RESPONSE_SHEET_NAME]["E2"] = "=1+1"
    book.save(path)
    with path.open("rb") as source:
        parsed = parse_response_book(
            source,
            sheet_name=RESPONSE_SHEET_NAME,
            session_id="import-session",
            source_sha256=sha256(path.read_bytes()).hexdigest(),
        )
    assert isinstance(parsed, Err)
    assert parsed.errors[0].code == "XLSX_FORMULA_FORBIDDEN"


def _package(entries: dict[str, str | bytes]) -> BytesIO:
    source = BytesIO()
    with ZipFile(source, "w") as archive:
        for name, payload in entries.items():
            archive.writestr(name, payload)
    source.seek(0)
    return source


@pytest.mark.parametrize(
    ("entry_name", "payload", "code"),
    (
        ("xl/VBAPROJECT.BIN", b"x", "XLSX_DANGEROUS_FEATURE"),
        ("ENCRYPTIONINFO", b"x", "XLSX_DANGEROUS_FEATURE"),
        ("xl/EXTERNALLINKS/link1.xml", "<root />", "XLSX_DANGEROUS_FEATURE"),
        ("xl/CONNECTIONS.xml", "<root />", "XLSX_DANGEROUS_FEATURE"),
        ("xl/OLEOBJECTS/object1.bin", b"x", "XLSX_DANGEROUS_FEATURE"),
        ("xl/ACTIVEX/activeX1.bin", b"x", "XLSX_DANGEROUS_FEATURE"),
        (
            "XL/_RELS/WORKBOOK.XML.RELS",
            '<Relationships><Relationship Id="rId1" Type="x" Target="x" '
            'targetmode="external" /></Relationships>',
            "XLSX_DANGEROUS_FEATURE",
        ),
        ("xl/worksheets/sheet1.xml", "<root><F>1+1</F></root>", "XLSX_FORMULA_FORBIDDEN"),
        ("xl/worksheets/sheet1.xml", "<root><DDELINK /></root>", "XLSX_DANGEROUS_FEATURE"),
        ("xl/_rels/workbook.xml.rels", "<not-relationships />", "XLSX_ARCHIVE_QUOTA"),
        (
            "xl/_rels/workbook.xml.rels",
            "<Relationships><unexpected /></Relationships>",
            "XLSX_ARCHIVE_QUOTA",
        ),
        (
            "xl/_rels/workbook.xml.rels",
            '<Relationships><Relationship Id="rId1" Type="x" /></Relationships>',
            "XLSX_ARCHIVE_QUOTA",
        ),
    ),
)
def test_response_package_rejects_casefolded_hostile_and_malformed_parts(
    entry_name: str, payload: str | bytes, code: str
) -> None:
    scanned = response_import._scan_package(_package({entry_name: payload}))

    assert isinstance(scanned, Err)
    assert scanned.errors[0].code == code


@pytest.mark.parametrize(
    ("kind", "value", "expected_code"),
    (
        ("serial", True, "XLSX_ROW_INVALID"),
        ("serial", date(2026, 1, 1), "XLSX_ROW_INVALID"),
        ("id", True, "XLSX_ID_TEXT_REQUIRED"),
        ("id", 12345678.0, "XLSX_ID_TEXT_REQUIRED"),
        ("id", date(2026, 1, 1), "XLSX_ID_TEXT_REQUIRED"),
        ("id_error", "#VALUE!", "XLSX_ID_TEXT_REQUIRED"),
        ("answer_error", "13", "XLSX_ROW_INVALID"),
        ("answer", True, "XLSX_ROW_INVALID"),
        ("answer", date(2026, 1, 1), "XLSX_ROW_INVALID"),
    ),
)
def test_response_workbook_rejects_physical_nontext_and_error_cells(
    tmp_path, kind: str, value: object, expected_code: str
) -> None:
    path = tmp_path / f"{kind}.xlsx"
    row = _row(1)
    if kind == "serial":
        row[0] = value
    elif kind.startswith("id"):
        row[2] = value
    else:
        row[4] = value
    _workbook(path, [row])
    if kind.endswith("_error"):
        book = load_workbook(path)
        cell = book[RESPONSE_SHEET_NAME]["C2" if kind == "id_error" else "E2"]
        cell.data_type = "e"
        book.save(path)

    with path.open("rb") as source:
        parsed = parse_response_book(
            source,
            sheet_name=RESPONSE_SHEET_NAME,
            session_id="import-session",
            source_sha256=sha256(path.read_bytes()).hexdigest(),
        )

    assert isinstance(parsed, Err)
    assert parsed.errors[0].code == expected_code


def test_response_workbook_normalizes_text_rejects_unicode_controls_and_preserves_provenance(
    tmp_path,
) -> None:
    path = tmp_path / "canonical.xlsx"
    normalized = _row(1)
    normalized[1] = "  scan\u00a0name.jpg  "
    _workbook(path, [normalized, _row(2)])

    def parse() -> Ok | Err:
        with path.open("rb") as source:
            return parse_response_book(
                source,
                sheet_name=RESPONSE_SHEET_NAME,
                session_id="import-session",
                source_sha256=sha256(path.read_bytes()).hexdigest(),
            )

    first = parse()
    second = parse()
    assert isinstance(first, Ok)
    assert isinstance(second, Ok)
    assert first.value == second.value
    assert first.value[0].source_filename == "scan name.jpg"
    assert [row.input_ordinal for row in first.value] == [0, 1]
    assert [row.row_number for row in first.value] == [2, 3]
    assert all(row.source_sha256 == sha256(path.read_bytes()).hexdigest() for row in first.value)

    _workbook(path, [_row(1, "00123456"), _row(2)])
    book = load_workbook(path)
    book[RESPONSE_SHEET_NAME]["B2"] = "scan\u0080name.jpg"
    book.save(path)
    with path.open("rb") as source:
        rejected = parse_response_book(
            source,
            sheet_name=RESPONSE_SHEET_NAME,
            session_id="import-session",
            source_sha256=sha256(path.read_bytes()).hexdigest(),
        )
    assert isinstance(rejected, Err)
    assert rejected.errors[0].code == "XLSX_ROW_INVALID"


@pytest.mark.parametrize(
    ("entry_count", "expected_code"),
    ((0, None), (1, None), (2, None), (3, None), (4, "XLSX_ARCHIVE_QUOTA")),
)
def test_response_package_zip_entry_quota_boundaries_are_compact_and_fail_closed(
    monkeypatch, entry_count: int, expected_code: str | None
) -> None:
    monkeypatch.setattr(response_import, "MAX_ZIP_ENTRIES", 3)
    scanned = response_import._scan_package(
        _package({f"part-{index}": b"x" for index in range(entry_count)})
    )

    if expected_code is None:
        assert isinstance(scanned, Ok)
    else:
        assert isinstance(scanned, Err)
        assert scanned.errors[0].code == expected_code


@pytest.mark.parametrize(
    ("relationship_count", "expected_code"),
    ((0, None), (1, None), (2, None), (3, None), (4, "XLSX_ARCHIVE_QUOTA")),
)
def test_response_package_relationship_quota_boundaries_are_compact_and_fail_closed(
    monkeypatch, relationship_count: int, expected_code: str | None
) -> None:
    monkeypatch.setattr(response_import, "MAX_RELATIONSHIPS", 3)
    relationships = "".join(
        f'<Relationship Id="r{index}" Type="x" Target="part-{index}" />'
        for index in range(relationship_count)
    )
    scanned = response_import._scan_package(
        _package(
            {"xl/_rels/workbook.xml.rels": (f"<Relationships>{relationships}</Relationships>")}
        )
    )

    if expected_code is None:
        assert isinstance(scanned, Ok)
    else:
        assert isinstance(scanned, Err)
        assert scanned.errors[0].code == expected_code
