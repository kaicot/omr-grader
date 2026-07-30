from dataclasses import replace

import pytest
from openpyxl import load_workbook
from tests.unit.test_grading import _key, _response

import omr_grader.workbooks.score_book as score_book_module
from omr_grader.application.dto import ScoreInput
from omr_grader.domain.grading import score_effective
from omr_grader.workbooks.score_book import (
    score_filename,
    write_final_book,
    write_score_book,
)


def test_score_and_final_workbooks_have_exact_projection_shape(tmp_path):
    key = _key()
    first = replace(
        _response("wi_first", (1, 2)),
        corrected_targets=("id_cell:0", "answer_cell:2"),
    )
    duplicate = _response("wi_duplicate", (1,))
    scores = score_effective(ScoreInput((first, duplicate), key))
    common = dict(
        exam_name="수학/1",
        committed_at="2026-07-28T01:02:03.123456Z",
        session_id="session-1",
        revision=2,
        manifest_sha256="b" * 64,
        responses=(first, duplicate),
        key=key,
        scores=scores,
        names_by_student_id={"12345678": "=수식아님"},
    )

    score_path = write_score_book(tmp_path, **common)
    final_path = write_final_book(tmp_path, **common)
    score_book = load_workbook(score_path, data_only=False)
    final_book = load_workbook(final_path, data_only=False)
    score = score_book.active
    final = final_book.active

    expected_score_headers = (
        "석차",
        "학번",
        "이름",
        "총점",
        *(f"Q{number}" for number in range(1, 101)),
        "비고",
    )
    assert score.title == "채점결과"
    assert tuple(cell.value for cell in score[1]) == expected_score_headers
    assert score.max_column == 105 and score.max_row == 3
    assert final.title == "최종성적표"
    assert tuple(cell.value for cell in final[1]) == (
        *expected_score_headers,
        "수정여부",
        "수정문항",
        "확정일시",
    )
    assert final.max_column == 108 and final.max_row == 3
    assert score.cell(2, 1).value == 1 and score.cell(2, 1).data_type == "n"
    assert score.cell(2, 4).value == 3.5 and score.cell(2, 4).data_type == "n"
    assert tuple(score.cell(2, column).value for column in range(5, 8)) == ("O", "O", "제외")
    assert score.cell(2, 3).value == "'=수식아님"
    assert {score.cell(row, 105).value for row in (2, 3)} == {"중복확인필요"}
    assert final.cell(2, 106).value is True and final.cell(2, 106).data_type == "b"
    assert final.cell(2, 107).value == "학번1,Q2"
    assert final.cell(2, 108).value == "2026-07-28T01:02:03.123456Z"
    assert score_book.sheetnames == ["채점결과", "응답내역"]
    responses = score_book["응답내역"]
    assert tuple(cell.value for cell in responses[1]) == expected_score_headers
    assert tuple(responses.cell(2, column).value for column in range(5, 8)) == (
        "1,2",
        None,
        None,
    )
    assert final_book.sheetnames == ["최종성적표", "응답내역"]
    assert {property.name: property.value for property in score_book.custom_doc_props} == {
        "schema": "1",
        "session_id": "session-1",
        "revision": "2",
        "manifest_sha256": "b" * 64,
    }
    assert score_path.name == "02_score_수학_1_260728_100203_채점결과.xlsx"
    assert final_path.name == "03_final_수학_1_260728_100203_최종성적표.xlsx"


def test_score_book_rejects_collisions_and_leaves_no_partial_file(tmp_path, monkeypatch):
    key = _key()
    response = _response("wi", (1, 2))
    common = dict(
        exam_name="시험",
        committed_at="2026-07-28T01:02:03.123456Z",
        session_id="session-1",
        revision=1,
        manifest_sha256="b" * 64,
        responses=(response,),
        key=key,
        scores=score_effective(ScoreInput((response,), key)),
    )

    original = write_score_book(tmp_path, **common)
    with pytest.raises(FileExistsError):
        write_score_book(tmp_path, **common)
    assert original.exists()

    def fail_save(self, filename):
        raise OSError("simulated write failure")

    monkeypatch.setattr(score_book_module.Workbook, "save", fail_save)
    failed_destination = tmp_path / "failed"
    with pytest.raises(OSError, match="simulated write failure"):
        write_score_book(failed_destination, **common)
    assert not (failed_destination / score_filename("시험", common["committed_at"])).exists()
    assert not tuple(failed_destination.glob(".*.tmp"))
