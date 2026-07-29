from __future__ import annotations

import json
from concurrent.futures import ThreadPoolExecutor
from decimal import Decimal
from hashlib import sha256
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from omr_grader.application.dashboard_use_case import DashboardUseCase, _publish
from omr_grader.application.dto import (
    CollisionPolicy,
    CombinedReportRequest,
    ScoreResult,
    ScoreSet,
    ScoreStatistics,
    SnapshotRef,
)
from omr_grader.domain.enums import (
    AnswerStatus,
    ExamTerm,
    RosterRowStatus,
    RosterSnapshotKind,
    SessionState,
    SourceKind,
    StudentIdStatus,
)
from omr_grader.domain.errors import Err, ErrorInfo, Ok
from omr_grader.domain.models import (
    AnswerValue,
    EffectiveResponse,
    EvidenceSummary,
    PageFailure,
    PageRef,
    RosterEntry,
    RosterSnapshot,
    SessionRecord,
)
from omr_grader.workbooks.combined_book import CombinedExam, build_combined_workbook

STAMP = "2026-07-28T12:00:00.000000Z"
SHA = "a" * 64


def _exam() -> CombinedExam:
    return CombinedExam(
        SnapshotRef("session-1", 1, "generation-1", SHA),
        SessionRecord(
            1,
            "session-1",
            1,
            SessionState.RECOGNIZED,
            "수학",
            2026,
            ExamTerm.FIRST,
            STAMP,
            None,
            STAMP,
        ),
        RosterSnapshot(1, RosterSnapshotKind.NONE, None, None, None, "v1", (), ()),
        (),
        None,
        (),
    )


def _answers() -> tuple[AnswerValue, ...]:
    return (AnswerValue((1,), AnswerStatus.NORMAL),) * 100


def _response(
    work_item_id: str,
    student_id: str | None,
    *,
    duplicate: bool = False,
    corrected: bool = False,
    source: str | None = None,
) -> EffectiveResponse:
    return EffectiveResponse(
        work_item_id,
        SourceKind.IMAGE,
        source or f"{work_item_id}.png",
        student_id,
        StudentIdStatus.DUPLICATE
        if duplicate
        else (StudentIdStatus.NORMAL if student_id is not None else StudentIdStatus.UNREADABLE),
        _answers(),
        ("id_cell:0",) if corrected else (),
    )


def _entry(
    number: int,
    student_id: str | None,
    name: str,
    status: RosterRowStatus,
    issues: tuple[str, ...],
) -> RosterEntry:
    return RosterEntry(
        f"row-{number}", number, number, student_id or "", student_id, name, status, issues
    )


def _scores(*rows: tuple[str, Decimal | None]) -> ScoreSet:
    values = tuple(score for _, score in rows if score is not None)
    ranks = {score: 1 + sum(other > score for other in values) for score in set(values)}
    results = tuple(
        ScoreResult(work_item_id, score, None if score is None else ranks[score])
        for work_item_id, score in rows
    )
    return ScoreSet(
        Decimal("10"),
        results,
        ScoreStatistics(
            len(values),
            None if not values else sum(values) / len(values),
            None if not values else max(values),
            None if not values else min(values),
        ),
    )


def _failure() -> PageFailure:
    return PageFailure(
        1,
        PageRef(
            1,
            "session-z",
            "failed",
            SourceKind.IMAGE,
            SHA,
            "failed.png",
            "failed.png",
            1,
            None,
            0,
            0,
            "failed",
        ),
        (ErrorInfo("PAGE_FAILED", "error.page_failed"),),
        EvidenceSummary(80, 500, 0),
    )


def test_combined_book_populated_multi_exam_golden_is_deterministic_and_typed() -> None:
    early = CombinedExam(
        SnapshotRef("session-a", 1, "generation-a", "b" * 64),
        SessionRecord(
            1,
            "session-a",
            1,
            SessionState.GRADED,
            "이전",
            2025,
            ExamTerm.FIRST,
            STAMP,
            STAMP,
            STAMP,
        ),
        RosterSnapshot(
            1,
            RosterSnapshotKind.WORKBOOK,
            "roster.xlsx",
            SHA,
            "명단",
            "v1",
            (_entry(1, "00000001", "=Alice", RosterRowStatus.NORMAL, ()),),
            (),
        ),
        (_response("early", "00000001"),),
        _scores(("early", Decimal("5"))),
        (),
    )
    current = CombinedExam(
        SnapshotRef("session-z", 2, "generation-z", "c" * 64),
        SessionRecord(
            1,
            "session-z",
            2,
            SessionState.GRADED,
            "+현재",
            2026,
            ExamTerm.SECOND,
            STAMP,
            STAMP,
            STAMP,
        ),
        RosterSnapshot(
            1,
            RosterSnapshotKind.WORKBOOK,
            "roster.xlsx",
            SHA,
            "명단",
            "v1",
            (
                _entry(1, "00000001", "=Alice", RosterRowStatus.NORMAL, ()),
                _entry(2, "00000002", "Roster only", RosterRowStatus.NORMAL, ()),
                _entry(
                    3,
                    "00000003",
                    "Kim",
                    RosterRowStatus.DUPLICATE_ID,
                    ("duplicate_id", "name_conflict"),
                ),
                _entry(
                    4,
                    "00000003",
                    "Lee",
                    RosterRowStatus.DUPLICATE_ID,
                    ("duplicate_id", "name_conflict"),
                ),
                _entry(5, None, "@broken", RosterRowStatus.INVALID_ID, ("invalid_id",)),
            ),
            (),
        ),
        (
            _response("ok", "00000001"),
            _response("unregistered", "00000009"),
            _response("duplicate-a", "00000003", duplicate=True, corrected=True),
            _response("duplicate-b", "00000003", duplicate=True),
            _response("unreadable", None, source="=unreadable.png"),
        ),
        _scores(
            ("ok", Decimal("7.5")),
            ("unregistered", None),
            ("duplicate-a", None),
            ("duplicate-b", None),
            ("unreadable", None),
        ),
        (_failure(),),
    )
    book = load_workbook(BytesIO(build_combined_workbook((current, early), STAMP)), data_only=False)

    assert [row[1].value for row in book["시험목록"].iter_rows(min_row=2)] == [
        "session-a",
        "session-z",
    ]
    assert book["시험목록"]["D3"].value == "'+현재"
    assert [row[1].value for row in book["통합성적"].iter_rows(min_row=2)] == [
        "00000001",
        "00000002",
        "00000003",
        "00000009",
    ]
    alice = list(book["통합성적"].iter_rows(min_row=2, max_row=2))[0]
    assert [cell.value for cell in alice] == [
        1,
        "00000001",
        "'=Alice",
        "정상",
        Decimal("5"),
        Decimal("10"),
        "응시",
        Decimal("7.5"),
        Decimal("10"),
        "응시",
        2,
        Decimal("12.5"),
        Decimal("20"),
        Decimal("62.5"),
    ]
    assert alice[-1].data_type == "n"
    roster_only = list(book["통합성적"].iter_rows(min_row=3, max_row=3))[0]
    assert [cell.value for cell in roster_only[4:10]] == [
        None,
        Decimal("10"),
        "미응시",
        None,
        Decimal("10"),
        "미응시",
    ]

    reviews = [tuple(cell.value for cell in row) for row in book["검토필요"].iter_rows(min_row=2)]
    assert ("session-z", 2, None, None, None, "명단학번오류", "'@broken", "명단학번오류") in reviews
    assert (
        "session-z",
        2,
        "unreadable",
        "'=unreadable.png",
        None,
        "학번미인식",
        None,
        "학번미인식",
    ) in reviews
    assert (
        "session-z",
        2,
        "unregistered",
        "unregistered.png",
        "00000009",
        "명단미등록",
        None,
        "명단미등록",
    ) in reviews
    assert (
        "session-z",
        2,
        "duplicate-a",
        "duplicate-a.png",
        "00000003",
        "학번중복",
        None,
        "학번중복,수동확인",
    ) in reviews
    assert (
        "session-z",
        2,
        "failed",
        "failed.png",
        None,
        "페이지처리실패",
        None,
        "페이지처리실패",
    ) in reviews
    assert all(
        cell.data_type != "f"
        for sheet in book.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )
    properties = {item.name: item.value for item in book.custom_doc_props}
    assert properties["schema"] == "1"
    assert properties["generation_time"] == STAMP
    assert (
        properties["provenance_tuple_digest"]
        == sha256(
            f"session-a:1:generation-a:{'b' * 64}\nsession-z:2:generation-z:{'c' * 64}".encode()
        ).hexdigest()
    )


def test_combined_book_golden_has_exact_formula_free_sheet_contract() -> None:
    book = load_workbook(BytesIO(build_combined_workbook((_exam(),), STAMP)), data_only=False)
    assert book.sheetnames == ["시험목록", "통합성적", "검토필요"]
    assert [cell.value for cell in book["시험목록"][1]] == [
        "시험순서",
        "시험ID",
        "리비전",
        "시험명",
        "연도",
        "학기",
        "채점일시",
        "만점",
        "ManifestSHA256",
    ]
    assert [cell.value for cell in book["통합성적"][1]] == [
        "연번",
        "학번",
        "이름",
        "이름상태",
        "1_점수",
        "1_만점",
        "1_응시상태",
        "응시시험수",
        "취득점수합계",
        "만점합계",
        "환산평균(%)",
    ]
    assert [cell.value for cell in book["검토필요"][1]] == [
        "시험ID",
        "리비전",
        "WorkItemID",
        "원본파일명",
        "유효학번",
        "학번상태",
        "이름후보",
        "사유",
    ]
    assert all(
        cell.data_type != "f"
        for sheet in book.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )


class _Lease:
    def __init__(self, payload: dict[str, object]) -> None:
        self.snapshot_ref = SnapshotRef("session-1", 1, "generation-1", SHA)
        self.manifest = type("Manifest", (), {"state": SessionState.RECOGNIZED})()
        self._payload = payload
        self.closed = False

    def open_allowlisted(self, path: str):
        assert path == "semantic_inputs.json"
        return Ok(BytesIO(json.dumps(self._payload).encode()))

    def close(self):
        self.closed = True
        return Ok(None)


class _Coordinator:
    def __init__(self, lease: _Lease) -> None:
        self.lease = lease

    def open_committed_snapshot(self, _request):
        return Ok(self.lease)


def test_combined_report_rejects_corrupt_score_semantics_before_publish(tmp_path: Path) -> None:
    session = _exam().session.to_dict()
    session["state"] = SessionState.GRADED.value
    session["graded_at"] = STAMP
    response = _response("response-1", "00000001").to_dict()
    valid_row = {"work_item_id": "response-1", "score": "5", "rank": 1}
    valid_statistics = {
        "participant_count": 1,
        "average_score": "5",
        "highest_score": "5",
        "lowest_score": "5",
    }
    empty_statistics = {
        "participant_count": 0,
        "average_score": None,
        "highest_score": None,
        "lowest_score": None,
    }

    def payload(scores: object) -> dict[str, object]:
        return {
            "combined": {
                "session": session,
                "roster": _exam().roster.to_dict(),
                "responses": [response],
                "scores": scores,
                "failures": [],
            }
        }

    invalid_scores = (
        {"maximum_score": "10", "rows": [valid_row, valid_row], "statistics": valid_statistics},
        {
            "maximum_score": "10",
            "rows": [{"work_item_id": "unknown", "score": "5", "rank": 1}],
            "statistics": valid_statistics,
        },
        {"maximum_score": "10", "rows": [], "statistics": empty_statistics},
        {
            "maximum_score": "10",
            "rows": [{"work_item_id": "response-1", "score": "5", "rank": True}],
            "statistics": valid_statistics,
        },
        {
            "maximum_score": "10",
            "rows": [{"work_item_id": "response-1", "score": "5", "rank": 0}],
            "statistics": valid_statistics,
        },
        {
            "maximum_score": "10",
            "rows": [{"work_item_id": "response-1", "score": "NaN", "rank": 1}],
            "statistics": valid_statistics,
        },
        {
            "maximum_score": "10",
            "rows": [{"work_item_id": "response-1", "score": "-1", "rank": 1}],
            "statistics": valid_statistics,
        },
        {
            "maximum_score": "10",
            "rows": [{"work_item_id": "response-1", "score": "11", "rank": 1}],
            "statistics": valid_statistics,
        },
        {
            "maximum_score": "10",
            "rows": [valid_row],
            "statistics": {**valid_statistics, "participant_count": True},
        },
        {
            "maximum_score": "10",
            "rows": [valid_row],
            "statistics": {**valid_statistics, "average_score": "4"},
        },
        None,
    )

    for index, scores in enumerate(invalid_scores):
        destination = tmp_path / f"corrupt-{index}.xlsx"
        lease = _Lease(payload(scores))
        result = DashboardUseCase(_Coordinator(lease)).build_combined_report(
            CombinedReportRequest(
                ("session-1",), False, str(destination), CollisionPolicy.ERROR, f"report-{index}"
            )
        )
        assert isinstance(result, Err)
        assert result.errors[0].code == "COMBINED_SEMANTICS_INVALID"
        assert not destination.exists()
        assert lease.closed


def test_combined_report_accepts_explicit_ungraded_and_manual_scores(tmp_path: Path) -> None:
    ungraded = _exam()
    manual_session = ungraded.session.to_dict()
    manual_session["state"] = SessionState.GRADED.value
    manual_session["graded_at"] = STAMP
    manual_scores = {
        "maximum_score": "10",
        "rows": [{"work_item_id": "response-1", "score": None, "rank": None}],
        "statistics": {
            "participant_count": 0,
            "average_score": None,
            "highest_score": None,
            "lowest_score": None,
        },
    }
    payloads = (
        {
            "combined": {
                "session": ungraded.session.to_dict(),
                "roster": ungraded.roster.to_dict(),
                "responses": [],
                "scores": None,
                "failures": [],
            }
        },
        {
            "combined": {
                "session": manual_session,
                "roster": ungraded.roster.to_dict(),
                "responses": [_response("response-1", "00000001").to_dict()],
                "scores": manual_scores,
                "failures": [],
            }
        },
    )
    for index, payload in enumerate(payloads):
        destination = tmp_path / f"valid-{index}.xlsx"
        result = DashboardUseCase(_Coordinator(_Lease(payload))).build_combined_report(
            CombinedReportRequest(
                ("session-1",), False, str(destination), CollisionPolicy.ERROR, f"valid-{index}"
            )
        )
        assert isinstance(result, Ok)
        assert destination.is_file()


def test_combined_report_holds_lease_through_atomic_publish(tmp_path: Path) -> None:
    exam = _exam()
    payload = {
        "combined": {
            "session": exam.session.to_dict(),
            "roster": exam.roster.to_dict(),
            "responses": [],
            "scores": None,
            "failures": [],
        }
    }
    lease = _Lease(payload)
    destination = tmp_path / "combined.xlsx"
    result = DashboardUseCase(_Coordinator(lease)).build_combined_report(
        CombinedReportRequest(
            ("session-1",), False, str(destination), CollisionPolicy.ERROR, "report-1"
        )
    )
    assert isinstance(result, Ok)
    assert destination.is_file()
    assert lease.closed


def test_combined_report_failed_publish_cleans_temporary_and_releases_every_lease(
    tmp_path: Path, monkeypatch
) -> None:
    exam = _exam()
    payload = {
        "combined": {
            "session": exam.session.to_dict(),
            "roster": exam.roster.to_dict(),
            "responses": [],
            "scores": None,
            "failures": [],
        }
    }
    first, second = _Lease(payload), _Lease(payload)

    class TwoLeaseCoordinator:
        def __init__(self) -> None:
            self._leases = iter((first, second))

        def open_committed_snapshot(self, _request):
            return Ok(next(self._leases))

    destination = tmp_path / "combined.xlsx"
    temporary_paths: list[Path] = []

    def fail_atomic_create(temporary: Path, published: Path) -> None:
        assert published == destination
        assert temporary.is_file()
        temporary_paths.append(temporary)
        raise OSError("simulated publish failure")

    monkeypatch.setattr("omr_grader.application.dashboard_use_case.os.link", fail_atomic_create)
    result = DashboardUseCase(TwoLeaseCoordinator()).build_combined_report(
        CombinedReportRequest(
            ("session-1", "session-2"),
            False,
            str(destination),
            CollisionPolicy.ERROR,
            "report-1",
        )
    )

    assert isinstance(result, Err)
    assert result.errors[0].code == "REPORT_PUBLISH_FAILED"
    assert result.errors[0].cause_type == "OSError"
    assert not destination.exists()
    assert len(temporary_paths) == 1
    assert not temporary_paths[0].exists()
    assert tuple(tmp_path.glob(".combined.xlsx.*.tmp")) == ()
    assert first.closed and second.closed

def test_publish_returns_cleanup_failure_when_published_temporary_remains(
    tmp_path: Path, monkeypatch
) -> None:
    destination = tmp_path / "combined.xlsx"
    payload = b"combined-report"
    original_unlink = Path.unlink
    temporary_paths: list[Path] = []

    def fail_temporary_cleanup(path: Path, *, missing_ok: bool = False) -> None:
        if path.name.startswith(".combined.xlsx.") and path.suffix == ".tmp":
            temporary_paths.append(path)
            raise OSError("simulated cleanup failure")
        original_unlink(path, missing_ok=missing_ok)

    monkeypatch.setattr(Path, "unlink", fail_temporary_cleanup)
    result = _publish(destination, payload, CollisionPolicy.ERROR)

    assert isinstance(result, Err)
    assert [error.code for error in result.errors] == ["REPORT_TEMPORARY_CLEANUP_FAILED"]
    assert result.errors[0].cause_type == "OSError"
    assert destination.read_bytes() == payload
    assert len(temporary_paths) == 1 and temporary_paths[0].is_file()
    original_unlink(temporary_paths[0])


def test_publish_preserves_primary_and_cleanup_failures_in_order(
    tmp_path: Path, monkeypatch
) -> None:
    destination = tmp_path / "combined.xlsx"
    original_unlink = Path.unlink
    temporary_paths: list[Path] = []

    def fail_atomic_create(temporary: Path, published: Path) -> None:
        assert published == destination
        assert temporary.is_file()
        raise OSError("simulated publish failure")

    def fail_temporary_cleanup(path: Path, *, missing_ok: bool = False) -> None:
        if path.name.startswith(".combined.xlsx.") and path.suffix == ".tmp":
            temporary_paths.append(path)
            raise OSError("simulated cleanup failure")
        original_unlink(path, missing_ok=missing_ok)

    monkeypatch.setattr("omr_grader.application.dashboard_use_case.os.link", fail_atomic_create)
    monkeypatch.setattr(Path, "unlink", fail_temporary_cleanup)
    result = _publish(destination, b"combined-report", CollisionPolicy.ERROR)

    assert isinstance(result, Err)
    assert [error.code for error in result.errors] == [
        "REPORT_PUBLISH_FAILED",
        "REPORT_TEMPORARY_CLEANUP_FAILED",
    ]
    assert all(error.cause_type == "OSError" for error in result.errors)
    assert not destination.exists()
    assert len(temporary_paths) == 1 and temporary_paths[0].is_file()
    original_unlink(temporary_paths[0])


def test_combined_report_collision_and_concurrent_publish_are_atomic(tmp_path: Path) -> None:
    exam = _exam()
    payload = {
        "combined": {
            "session": exam.session.to_dict(),
            "roster": exam.roster.to_dict(),
            "responses": [],
            "scores": None,
            "failures": [],
        }
    }
    destination = tmp_path / "combined.xlsx"

    def publish() -> object:
        return DashboardUseCase(_Coordinator(_Lease(payload))).build_combined_report(
            CombinedReportRequest(
                ("session-1",), False, str(destination), CollisionPolicy.ERROR, "report-1"
            )
        )

    with ThreadPoolExecutor(max_workers=2) as pool:
        results = tuple(pool.map(lambda _: publish(), range(2)))
    assert sum(isinstance(result, Ok) for result in results) == 1
    assert sum(isinstance(result, Err) for result in results) == 1
    assert destination.is_file()

    replaced = DashboardUseCase(_Coordinator(_Lease(payload))).build_combined_report(
        CombinedReportRequest(
            ("session-1",), False, str(destination), CollisionPolicy.REPLACE, "report-2"
        )
    )
    assert isinstance(replaced, Ok)


def test_combined_report_releases_all_leases_when_cleanup_raises(tmp_path: Path) -> None:
    class ExplodingLease(_Lease):
        def close(self):
            self.closed = True
            raise OSError("close failed")

    class TwoLeaseCoordinator:
        def __init__(self, first: _Lease, second: _Lease) -> None:
            self._leases = iter((first, second))

        def open_committed_snapshot(self, _request):
            return Ok(next(self._leases))

    exam = _exam()
    payload = {
        "combined": {
            "session": exam.session.to_dict(),
            "roster": exam.roster.to_dict(),
            "responses": [],
            "scores": None,
            "failures": [],
        }
    }
    first, second = _Lease(payload), ExplodingLease(payload)
    result = DashboardUseCase(TwoLeaseCoordinator(first, second)).build_combined_report(
        CombinedReportRequest(
            ("session-1", "session-2"),
            False,
            str(tmp_path / "combined.xlsx"),
            CollisionPolicy.ERROR,
            "report-1",
        )
    )
    assert isinstance(result, Ok)
    assert [warning.code for warning in result.warnings] == ["COMBINED_LEASE_CLOSE_FAILED"]
    assert first.closed and second.closed
def test_combined_report_reports_returned_close_errors_as_warnings(tmp_path: Path) -> None:
    class ReturnedErrorLease(_Lease):
        def close(self):
            self.closed = True
            return Err(
                (
                    ErrorInfo(
                        "SESSION_LEASE_CLOSE_FAILED",
                        "error.session_lease_close_failed",
                        context={"reason": "close returned error"},
                    ),
                )
            )

    exam = _exam()
    payload = {
        "combined": {
            "session": exam.session.to_dict(),
            "roster": exam.roster.to_dict(),
            "responses": [],
            "scores": None,
            "failures": [],
        }
    }
    lease = ReturnedErrorLease(payload)
    result = DashboardUseCase(_Coordinator(lease)).build_combined_report(
        CombinedReportRequest(
            ("session-1",),
            False,
            str(tmp_path / "combined.xlsx"),
            CollisionPolicy.ERROR,
            "report-1",
        )
    )

    assert isinstance(result, Ok)
    assert [warning.code for warning in result.warnings] == ["SESSION_LEASE_CLOSE_FAILED"]
    assert lease.closed


def test_combined_report_appends_returned_close_errors_to_primary_failure(
    tmp_path: Path, monkeypatch
) -> None:
    class ReturnedErrorLease(_Lease):
        def close(self):
            self.closed = True
            return Err(
                (
                    ErrorInfo(
                        "SESSION_LEASE_CLOSE_FAILED",
                        "error.session_lease_close_failed",
                        context={"reason": "close returned error"},
                    ),
                )
            )

    class TwoLeaseCoordinator:
        def __init__(self, first: _Lease, second: _Lease) -> None:
            self._leases = iter((first, second))

        def open_committed_snapshot(self, _request):
            return Ok(next(self._leases))

    def fail_atomic_create(_temporary: Path, _published: Path) -> None:
        raise OSError("simulated publish failure")

    exam = _exam()
    payload = {
        "combined": {
            "session": exam.session.to_dict(),
            "roster": exam.roster.to_dict(),
            "responses": [],
            "scores": None,
            "failures": [],
        }
    }
    first, second = ReturnedErrorLease(payload), ReturnedErrorLease(payload)
    monkeypatch.setattr("omr_grader.application.dashboard_use_case.os.link", fail_atomic_create)
    result = DashboardUseCase(TwoLeaseCoordinator(first, second)).build_combined_report(
        CombinedReportRequest(
            ("session-1", "session-2"),
            False,
            str(tmp_path / "combined.xlsx"),
            CollisionPolicy.ERROR,
            "report-1",
        )
    )

    assert isinstance(result, Err)
    assert [error.code for error in result.errors] == [
        "REPORT_PUBLISH_FAILED",
        "SESSION_LEASE_CLOSE_FAILED",
        "SESSION_LEASE_CLOSE_FAILED",
    ]
    assert first.closed and second.closed
