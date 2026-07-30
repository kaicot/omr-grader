"""Deterministic real-data QA for compact OMR session retention."""

from __future__ import annotations

import argparse
import glob
import json
import shutil
from pathlib import Path

import cv2
import numpy as np
from openpyxl import load_workbook

from omr_grader.application.answer_key_use_case import AnswerKeyWorkbookUseCase
from omr_grader.application.dto import RegradeCommand, ScanCommand, ScanSource
from omr_grader.application.grading_use_case import GradingUseCase
from omr_grader.domain.enums import ExamTerm
from omr_grader.domain.errors import Err
from omr_grader.infrastructure.grading_runtime import CommittedGradingSnapshotReader
from omr_grader.infrastructure.paths import ManagedPaths
from omr_grader.infrastructure.profile_store import ProfileStore
from omr_grader.infrastructure.scan_runtime import ScanRuntime, bind_scan_runtime
from omr_grader.infrastructure.session_store import SessionStore


def _physical_bytes(root: Path) -> int:
    seen: set[tuple[int, int]] = set()
    total = 0
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        stat = path.stat()
        identity = (stat.st_dev, stat.st_ino)
        if identity in seen:
            continue
        seen.add(identity)
        total += stat.st_size
    return total


def _require_ok(result: object, operation: str):
    if isinstance(result, Err):
        errors = ", ".join(f"{item.code}:{item.context}" for item in result.errors)
        raise RuntimeError(f"{operation} failed: {errors}")
    return result


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--profile", required=True)
    parser.add_argument("--source-glob", required=True)
    parser.add_argument("--answer-key", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    output = Path(args.output).resolve()
    runtime_root = output / "runtime"
    if runtime_root.exists():
        shutil.rmtree(runtime_root)
    runtime_root.mkdir(parents=True)
    paths = ManagedPaths.from_root(runtime_root)
    paths.profiles_dir.mkdir()
    paths.data_dir.mkdir()
    paths.logs_dir.mkdir()
    profile_source = Path(args.profile).resolve()
    profile_target = paths.profiles_dir / profile_source.name
    shutil.copy2(profile_source, profile_target)
    sources = tuple(str(Path(path).resolve()) for path in sorted(glob.glob(args.source_glob)))
    if not sources:
        raise RuntimeError("source glob matched no files")
    answer_key = Path(args.answer_key).resolve()
    workbook = load_workbook(answer_key, read_only=True, data_only=False)
    try:
        answer_sheet = workbook.sheetnames[0]
    finally:
        workbook.close()

    store = SessionStore(paths)
    scan_runtime = ScanRuntime(ProfileStore(paths), store, app_version="qa")
    scan = bind_scan_runtime(scan_runtime)
    scanned = _require_ok(
        scan.run_scan(
            ScanCommand(
                "qa-minimal-storage",
                "qa-scan",
                0,
                "용량최적화_QA",
                2026,
                ExamTerm.SECOND,
                profile_target.name,
                None,
                ScanSource(sources),
                5,
                False,
            )
        ),
        "scan",
    )
    grader = GradingUseCase(
        CommittedGradingSnapshotReader(store),
        AnswerKeyWorkbookUseCase(),
        store,
    )
    first = _require_ok(
        grader.regrade(
            RegradeCommand(
                scanned.value.session_id,
                1,
                str(answer_key),
                answer_sheet,
                "qa-grade-1",
            )
        ),
        "first grade",
    )
    session = next(
        path
        for path in paths.data_dir.iterdir()
        if path.is_dir() and (path / "CURRENT.json").is_file()
    )
    first_pointer = json.loads((session / "CURRENT.json").read_text(encoding="utf-8"))
    first_generation = session / first_pointer["generation_relpath"]
    first_reviews = tuple(sorted((first_generation / "02_채점결과_이미지").glob("*.jpg")))
    if first_reviews:
        shutil.copy2(first_reviews[0], output / "representative-review.jpg")

    second = _require_ok(
        grader.regrade(
            RegradeCommand(
                scanned.value.session_id,
                first.value.revision,
                str(answer_key),
                answer_sheet,
                "qa-grade-2",
            )
        ),
        "second grade",
    )
    pointer = json.loads((session / "CURRENT.json").read_text(encoding="utf-8"))
    generation = session / pointer["generation_relpath"]
    generations = tuple(path for path in (session / "generations").iterdir() if path.is_dir())
    reviews = tuple(sorted((generation / "02_채점결과_이미지").glob("*.jpg")))
    review_details = []
    for review in reviews:
        payload = review.read_bytes()
        raster = cv2.imdecode(np.frombuffer(payload, dtype=np.uint8), cv2.IMREAD_COLOR)
        review_details.append(
            {
                "name": review.name,
                "bytes": len(payload),
                "jpeg": payload.startswith(b"\xff\xd8") and payload.endswith(b"\xff\xd9"),
                "max_dimension": None if raster is None else max(raster.shape[:2]),
            }
        )
    forbidden = (
        "recognition",
        "evidence",
        "automatic",
        "좌표데이터",
        "01_인식결과_이미지",
        "corrections",
        "correction_events.json",
        "correction_history.json",
        "projection_request.json",
        "artifacts",
    )
    checks = {
        "revision_is_three": second.value.revision == 3 and pointer["revision"] == 3,
        "one_generation": len(generations) == 1,
        "source_scans_retained": len(tuple((generation / "images").glob("*.png"))) == 10,
        "response_workbook_retained": len(tuple(generation.glob("01_ocr_*_응답결과.xlsx"))) == 1,
        "answer_key_retained": len(tuple(generation.glob("정답표_*.xlsx"))) == 1,
        "score_workbook_retained": len(tuple(generation.glob("02_score_*_채점결과.xlsx"))) == 1,
        "review_jpegs_retained": len(reviews) == 10,
        "review_jpegs_bounded": bool(review_details)
        and all(item["jpeg"] and item["max_dimension"] <= 2000 for item in review_details),
        "forbidden_payloads_absent": not any((generation / name).exists() for name in forbidden),
    }
    actual_bytes = _physical_bytes(session)
    checks["physical_bytes_below_120mb"] = actual_bytes <= 120 * 1024 * 1024
    report = {
        "passed": all(checks.values()),
        "checks": checks,
        "session_path": str(session),
        "generation_path": str(generation),
        "generation_count": len(generations),
        "revision": pointer["revision"],
        "physical_bytes": actual_bytes,
        "legacy_reference_bytes": 351 * 1024 * 1024,
        "saved_percent": round(100 * (1 - actual_bytes / (351 * 1024 * 1024)), 1),
        "review_images": review_details,
        "first_review_count": len(first_reviews),
        "required_files": sorted(
            path.relative_to(generation).as_posix()
            for path in generation.rglob("*")
            if path.is_file()
        ),
    }
    output.mkdir(parents=True, exist_ok=True)
    report_path = output / "report.json"
    report_path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(
        "PASS" if report["passed"] else "FAIL",
        f"bytes={actual_bytes}",
        f"generations={len(generations)}",
        f"reviews={len(reviews)}",
        f"report={report_path}",
    )
    return 0 if report["passed"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
