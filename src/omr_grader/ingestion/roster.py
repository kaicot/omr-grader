"""Fail-closed import policy for the optional student roster workbook."""

from __future__ import annotations

from hashlib import sha256
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZIP_STORED, BadZipFile, ZipFile, ZipInfo

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from omr_grader.domain.enums import RosterRowStatus, RosterSnapshotKind
from omr_grader.domain.errors import Err, ErrorInfo, Ok, Result
from omr_grader.domain.models import SCHEMA_VERSION, RosterEntry, RosterSnapshot
from omr_grader.workbooks.schemas import validate_opc_relationships

NORMALIZATION_VERSION = "roster-xlsx-v1"
ROSTER_HEADERS = ("연번", "학번", "이름")
MAX_PACKAGE_BYTES = 16 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 64 * 1024 * 1024
MAX_PACKAGE_MEMBERS = 2_000
MAX_ROWS = 100_000
MAX_RELATIONSHIP_BYTES = 1024 * 1024


def _error(code: str, field: str | None = None) -> ErrorInfo:
    return ErrorInfo(code, f"error.{code.lower()}", field)


def _has_external_relationship(package: ZipFile, member: ZipInfo) -> Result[bool]:
    if member.file_size > MAX_RELATIONSHIP_BYTES:
        return Err((_error("XLSX_PACKAGE_QUOTA", "path"),))
    payload = package.read(member)
    if len(payload) > MAX_RELATIONSHIP_BYTES:
        return Err((_error("XLSX_PACKAGE_QUOTA", "path"),))
    return validate_opc_relationships(payload)


def _read_package(path: str) -> Result[tuple[bytes, str]]:
    try:
        data = Path(path).read_bytes()
    except OSError:
        return Err((_error("XLSX_READ_FAILED", "path"),))
    if not data or len(data) > MAX_PACKAGE_BYTES:
        return Err((_error("XLSX_PACKAGE_QUOTA", "path"),))
    try:
        with ZipFile(BytesIO(data)) as package:
            members = package.infolist()
            if any(
                item.flag_bits & 0x1 or item.compress_type not in (ZIP_STORED, ZIP_DEFLATED)
                for item in members
            ):
                return Err((_error("XLSX_INVALID_PACKAGE", "path"),))
            if (
                len(members) > MAX_PACKAGE_MEMBERS
                or sum(item.file_size for item in members) > MAX_UNCOMPRESSED_BYTES
            ):
                return Err((_error("XLSX_PACKAGE_QUOTA", "path"),))
            names = {item.filename.lower() for item in members}
            if any(
                b"<f" in package.read(item).lower()
                for item in members
                if item.filename.lower().startswith("xl/worksheets/")
            ):
                return Err((_error("XLSX_FORMULA_FORBIDDEN", "path"),))
            for item in members:
                if not item.filename.lower().endswith(".rels"):
                    continue
                external_relationship = _has_external_relationship(package, item)
                if isinstance(external_relationship, Err):
                    return external_relationship
                if external_relationship.value:
                    return Err((_error("XLSX_FORBIDDEN_FEATURE", "path"),))
    except (BadZipFile, NotImplementedError, RuntimeError):
        return Err((_error("XLSX_INVALID_PACKAGE", "path"),))
    forbidden = (
        "vbaProject.bin".lower(),
        "encryptioninfo",
        "encryptedpackage",
        "externalLinks/".lower(),
        "connections",
        "oleobject",
        "dde",
        "activex",
    )
    if any(any(token in name for token in forbidden) for name in names):
        return Err((_error("XLSX_FORBIDDEN_FEATURE", "path"),))
    return Ok((data, sha256(data).hexdigest()))


def _cell_text(cell: object, field: str, *, allow_blank: bool = False) -> Result[str]:
    value = getattr(cell, "value", None)
    data_type = getattr(cell, "data_type", None)
    if data_type == "f":
        return Err((_error("XLSX_FORMULA_FORBIDDEN", field),))
    if value is None and allow_blank:
        return Ok("")
    if data_type not in {"s", "inlineStr"} or type(value) is not str:
        return Err((_error("XLSX_CELL_TYPE", field),))
    return Ok(value)


def import_roster(path: str, sheet_name: str) -> Result[RosterSnapshot]:
    """Read a roster as immutable source data; invalid workbooks are never normalized."""
    package = _read_package(path)
    if isinstance(package, Err):
        return package
    data, digest = package.value
    workbook = None
    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=False, keep_links=False)
    except (BadZipFile, InvalidFileException, OSError):
        return Err((_error("XLSX_INVALID_WORKBOOK", "path"),))
    try:
        if sheet_name not in workbook.sheetnames:
            return Err((_error("XLSX_SHEET_NOT_FOUND", "sheet_name"),))
        sheet = workbook[sheet_name]
        if sheet.max_row > MAX_ROWS or sheet.max_column != len(ROSTER_HEADERS):
            return Err((_error("XLSX_DIMENSION_QUOTA", "sheet_name"),))
        header_results = tuple(
            _cell_text(cell, f"A1:{index}")
            for index, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1)), 1)
        )
        header: list[str] = []
        for result in header_results:
            if isinstance(result, Err):
                return Err((_error("XLSX_HEADERS_INVALID", "header"),))
            header.append(result.value)
        if tuple(header) != ROSTER_HEADERS:
            return Err((_error("XLSX_HEADERS_INVALID", "header"),))
        raw_rows: list[tuple[int, int, str, str]] = []
        for ordinal, row in enumerate(sheet.iter_rows(min_row=2, max_col=3), 0):
            row_number = ordinal + 2
            if any(cell.data_type == "f" for cell in row):
                return Err((_error("XLSX_FORMULA_FORBIDDEN", f"row[{row_number}]"),))
            if all(cell.value is None for cell in row):
                continue
            values = (
                _cell_text(row[0], f"A{row_number}"),
                _cell_text(row[1], f"B{row_number}", allow_blank=True),
                _cell_text(row[2], f"C{row_number}", allow_blank=True),
            )
            row_values: list[str] = []
            for result in values:
                if isinstance(result, Err):
                    return Err((_error("XLSX_CELL_TYPE", f"row[{row_number}]"),))
                row_values.append(result.value)
            serial, parsed_student_id, name = row_values
            if not serial and not parsed_student_id and not name:
                continue
            raw_rows.append((ordinal, row_number, parsed_student_id, name))
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except OSError:
                pass

    groups: dict[str, list[tuple[int, int, str, str]]] = {}
    for record in raw_rows:
        if len(record[2]) == 8 and record[2].isdigit():
            groups.setdefault(record[2], []).append(record)
    entries: list[RosterEntry] = []
    for ordinal, row_number, raw_id, name in raw_rows:
        effective_student_id = (
            raw_id if len(raw_id) == 8 and raw_id.isascii() and raw_id.isdigit() else None
        )
        issues: list[str] = []
        if effective_student_id is None:
            issues.append(RosterRowStatus.INVALID_ID.value)
        else:
            group = groups[effective_student_id]
            if len(group) > 1:
                issues.append(RosterRowStatus.DUPLICATE_ID.value)
            if len({item[3] for item in group}) > 1:
                issues.append(RosterRowStatus.NAME_CONFLICT.value)
        ordered = tuple(sorted(issues))
        status = next(
            (
                item
                for item in (
                    RosterRowStatus.INVALID_ID,
                    RosterRowStatus.DUPLICATE_ID,
                    RosterRowStatus.NAME_CONFLICT,
                )
                if item.value in ordered
            ),
            RosterRowStatus.NORMAL,
        )
        row_id = sha256(f"{digest}:{sheet_name}:{row_number}".encode()).hexdigest()
        entries.append(
            RosterEntry(
                row_id,
                row_number,
                ordinal,
                raw_id,
                effective_student_id,
                name,
                status,
                ordered,
            )
        )
    return Ok(
        RosterSnapshot(
            SCHEMA_VERSION,
            RosterSnapshotKind.WORKBOOK,
            Path(path).name,
            digest,
            sheet_name,
            NORMALIZATION_VERSION,
            tuple(entries),
            (),
        )
    )


validate_roster_workbook = import_roster
