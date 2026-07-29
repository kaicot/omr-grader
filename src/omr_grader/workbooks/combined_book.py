"""Formula-free, deterministic three-sheet combined report projection."""

from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from io import BytesIO
from typing import Any, cast

from openpyxl import Workbook
from openpyxl.packaging.custom import StringProperty
from openpyxl.worksheet.worksheet import Worksheet

from omr_grader.application.dto import ScoreResult, ScoreSet, ScoreStatistics, SnapshotRef
from omr_grader.domain.enums import ExamTerm, RosterRowStatus, SessionState, StudentIdStatus
from omr_grader.domain.models import EffectiveResponse, PageFailure, RosterSnapshot, SessionRecord

type ReviewRow = tuple[str, int, str | None, str, str | None, str, str, list[str]]
type RosterRecord = tuple[str, RosterRowStatus, tuple[str, ...]]


_REASONS = (
    "학번미인식",
    "학번중복",
    "명단미등록",
    "명단학번오류",
    "명단중복",
    "이름불일치",
    "수동확인",
    "페이지처리실패",
    "미채점",
)
_REASON_ORDER = {reason: index for index, reason in enumerate(_REASONS)}
_TERM_ORDER = {
    ExamTerm.FIRST: 0,
    ExamTerm.SECOND: 1,
    ExamTerm.SUMMER: 2,
    ExamTerm.WINTER: 3,
    ExamTerm.OTHER: 4,
    ExamTerm.UNSPECIFIED: 5,
}


@dataclass(frozen=True, slots=True)
class CombinedExam:
    snapshot: SnapshotRef
    session: SessionRecord
    roster: RosterSnapshot
    responses: tuple[EffectiveResponse, ...]
    scores: ScoreSet | None
    failures: tuple[PageFailure, ...]

    @property
    def maximum(self) -> Decimal | None:
        return None if self.scores is None else self.scores.maximum_score


def _text(value: object) -> str:
    text = "" if value is None else str(value)
    return "'" + text if text.startswith(("=", "+", "-", "@")) else text


def _decimal(value: Decimal) -> Decimal:
    if not value.is_finite():
        raise ValueError("score must be finite")
    return value


def _exam_key(exam: CombinedExam) -> tuple[object, ...]:
    session = exam.session
    return (
        session.exam_year is None,
        session.exam_year or 0,
        _TERM_ORDER[session.exam_term],
        session.graded_at is None,
        session.graded_at or "",
        session.session_id,
    )


def _score_decimal(value: object, name: str) -> Decimal:
    if isinstance(value, bool):
        raise ValueError(f"{name} must be a finite nonnegative decimal")
    try:
        parsed = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError) as error:
        raise ValueError(f"{name} must be a finite nonnegative decimal") from error
    if not parsed.is_finite() or parsed < 0:
        raise ValueError(f"{name} must be a finite nonnegative decimal")
    return parsed


def _score_set(value: object) -> ScoreSet | None:
    if value is None:
        return None
    if not isinstance(value, dict) or set(value) != {"maximum_score", "rows", "statistics"}:
        raise ValueError("combined scores are invalid")
    rows = value["rows"]
    statistics = value["statistics"]
    if (
        not isinstance(rows, list)
        or not isinstance(statistics, dict)
        or set(statistics)
        != {
            "participant_count",
            "average_score",
            "highest_score",
            "lowest_score",
        }
    ):
        raise ValueError("combined scores are invalid")

    parsed_rows: list[ScoreResult] = []
    for row in rows:
        if not isinstance(row, dict) or set(row) != {"work_item_id", "score", "rank"}:
            raise ValueError("combined scores are invalid")
        parsed_rows.append(
            ScoreResult(
                row["work_item_id"],
                None if row["score"] is None else _score_decimal(row["score"], "score"),
                row["rank"],
            )
        )

    def statistic(name: str) -> Decimal | None:
        raw = statistics[name]
        return None if raw is None else _score_decimal(raw, name)

    try:
        return ScoreSet(
            _score_decimal(value["maximum_score"], "maximum_score"),
            tuple(parsed_rows),
            ScoreStatistics(
                statistics["participant_count"],
                statistic("average_score"),
                statistic("highest_score"),
                statistic("lowest_score"),
            ),
        )
    except (TypeError, ValueError) as error:
        raise ValueError("combined scores are invalid") from error


def _validate_score_projection(
    session: SessionRecord,
    responses: tuple[EffectiveResponse, ...],
    scores: ScoreSet | None,
) -> None:
    """Enforce session scoring authority before projection.

    RECOGNIZED sessions are explicitly ungraded and have no ScoreSet. GRADED
    and FINALIZED sessions have one row per EffectiveResponse; a (None, None)
    row is an explicit manual-review score. PageFailure values have no
    EffectiveResponse and are therefore intentionally outside the score set.
    """
    if session.state is SessionState.RECOGNIZED:
        if scores is not None:
            raise ValueError("recognized sessions must be explicitly ungraded")
        return
    if scores is None:
        raise ValueError("graded sessions require explicit score rows")
    response_ids = {response.work_item_id for response in responses}
    score_ids = {row.work_item_id for row in scores.rows}
    if score_ids != response_ids or len(scores.rows) != len(response_ids):
        raise ValueError("score rows must match responses exactly")


def parse_combined_semantics(value: object, snapshot: SnapshotRef) -> CombinedExam:
    """Decode the explicit report payload; no workbook or excluded control file is consulted."""
    if (
        not isinstance(value, dict)
        or set(value) != {"combined"}
        or not isinstance(value["combined"], dict)
    ):
        raise ValueError("combined semantic payload is missing")
    payload = value["combined"]
    if set(payload) != {"session", "roster", "responses", "scores", "failures"}:
        raise ValueError("combined semantic payload fields are invalid")
    responses = payload["responses"]
    failures = payload["failures"]
    if not isinstance(responses, list) or not isinstance(failures, list):
        raise ValueError("combined response payload is invalid")
    session = SessionRecord.from_dict(payload["session"])
    roster = RosterSnapshot.from_dict(payload["roster"])
    parsed_responses = tuple(EffectiveResponse.from_dict(item) for item in responses)
    parsed_failures = tuple(PageFailure.from_dict(item) for item in failures)
    if len({item.work_item_id for item in parsed_responses}) != len(parsed_responses):
        raise ValueError("combined response work item IDs are not unique")
    if session.session_id != snapshot.session_id or session.revision != snapshot.revision:
        raise ValueError("combined session does not match frozen snapshot")
    if session.state not in (SessionState.RECOGNIZED, SessionState.GRADED, SessionState.FINALIZED):
        raise ValueError("combined session state is invalid")
    scores = _score_set(payload["scores"])
    _validate_score_projection(session, parsed_responses, scores)
    return CombinedExam(snapshot, session, roster, parsed_responses, scores, parsed_failures)


def _review(
    rows: list[ReviewRow],
    exam: CombinedExam,
    work_item_id: str | None,
    source: str,
    student_id: str | None,
    status: str,
    names: Iterable[str],
    reasons: Iterable[str],
) -> None:
    reason_set = tuple(sorted(set(reasons), key=_REASON_ORDER.__getitem__))
    if reason_set:
        rows.append(
            (
                exam.snapshot.session_id,
                exam.snapshot.revision,
                work_item_id,
                source,
                student_id,
                status,
                " | ".join(sorted(set(filter(None, names)))),
                list(reason_set),
            )
        )


def build_combined_workbook(exams: Iterable[CombinedExam], generated_at: str) -> bytes:
    ordered = tuple(sorted(exams, key=_exam_key))
    if not ordered:
        raise ValueError("at least one exam is required")
    for exam in ordered:
        _validate_score_projection(exam.session, exam.responses, exam.scores)
    score_rows_by_exam: tuple[dict[str, Decimal | None], ...] = tuple(
        {} if exam.scores is None else {row.work_item_id: row.score for row in exam.scores.rows}
        for exam in ordered
    )
    workbook = Workbook()
    exams_sheet = workbook.active
    if not isinstance(exams_sheet, Worksheet):
        raise RuntimeError("active workbook sheet is not a worksheet")
    exams_sheet.title = "시험목록"
    score_sheet = workbook.create_sheet("통합성적")
    review_sheet = workbook.create_sheet("검토필요")
    exams_sheet.append(
        [
            "시험순서",
            "시험ID",
            "리비전",
            "시험명",
            "연도",
            "학기",
            "채점일시",
            "만점",
            "ManifestSHA256",
        ]
    )
    for ordinal, exam in enumerate(ordered, 1):
        session = exam.session
        exams_sheet.append(
            [
                ordinal,
                session.session_id,
                exam.snapshot.revision,
                _text(session.exam_name),
                session.exam_year,
                session.exam_term.value,
                session.graded_at,
                exam.maximum,
                exam.snapshot.manifest_sha256,
            ]
        )

    roster_by_exam: list[dict[str, list[RosterRecord]]] = []
    review_rows: list[ReviewRow] = []
    ids: set[str] = set()
    for exam in ordered:
        roster: dict[str, list[RosterRecord]] = {}
        for roster_row in exam.roster.rows:
            if roster_row.student_id is None:
                _review(
                    review_rows,
                    exam,
                    None,
                    "",
                    None,
                    "명단학번오류",
                    (roster_row.name,),
                    ("명단학번오류",),
                )
                continue
            ids.add(roster_row.student_id)
            roster.setdefault(roster_row.student_id, []).append(
                (roster_row.name, roster_row.status, roster_row.issues)
            )
            roster_reasons: list[str] = []
            if roster_row.status is RosterRowStatus.DUPLICATE_ID:
                roster_reasons.append("명단중복")
            if roster_row.status is RosterRowStatus.NAME_CONFLICT:
                roster_reasons.append("이름불일치")
            _review(
                review_rows,
                exam,
                None,
                "",
                roster_row.student_id,
                "명단확인필요",
                (roster_row.name,),
                roster_reasons,
            )
        roster_by_exam.append(roster)
        seen: dict[str, list[EffectiveResponse]] = {}
        for response in exam.responses:
            if response.student_id is not None:
                ids.add(response.student_id)
                seen.setdefault(response.student_id, []).append(response)
        for response in exam.responses:
            response_reasons: list[str] = []
            status = "정상"
            if response.student_id is None:
                status, response_reasons = "학번미인식", ["학번미인식"]
            elif (
                len(seen[response.student_id]) > 1
                or response.student_id_status is StudentIdStatus.DUPLICATE
            ):
                status, response_reasons = "학번중복", ["학번중복"]
            elif response.student_id not in roster:
                status, response_reasons = "명단미등록", ["명단미등록"]
            if response.corrected_targets:
                response_reasons.append("수동확인")
            _review(
                review_rows,
                exam,
                response.work_item_id,
                response.source_label,
                response.student_id,
                status,
                (),
                response_reasons,
            )
        for failure in exam.failures:
            _review(
                review_rows,
                exam,
                failure.page_ref.work_item_id,
                failure.page_ref.source_label,
                None,
                "페이지처리실패",
                (),
                ("페이지처리실패",),
            )

    headers = ["연번", "학번", "이름", "이름상태"]
    for number in range(1, len(ordered) + 1):
        headers.extend((f"{number}_점수", f"{number}_만점", f"{number}_응시상태"))
    score_sheet.append(headers + ["응시시험수", "취득점수합계", "만점합계", "환산평균(%)"])
    for ordinal, student_id in enumerate(sorted(ids), 1):
        candidates: list[str] = [
            entry[0]
            for roster in roster_by_exam
            for entry in roster.get(student_id, ())
            if entry[0]
        ]
        name: str = candidates[0] if candidates else "미등록"
        name_status: str = (
            "이름불일치"
            if len(set(candidates)) > 1
            else ("명단미등록" if not candidates else "정상")
        )
        values: list[object] = [ordinal, student_id, _text(name), name_status]
        score_sum = Decimal("0")
        maximum_sum = Decimal("0")
        attended = 0
        for exam, score_by_work_item in zip(ordered, score_rows_by_exam, strict=True):
            matches = [item for item in exam.responses if item.student_id == student_id]
            duplicate = len(matches) > 1 or any(
                item.student_id_status is StudentIdStatus.DUPLICATE for item in matches
            )
            if duplicate:
                values.extend((None, None, "중복확인필요"))
                continue
            if not matches:
                values.extend((None, exam.maximum if exam.scores else None, "미응시"))
                continue
            response = matches[0]
            if exam.scores is None:
                values.extend((None, None, "미채점"))
                _review(
                    review_rows,
                    exam,
                    response.work_item_id,
                    response.source_label,
                    student_id,
                    "미채점",
                    (),
                    ("미채점",),
                )
                continue
            score = score_by_work_item[response.work_item_id]
            if score is None:
                values.extend((None, exam.maximum, "검토필요"))
                _review(
                    review_rows,
                    exam,
                    response.work_item_id,
                    response.source_label,
                    student_id,
                    "검토필요",
                    (),
                    ("수동확인",),
                )
                continue
            score = _decimal(score)
            maximum = _decimal(exam.maximum or Decimal("0"))
            values.extend((score, maximum, "응시"))
            score_sum += score
            maximum_sum += maximum
            attended += 1
        values.extend(
            (
                attended,
                score_sum,
                maximum_sum,
                None if maximum_sum == 0 else score_sum / maximum_sum * Decimal("100"),
            )
        )
        score_sheet.append(values)
        if name_status == "이름불일치":
            for exam in ordered:
                _review(
                    review_rows,
                    exam,
                    None,
                    "",
                    student_id,
                    "이름불일치",
                    candidates,
                    ("이름불일치",),
                )

    review_sheet.append(
        ["시험ID", "리비전", "WorkItemID", "원본파일명", "유효학번", "학번상태", "이름후보", "사유"]
    )
    for review_row in sorted(
        review_rows,
        key=lambda item: (
            _exam_key(next(exam for exam in ordered if exam.snapshot.session_id == item[0])),
            item[4] is None,
            item[4] or "",
            item[2] is None,
            item[2] or "",
            min(_REASON_ORDER[reason] for reason in item[7]),
        ),
    ):
        review_sheet.append(
            [
                _text(cell) if isinstance(cell, str) else cell
                for cell in (*review_row[:-1], ",".join(review_row[-1]))
            ]
        )

    digest = sha256(
        "\n".join(
            f"{exam.snapshot.session_id}:{exam.snapshot.revision}:{exam.snapshot.generation_id}:{exam.snapshot.manifest_sha256}"
            for exam in ordered
        ).encode()
    ).hexdigest()
    custom_doc_props = cast(Any, workbook).custom_doc_props
    custom_doc_props.append(StringProperty(name="schema", value="1"))
    custom_doc_props.append(StringProperty(name="provenance_tuple_digest", value=digest))
    custom_doc_props.append(StringProperty(name="generation_time", value=generated_at))
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
