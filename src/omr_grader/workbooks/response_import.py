"""Fail-closed response-XLSX parser for response-xlsx-policy-v1."""

from __future__ import annotations

import re
import xml.etree.ElementTree as element_tree
from hashlib import sha256
from io import BytesIO
from typing import BinaryIO, Protocol, runtime_checkable
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from omr_grader.domain.enums import AnswerStatus, SourceKind
from omr_grader.domain.errors import Err, ErrorInfo, Ok, Result
from omr_grader.domain.models import SCHEMA_VERSION, AnswerValue, ImportedResponseRef
from omr_grader.domain.session import stable_work_item_id

from .schemas import (
    MAX_COMPRESSED_BYTES,
    MAX_COMPRESSION_RATIO,
    MAX_DEFINED_NAMES,
    MAX_ENTRY_UNCOMPRESSED_BYTES,
    MAX_FORMULA_BYTES,
    MAX_FORMULA_CELLS,
    MAX_OTHER_COLUMNS,
    MAX_OTHER_ROWS,
    MAX_RELATIONSHIPS,
    MAX_SELECTED_CELLS,
    MAX_SELECTED_ROWS,
    MAX_SHARED_STRING_BYTES,
    MAX_SHARED_STRING_TOTAL_BYTES,
    MAX_SHARED_STRINGS,
    MAX_SHEETS,
    MAX_TOTAL_UNCOMPRESSED_BYTES,
    MAX_WORKBOOK_CELLS,
    MAX_ZIP_ENTRIES,
    RESPONSE_COLUMN_COUNT,
    RESPONSE_HEADERS,
    has_prohibited_control,
    normalize_header,
    normalize_text,
)

_CELL_REFERENCE = re.compile(r"([A-Z]+)")
_ID = re.compile(r"^[0-9]{8}$")
_DANGEROUS_PARTS = (
    "vbaproject",
    "externallink",
    "connections",
    "oleobject",
    "activex",
    "encryptedpackage",
    "encryptioninfo",
)


def _error(code: str, path: str, **context: str | int | bool | None) -> Err:
    return Err((ErrorInfo(code, f"error.{code.lower()}", path, context),))


def _column_index(reference: str) -> int:
    letters = _CELL_REFERENCE.match(reference)
    if letters is None:
        return 0
    value = 0
    for character in letters.group(1):
        value = value * 26 + ord(character) - ord("A") + 1
    return value


def _local_name(name: str) -> str:
    """Return a case-insensitive XML local name."""
    return name.rsplit("}", 1)[-1].rsplit(":", 1)[-1].casefold()


def _xml_root(payload: bytes, path: str) -> element_tree.Element | Err:
    try:
        return element_tree.fromstring(payload)
    except element_tree.ParseError:
        return _error("XLSX_ARCHIVE_QUOTA", path)


def _scan_package(source: BinaryIO) -> Result[None]:
    source.seek(0)
    payload = source.read(MAX_COMPRESSED_BYTES + 1)
    if len(payload) > MAX_COMPRESSED_BYTES or source.read(1):
        return _error("XLSX_ARCHIVE_QUOTA", "source")
    try:
        archive = ZipFile(BytesIO(payload))
    except BadZipFile:
        return _error("XLSX_ARCHIVE_QUOTA", "source")
    with archive:
        entries = archive.infolist()
        if len(entries) > MAX_ZIP_ENTRIES:
            return _error("XLSX_ARCHIVE_QUOTA", "zip.entries")
        total = 0
        formulas = 0
        formula_bytes = 0
        relationships = 0
        worksheet_cells: dict[str, tuple[int, int, int]] = {}
        for entry in entries:
            if entry.flag_bits & 0x1 or entry.file_size > MAX_ENTRY_UNCOMPRESSED_BYTES:
                return _error("XLSX_ARCHIVE_QUOTA", entry.filename)
            total += entry.file_size
            if total > MAX_TOTAL_UNCOMPRESSED_BYTES or (
                entry.compress_size
                and entry.file_size > entry.compress_size * MAX_COMPRESSION_RATIO
            ):
                return _error("XLSX_ARCHIVE_QUOTA", entry.filename)
            filename = entry.filename.casefold()
            if any(part in filename for part in _DANGEROUS_PARTS):
                return _error("XLSX_DANGEROUS_FEATURE", entry.filename)
            if filename.endswith(".rels"):
                root = _xml_root(archive.read(entry), entry.filename)
                if isinstance(root, Err):
                    return root
                if _local_name(root.tag) != "relationships" or any(
                    _local_name(child.tag) != "relationship"
                    or len(child) != 0
                    or bool((child.text or "").strip())
                    or not all(
                        isinstance(value, str) and value
                        for value in (
                            child.attrib.get("Id"),
                            child.attrib.get("Type"),
                            child.attrib.get("Target"),
                        )
                    )
                    for child in root
                ):
                    return _error("XLSX_ARCHIVE_QUOTA", entry.filename)
                if any(
                    value.casefold() == "external"
                    for child in root
                    for name, value in child.attrib.items()
                    if _local_name(name) == "targetmode"
                ):
                    return _error("XLSX_DANGEROUS_FEATURE", entry.filename)
                relationships += len(root)
                if relationships > MAX_RELATIONSHIPS:
                    return _error("XLSX_ARCHIVE_QUOTA", "relationships")
            if filename.endswith((".xml", ".rels")):
                root = _xml_root(archive.read(entry), entry.filename)
                if isinstance(root, Err):
                    return root
                for item in root.iter():
                    local_name = _local_name(item.tag)
                    if local_name == "f":
                        formulas += 1
                        formula_bytes += len((item.text or "").encode("utf-8"))
                    if local_name == "ddelink":
                        return _error("XLSX_DANGEROUS_FEATURE", entry.filename)
            if filename == "xl/workbook.xml":
                root = _xml_root(archive.read(entry), entry.filename)
                if isinstance(root, Err):
                    return root
                if (
                    sum(1 for item in root.iter() if _local_name(item.tag) == "sheet") > MAX_SHEETS
                    or sum(1 for item in root.iter() if _local_name(item.tag) == "definedname")
                    > MAX_DEFINED_NAMES
                ):
                    return _error("XLSX_SHEET_QUOTA", entry.filename)
            if filename == "xl/sharedstrings.xml":
                root = _xml_root(archive.read(entry), entry.filename)
                if isinstance(root, Err):
                    return root
                strings = [item for item in root.iter() if _local_name(item.tag) == "si"]
                total_text = 0
                if len(strings) > MAX_SHARED_STRINGS:
                    return _error("XLSX_SHARED_STRING_QUOTA", entry.filename)
                for item in strings:
                    text = "".join(item.itertext()).encode("utf-8")
                    total_text += len(text)
                    if (
                        len(text) > MAX_SHARED_STRING_BYTES
                        or total_text > MAX_SHARED_STRING_TOTAL_BYTES
                    ):
                        return _error("XLSX_SHARED_STRING_QUOTA", entry.filename)
            if entry.filename.startswith("xl/worksheets/") and entry.filename.endswith(".xml"):
                root = _xml_root(archive.read(entry), entry.filename)
                if isinstance(root, Err):
                    return root
                cells = list(root.iter("{*}c"))
                rows = {
                    match.group(0)
                    for cell in cells
                    if (match := re.search(r"\d+$", cell.attrib.get("r", ""))) is not None
                }
                columns = max(
                    (_column_index(cell.attrib.get("r", "")) for cell in cells), default=0
                )
                worksheet_cells[entry.filename] = (len(rows), columns, len(cells))
        if formulas:
            return _error(
                "XLSX_FORMULA_QUOTA"
                if formulas > MAX_FORMULA_CELLS or formula_bytes > MAX_FORMULA_BYTES
                else "XLSX_FORMULA_FORBIDDEN",
                "workbook",
            )
        if sum(item[2] for item in worksheet_cells.values()) > MAX_WORKBOOK_CELLS:
            return _error("XLSX_CELL_QUOTA", "workbook")
    source.seek(0)
    return Ok(None)


def _cell_text(cell: object, path: str) -> str | Err:
    if not isinstance(cell, _ValueCell):
        return _error("XLSX_ROW_INVALID", path)
    value = cell.value
    if value is None:
        return ""
    if not isinstance(value, str) or cell.data_type != "s":
        return _error("XLSX_ROW_INVALID", path)
    normalized = normalize_text(value)
    if has_prohibited_control(normalized):
        return _error("XLSX_ROW_INVALID", path)
    return normalized


@runtime_checkable
class _ValueCell(Protocol):
    @property
    def value(self) -> object: ...

    @property
    def data_type(self) -> str: ...


def _answer(cell: object, path: str) -> AnswerValue | Err:
    if not isinstance(cell, _ValueCell):
        return _error("XLSX_ROW_INVALID", path)
    value = cell.value
    if value is None:
        return AnswerValue((), AnswerStatus.BLANK)
    if isinstance(value, bool):
        return _error("XLSX_ROW_INVALID", path)
    if type(value) is int and 1 <= value <= 5:
        return AnswerValue((value,), AnswerStatus.NORMAL)
    if not isinstance(value, str) or cell.data_type != "s":
        return _error("XLSX_ROW_INVALID", path)
    text = normalize_text(value)
    if has_prohibited_control(text) or any(character not in "12345" for character in text):
        return _error("XLSX_ROW_INVALID", path)
    choices = tuple(int(character) for character in text)
    if tuple(sorted(set(choices))) != choices:
        return _error("XLSX_ROW_INVALID", path)
    return AnswerValue(
        choices,
        AnswerStatus.BLANK
        if not choices
        else AnswerStatus.NORMAL
        if len(choices) == 1
        else AnswerStatus.MULTIPLE,
    )


def parse_response_book(
    source: BinaryIO, *, sheet_name: str, session_id: str, source_sha256: str
) -> Result[tuple[ImportedResponseRef, ...]]:
    """Parse one already-pinned source; no filesystem path is reopened."""
    scanned = _scan_package(source)
    if isinstance(scanned, Err):
        return scanned
    workbook = None
    try:
        source.seek(0)
        workbook = load_workbook(source, read_only=True, data_only=False, keep_links=False)
    except (BadZipFile, InvalidFileException, OSError) as error:
        return _error("XLSX_ARCHIVE_QUOTA", "source", cause=type(error).__name__)
    try:
        if len(workbook.sheetnames) > MAX_SHEETS or sheet_name not in workbook.sheetnames:
            return _error("XLSX_SHEET_QUOTA", "sheet_name")
        sheet = workbook[sheet_name]
        if getattr(sheet, "merged_cells", None) and sheet.merged_cells.ranges:
            return _error("XLSX_HEADER_INVALID", "sheet.merged_cells")
        header = tuple(
            normalize_header(cell.value)
            for cell in next(sheet.iter_rows(min_row=1, max_row=1, max_col=RESPONSE_COLUMN_COUNT))
        )
        if (
            header != RESPONSE_HEADERS
            or sheet.max_column != RESPONSE_COLUMN_COUNT
            or sheet.max_row > MAX_SELECTED_ROWS
        ):
            return _error("XLSX_HEADER_INVALID", "A1:DA1")
        if sheet.max_row * RESPONSE_COLUMN_COUNT > MAX_SELECTED_CELLS:
            return _error("XLSX_CELL_QUOTA", sheet_name)
        for other_name in workbook.sheetnames:
            other = workbook[other_name]
            if other_name != sheet_name and (
                other.max_row > MAX_OTHER_ROWS or other.max_column > MAX_OTHER_COLUMNS
            ):
                return _error("XLSX_SHEET_QUOTA", other_name)
        rows: list[ImportedResponseRef] = []
        serials: set[int] = set()
        for row_number, cells in enumerate(
            sheet.iter_rows(min_row=2, max_col=RESPONSE_COLUMN_COUNT), 2
        ):
            if all(cell.value is None for cell in cells):
                if rows:
                    return _error("XLSX_ROW_INVALID", f"{sheet_name}!{row_number}")
                continue
            serial = cells[0].value
            if type(serial) is not int or serial < 1 or serial in serials:
                return _error("XLSX_ROW_INVALID", f"{sheet_name}!{row_number}:serial")
            serials.add(serial)
            raw_id = cells[2].value
            if raw_id is not None and (
                cells[2].data_type != "s"
                or not isinstance(raw_id, str)
                or (_ID.fullmatch(raw_id) is None and raw_id != "")
            ):
                return _error("XLSX_ID_TEXT_REQUIRED", f"{sheet_name}!{row_number}:C")
            fields = []
            for index in (1, 2, 3, 104):
                parsed_field = _cell_text(cells[index], f"{sheet_name}!{row_number}:{index + 1}")
                if isinstance(parsed_field, Err):
                    return parsed_field
                fields.append(parsed_field)
            answers: list[AnswerValue] = []
            for question, cell in enumerate(cells[4:104], 1):
                parsed_answer = _answer(cell, f"{sheet_name}!{row_number}:Q{question}")
                if isinstance(parsed_answer, Err):
                    return parsed_answer
                answers.append(parsed_answer)
            item_id = stable_work_item_id(
                session_id, SourceKind.IMPORTED_XLSX, source_sha256, None, None, len(rows), 0
            )
            if not isinstance(item_id, Ok):
                return item_id
            rows.append(
                ImportedResponseRef(
                    schema_version=SCHEMA_VERSION,
                    work_item_id=item_id.value,
                    source_sha256=source_sha256,
                    sheet_name=sheet_name,
                    row_number=row_number,
                    input_ordinal=len(rows),
                    serial=serial,
                    source_filename=fields[0],
                    raw_student_id=fields[1],
                    name=fields[2],
                    answers=tuple(answers),
                    note=fields[3],
                )
            )
        return Ok(tuple(rows))
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except OSError:
                pass


def response_source_sha256(source: BinaryIO) -> str:
    source.seek(0)
    digest = sha256(source.read()).hexdigest()
    source.seek(0)
    return digest


__all__ = ["parse_response_book", "response_source_sha256"]
