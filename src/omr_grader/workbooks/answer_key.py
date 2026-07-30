"""Strict XLSX policy and sample generator for answer-key workbooks."""

from __future__ import annotations

import os
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from io import BytesIO
from pathlib import Path
from tempfile import NamedTemporaryFile
from zipfile import ZIP_DEFLATED, ZIP_STORED, BadZipFile, ZipFile, ZipInfo

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from omr_grader.domain.enums import AnswerKeySnapshotKind, AnswerStatus, KeyQuestionStatus
from omr_grader.domain.errors import Err, ErrorInfo, Ok, Result
from omr_grader.domain.models import SCHEMA_VERSION, AnswerKeyEntry, AnswerKeySnapshot, AnswerValue
from omr_grader.workbooks.schemas import validate_opc_relationships

NORMALIZATION_VERSION = "answer-key-xlsx-v1"
ANSWER_KEY_HEADERS = ("문항번호", "정답", "배점")
MAX_PACKAGE_BYTES = 16 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 64 * 1024 * 1024
MAX_PACKAGE_MEMBERS = 2_000
MAX_ROWS = 102
MAX_RELATIONSHIP_BYTES = 1024 * 1024


def _error(code: str, field: str | None = None) -> ErrorInfo:
    return ErrorInfo(code, f"error.{code.lower()}", field)


def _has_external_relationship(archive: ZipFile, member: ZipInfo) -> Result[bool]:
    if member.file_size > MAX_RELATIONSHIP_BYTES:
        return Err((_error("XLSX_PACKAGE_QUOTA", "path"),))
    payload = archive.read(member)
    if len(payload) > MAX_RELATIONSHIP_BYTES:
        return Err((_error("XLSX_PACKAGE_QUOTA", "path"),))
    return validate_opc_relationships(payload)


def _package(path: str) -> Result[tuple[bytes, str]]:
    try:
        data = Path(path).read_bytes()
    except OSError:
        return Err((_error("XLSX_READ_FAILED", "path"),))
    if not data or len(data) > MAX_PACKAGE_BYTES:
        return Err((_error("XLSX_PACKAGE_QUOTA", "path"),))
    try:
        with ZipFile(BytesIO(data)) as archive:
            members = archive.infolist()
            if any(
                item.flag_bits & 0x1 or item.compress_type not in (ZIP_STORED, ZIP_DEFLATED)
                for item in members
            ):
                return Err((_error("XLSX_INVALID_PACKAGE", "path"),))
            if (
                len(members) > MAX_PACKAGE_MEMBERS
                or sum(item.file_size for item in members) > MAX_UNCOMPRESSED_BYTES
            ):
                return Err((_error("XLSX_PACKAGE_QUOTA", "path"),))
            names = {item.filename.lower() for item in members}
            if any(
                b"<f" in archive.read(item).lower()
                for item in members
                if item.filename.lower().startswith("xl/worksheets/")
            ):
                return Err((_error("XLSX_FORMULA_FORBIDDEN", "path"),))
            for item in members:
                if not item.filename.lower().endswith(".rels"):
                    continue
                external_relationship = _has_external_relationship(archive, item)
                if isinstance(external_relationship, Err):
                    return external_relationship
                if external_relationship.value:
                    return Err((_error("XLSX_FORBIDDEN_FEATURE", "path"),))
    except (BadZipFile, NotImplementedError, RuntimeError):
        return Err((_error("XLSX_INVALID_PACKAGE", "path"),))
    forbidden = (
        "vbaproject.bin",
        "encryptioninfo",
        "encryptedpackage",
        "externallinks/",
        "connections",
        "oleobject",
        "dde",
        "activex",
    )
    if any(any(token in name for token in forbidden) for name in names):
        return Err((_error("XLSX_FORBIDDEN_FEATURE", "path"),))
    return Ok((data, sha256(data).hexdigest()))


def _canonical_points(value: object, field: str) -> Result[str]:
    if type(value) not in (int, float) or isinstance(value, bool):
        return Err((_error("XLSX_CELL_TYPE", field),))
    try:
        number = Decimal(str(value))
    except InvalidOperation:
        return Err((_error("XLSX_POINTS_INVALID", field),))
    if not number.is_finite() or number < 0:
        return Err((_error("XLSX_POINTS_INVALID", field),))
    text = (
        format(number, "f").rstrip("0").rstrip(".")
        if "." in format(number, "f")
        else format(number, "f")
    )
    if "." in text and len(text.split(".", 1)[1]) > 12:
        return Err((_error("XLSX_POINTS_INVALID", field),))
    return Ok("0" if text in {"", "-0"} else text)


def _answer(value: str, field: str) -> Result[tuple[AnswerValue, KeyQuestionStatus]]:
    if value == "":
        return Ok((AnswerValue((), AnswerStatus.UNASKED), KeyQuestionStatus.UNASKED))
    if value in {"0", "전체"}:
        return Ok((AnswerValue((), AnswerStatus.ALL), KeyQuestionStatus.ALL))
    if (
        not value.isascii()
        or not value.isdigit()
        or "0" in value
        or any(char not in "12345" for char in value)
    ):
        return Err((_error("XLSX_ANSWER_INVALID", field),))
    choices = tuple(sorted({int(char) for char in value}))
    if len(choices) != len(value):
        return Err((_error("XLSX_ANSWER_INVALID", field),))
    status = AnswerStatus.NORMAL if len(choices) == 1 else AnswerStatus.MULTIPLE
    return Ok((AnswerValue(choices, status), KeyQuestionStatus.ANSWER))


def import_answer_key(path: str, sheet_name: str) -> Result[AnswerKeySnapshot]:
    """Validate an answer key without executing formulas or trusting workbook metadata."""
    package = _package(path)
    if isinstance(package, Err):
        return package
    data, digest = package.value
    workbook = None
    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=False, keep_links=False)
    except (BadZipFile, InvalidFileException, OSError):
        return Err((_error("XLSX_INVALID_WORKBOOK", "path"),))
    try:
        if sheet_name not in workbook.sheetnames:
            return Err((_error("XLSX_SHEET_NOT_FOUND", "sheet_name"),))
        sheet = workbook[sheet_name]
        if sheet.max_row > MAX_ROWS or sheet.max_column != 3:
            return Err((_error("XLSX_DIMENSION_QUOTA", "sheet_name"),))
        header = next(sheet.iter_rows(min_row=1, max_row=1))
        if (
            any(cell.data_type == "f" or type(cell.value) is not str for cell in header)
            or tuple(cell.value for cell in header) != ANSWER_KEY_HEADERS
        ):
            return Err((_error("XLSX_HEADERS_INVALID", "header"),))
        parsed: dict[int, AnswerKeyEntry] = {}
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, max_col=3), 2):
            question_cell, answer_cell, points_cell = row
            if any(cell.data_type == "f" for cell in row):
                return Err((_error("XLSX_FORMULA_FORBIDDEN", f"row[{row_number}]"),))
            if all(cell.value is None for cell in row):
                continue
            if type(question_cell.value) is not int or isinstance(question_cell.value, bool):
                return Err((_error("XLSX_CELL_TYPE", f"A{row_number}"),))
            question = question_cell.value
            if not 1 <= question <= 100 or question in parsed:
                return Err((_error("XLSX_QUESTION_INVALID", f"A{row_number}"),))
            if answer_cell.value is None:
                answer = _answer("", f"B{row_number}")
            elif type(answer_cell.value) is str:
                answer = _answer(answer_cell.value, f"B{row_number}")
            elif type(answer_cell.value) is int:
                answer = _answer(str(answer_cell.value), f"B{row_number}")
            else:
                return Err((_error("XLSX_CELL_TYPE", f"B{row_number}"),))
            if isinstance(answer, Err):
                return answer
            points = (
                Ok("0")
                if answer.value[1] is KeyQuestionStatus.UNASKED
                and points_cell.value is None
                else _canonical_points(points_cell.value, f"C{row_number}")
            )
            if isinstance(points, Err):
                return points
            value, status = answer.value
            if status is KeyQuestionStatus.UNASKED and points.value != "0":
                return Err((_error("XLSX_UNASKED_POINTS", f"C{row_number}"),))
            parsed[question] = AnswerKeyEntry(question, value, points.value, status)
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except OSError:
                pass
    entries = tuple(
        parsed.get(
            question,
            AnswerKeyEntry(
                question, AnswerValue((), AnswerStatus.UNASKED), "0", KeyQuestionStatus.UNASKED
            ),
        )
        for question in range(1, 101)
    )
    return Ok(
        AnswerKeySnapshot(
            SCHEMA_VERSION,
            AnswerKeySnapshotKind.WORKBOOK,
            Path(path).name,
            digest,
            sheet_name,
            NORMALIZATION_VERSION,
            entries,
            (),
        )
    )


def answer_key_sample_bytes(sheet_name: str = "정답표") -> bytes:
    """Return a deterministic, formula-free 50-question sample workbook."""
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(ANSWER_KEY_HEADERS)
    for question in range(1, 51):
        sheet.append((question, 1, 1))
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def write_answer_key_sample(path: str, sheet_name: str = "정답표") -> Result[None]:
    """Atomically write the sample; callers never observe a partial workbook."""
    target = Path(path)
    temporary_name: str | None = None
    try:
        target.parent.mkdir(parents=True, exist_ok=True)
        with NamedTemporaryFile("wb", dir=target.parent, delete=False) as temporary:
            temporary_name = temporary.name
            temporary.write(answer_key_sample_bytes(sheet_name))
            temporary.flush()
            os.fsync(temporary.fileno())
        os.replace(temporary_name, target)
        temporary_name = None
    except (OSError, TypeError, ValueError):
        return Err((_error("XLSX_WRITE_FAILED", "path"),))
    finally:
        if temporary_name is not None:
            try:
                Path(temporary_name).unlink(missing_ok=True)
            except OSError:
                pass
    return Ok(None)


validate_answer_key_workbook = import_answer_key
