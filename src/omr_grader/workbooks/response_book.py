"""Deterministic response-result workbook projection; never an authority reader."""

from __future__ import annotations

import os
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, cast
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.packaging.custom import StringProperty
from openpyxl.worksheet.worksheet import Worksheet

from omr_grader.domain.models import EffectiveResponse, ImportedResponseRef

from .schemas import RESPONSE_HEADERS, RESPONSE_SHEET_NAME, escape_formula_text


def response_projection_filename(exam_name: str, committed_at: datetime) -> str:
    """Use the immutable generation time once, displayed in Korean local time."""
    if committed_at.tzinfo is None:
        raise ValueError("committed_at must be timezone-aware")
    stamp = committed_at.astimezone(ZoneInfo("Asia/Seoul")).strftime("%y%m%d_%H%M%S")
    return f"01_ocr_{exam_name}_{stamp}_응답결과.xlsx"


def _text(cell: Cell, value: str) -> None:
    cell.value = escape_formula_text(value)
    cell.data_type = "s"


def _cell(sheet: Worksheet, row: int, column: int) -> Cell:
    return cast(Cell, sheet.cell(row, column))


def write_response_projection(
    destination: Path,
    rows: tuple[ImportedResponseRef, ...],
    *,
    session_id: str,
    revision: int,
    manifest_sha256: str,
) -> None:
    """Write the exact A1:DA response projection from immutable imported truth."""
    workbook = Workbook(write_only=False)
    sheet = workbook.active
    if not isinstance(sheet, Worksheet):
        raise RuntimeError("new workbook must have an active worksheet")
    sheet.title = RESPONSE_SHEET_NAME
    for column, header in enumerate(RESPONSE_HEADERS, 1):
        _text(_cell(sheet, 1, column), header)
    for output_row, row in enumerate(
        sorted(rows, key=lambda item: (item.input_ordinal, item.work_item_id)), 2
    ):
        _cell(sheet, output_row, 1).value = row.serial
        _text(_cell(sheet, output_row, 2), row.source_filename)
        _text(_cell(sheet, output_row, 3), row.raw_student_id)
        _text(_cell(sheet, output_row, 4), row.name)
        for question, answer in enumerate(row.answers, 5):
            _text(
                _cell(sheet, output_row, question),
                "".join(str(choice) for choice in answer.choices),
            )
        _text(_cell(sheet, output_row, 105), row.note)
    workbook.properties.creator = "OMR Grader"
    custom_doc_props = cast(Any, workbook).custom_doc_props
    custom_doc_props.append(StringProperty(name="schema", value="1"))
    custom_doc_props.append(StringProperty(name="session_id", value=session_id))
    custom_doc_props.append(StringProperty(name="revision", value=str(revision)))
    custom_doc_props.append(StringProperty(name="manifest_sha256", value=manifest_sha256))
    with tempfile.NamedTemporaryFile(
        dir=destination.parent, prefix=f".{destination.name}.", suffix=".tmp", delete=False
    ) as temporary:
        temporary_path = Path(temporary.name)
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, destination)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()


def write_effective_response_projection(
    destination: Path,
    rows: tuple[EffectiveResponse, ...],
    *,
    session_id: str,
    revision: int,
    manifest_sha256: str,
    names_by_student_id: dict[str, str],
) -> None:
    """Write corrected effective responses as the durable regrade input workbook."""
    workbook = Workbook(write_only=False)
    sheet = workbook.active
    if not isinstance(sheet, Worksheet):
        raise RuntimeError("new workbook must have an active worksheet")
    sheet.title = RESPONSE_SHEET_NAME
    for column, header in enumerate(RESPONSE_HEADERS, 1):
        _text(_cell(sheet, 1, column), header)
    for serial, row in enumerate(rows, 1):
        output_row = serial + 1
        _cell(sheet, output_row, 1).value = serial
        _text(_cell(sheet, output_row, 2), row.source_label)
        _text(_cell(sheet, output_row, 3), row.student_id or "")
        _text(
            _cell(sheet, output_row, 4),
            names_by_student_id.get(row.student_id, "") if row.student_id is not None else "",
        )
        for question, answer in enumerate(row.answers, 5):
            _text(
                _cell(sheet, output_row, question),
                "".join(str(choice) for choice in answer.choices),
            )
        _text(
            _cell(sheet, output_row, 105),
            "수동 수정 반영" if row.corrected_targets else "",
        )
    workbook.properties.creator = "OMR Grader"
    custom_doc_props = cast(Any, workbook).custom_doc_props
    custom_doc_props.append(StringProperty(name="schema", value="1"))
    custom_doc_props.append(StringProperty(name="session_id", value=session_id))
    custom_doc_props.append(StringProperty(name="revision", value=str(revision)))
    custom_doc_props.append(StringProperty(name="manifest_sha256", value=manifest_sha256))
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)


__all__ = [
    "response_projection_filename",
    "write_effective_response_projection",
    "write_response_projection",
]
