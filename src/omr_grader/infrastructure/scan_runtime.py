"""Filesystem-backed scan task construction and the single scan publication transaction."""

from __future__ import annotations

import hashlib
import json
import tempfile
import uuid
from dataclasses import asdict
from datetime import UTC, datetime
from decimal import Decimal
from enum import Enum
from pathlib import Path
from typing import Any, cast
from zipfile import BadZipFile

import cv2
from openpyxl.utils.exceptions import InvalidFileException

from omr_grader.application.dto import (
    CancelOperationCommand,
    EffectiveResponseProjection,
    ScanCommand,
    SessionCreateResult,
)
from omr_grader.application.ports import ScanUseCase as ScanUseCasePort
from omr_grader.application.scan_use_case import ScanUseCase
from omr_grader.domain.corrections import project_effective_responses
from omr_grader.domain.enums import (
    AnswerKeySnapshotKind,
    AnswerStatus,
    CreationKind,
    KeyQuestionStatus,
    OperationKind,
    RosterSnapshotKind,
    SessionState,
)
from omr_grader.domain.errors import Err, ErrorInfo, Ok, Result
from omr_grader.domain.models import (
    AnswerKeyEntry,
    AnswerKeySnapshot,
    AnswerValue,
    AutomaticPage,
    EffectiveResponse,
    IdentityRecord,
    ImportedResponseRef,
    ManifestFile,
    ManifestSummary,
    PageRef,
    RosterSnapshot,
    SessionManifest,
    SessionRecord,
)
from omr_grader.domain.profile import Profile
from omr_grader.domain.session import build_page_ref
from omr_grader.infrastructure.profile_store import ProfileStore
from omr_grader.infrastructure.result_layout import (
    COORDINATE_DIR,
    OCR_IMAGE_DIR,
    REVIEW_DIR,
    ocr_filename,
    result_base_name,
    safe_exam_name,
)
from omr_grader.infrastructure.session_store import SessionStore
from omr_grader.ingestion.images import ScanInput, enumerate_image_folder, enumerate_image_paths
from omr_grader.ingestion.pdf import PdfInput, enumerate_pdf, render_pdf_page
from omr_grader.ingestion.roster import NORMALIZATION_VERSION, import_roster
from omr_grader.recognition.pipeline import (
    PipelineFailure,
    PipelineInput,
    PipelineResult,
    PipelineSuccess,
)
from omr_grader.recognition.thresholds import (
    CALIBRATION_PROVENANCE,
    RecognitionThresholds,
    thresholds_for_sensitivity,
)
from omr_grader.ui.workers import WorkerTask
from omr_grader.workbooks.response_book import write_response_projection


def _error(code: str, field: str) -> Err:
    return Err((ErrorInfo(code, f"error.{code.lower()}", field),))


def _utc() -> str:
    return datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%S.%fZ")


def _json_bytes(value: object) -> bytes:
    return (
        json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")) + "\n"
    ).encode()


def _semantic_sha(value: object) -> str:
    def wire(item: object) -> object:
        if isinstance(item, Enum):
            return item.value
        if isinstance(item, Decimal):
            text = format(item, "f").rstrip("0").rstrip(".")
            return text or "0"
        if isinstance(item, dict):
            return {str(key): wire(entry) for key, entry in item.items()}
        if isinstance(item, list | tuple):
            return [wire(entry) for entry in item]
        return item

    return hashlib.sha256(
        json.dumps(
            wire(asdict(cast(Any, value)) if hasattr(value, "__dataclass_fields__") else value),
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode()
    ).hexdigest()


def _unset_answer_key() -> AnswerKeySnapshot:
    return AnswerKeySnapshot(
        1,
        AnswerKeySnapshotKind.UNSET,
        None,
        None,
        None,
        "unset-v1",
        tuple(
            AnswerKeyEntry(
                question, AnswerValue((), AnswerStatus.UNASKED), "0", KeyQuestionStatus.UNASKED
            )
            for question in range(1, 101)
        ),
        (),
    )


def _mime(path: str) -> str:
    if path.endswith(".json"):
        return "application/json"
    if path.endswith(".png"):
        return "image/png"
    if path.endswith(".xlsx"):
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return "application/octet-stream"


class ScanRuntime:
    """Concrete main-process adapters. Worker payloads contain values only."""

    def __init__(
        self, profiles: ProfileStore, sessions: SessionStore, *, app_version: str = "1"
    ) -> None:
        self._profiles = profiles
        self._sessions = sessions
        self._prepared: dict[
            str, tuple[Profile, RosterSnapshot, RecognitionThresholds, tuple[PageRef, ...]]
        ] = {}
        self._app_version = app_version

    def build_tasks(self, command: ScanCommand) -> Result[tuple[WorkerTask, ...]]:
        profile = self._profiles.load(command.profile_path)
        if isinstance(profile, Err):
            return profile
        if Path(command.profile_path).suffix.casefold() != ".omrtemplate":
            return _error("PROFILE_SOURCE_INVALID", "profile_path")
        roster = self._roster(command)
        if isinstance(roster, Err):
            return roster
        thresholds = thresholds_for_sensitivity(
            command.sensitivity,
            calibrated=True,
            calibration_provenance=CALIBRATION_PROVENANCE,
        )
        if isinstance(thresholds, Err):
            return thresholds

        inputs: list[ScanInput | PdfInput] = []
        ordinal = 0
        duplicate_names: dict[str, int] = {}
        for raw_path in command.source.paths:
            path = Path(raw_path)
            if path.is_dir():
                image_batch = enumerate_image_folder(path, command.session_id)
                if isinstance(image_batch, Err):
                    return image_batch
                if image_batch.value.failures:
                    return _error(image_batch.value.failures[0].code, "source")
                inputs.extend(image_batch.value.inputs)
            elif path.suffix.casefold() == ".pdf":
                duplicate = duplicate_names.get(path.name, 0)
                duplicate_names[path.name] = duplicate + 1
                pdf_batch = enumerate_pdf(
                    path, command.session_id, input_ordinal=ordinal, duplicate_ordinal=duplicate
                )
                if isinstance(pdf_batch, Err):
                    return pdf_batch
                inputs.extend(pdf_batch.value.inputs)
                ordinal += len(pdf_batch.value.inputs)
                continue
            else:
                image_batch = enumerate_image_paths((path,), command.session_id)
                if isinstance(image_batch, Err):
                    return image_batch
                if image_batch.value.failures or not image_batch.value.inputs:
                    return _error(
                        image_batch.value.failures[0].code
                        if image_batch.value.failures
                        else "SCAN_SOURCE_INVALID",
                        "source",
                    )
                inputs.extend(image_batch.value.inputs)
            ordinal += 1
        if not inputs:
            return _error("SCAN_SOURCE_EMPTY", "source")

        tasks: list[WorkerTask] = []
        source_duplicates: dict[Path, int] = {}
        display_duplicates: dict[str, int] = {}
        for task_ordinal, item in enumerate(inputs):
            path = item.source_path
            source_key = path.resolve(strict=False)
            task_duplicate = source_duplicates.get(source_key)
            if task_duplicate is None:
                task_duplicate = display_duplicates.get(item.page_ref.source_display_name, 0)
                display_duplicates[item.page_ref.source_display_name] = task_duplicate + 1
                source_duplicates[source_key] = task_duplicate
            reference = build_page_ref(
                session_id=command.session_id,
                source_kind=item.page_ref.source_kind,
                source_sha256=item.source_sha256,
                source_display_name=item.page_ref.source_display_name,
                page_number=item.page_ref.page_number,
                frame_number=item.page_ref.frame_number,
                input_ordinal=task_ordinal,
                duplicate_ordinal=task_duplicate,
            )
            if isinstance(reference, Err):
                return reference
            if isinstance(item, PdfInput):
                rendered = render_pdf_page(item)
                if isinstance(rendered, Err):
                    return rendered
                encoded_ok, encoded = cv2.imencode(".png", rendered.value.pixels)
                if not encoded_ok:
                    return _error("PDF_PAGE_RENDER_FAILED", "source")
                payload = bytes(encoded)
            else:
                try:
                    payload = path.read_bytes()
                except OSError:
                    return _error("SCAN_SOURCE_UNREADABLE", "source")
                if hashlib.sha256(payload).hexdigest() != item.source_sha256:
                    return _error("SCAN_SOURCE_CHANGED", "source")
            try:
                tasks.append(
                    WorkerTask(
                        task_ordinal,
                        PipelineInput(reference.value, payload, profile.value, thresholds.value),
                    )
                )
            except (TypeError, ValueError):
                return _error("SCAN_SOURCE_INVALID", "source")
        self._prepared[command.operation_id] = (
            profile.value,
            roster.value,
            thresholds.value,
            tuple(task.pipeline_input.page_ref for task in tasks),
        )
        return Ok(tuple(tasks))

    def commit_scan(
        self, command: ScanCommand, results: tuple[PipelineResult, ...]
    ) -> Result[SessionCreateResult]:
        prepared = self._prepared.pop(command.operation_id, None)
        if prepared is None:
            return _error("SCAN_PREPARATION_MISSING", "operation_id")
        if not results or not any(isinstance(item, PipelineSuccess) for item in results):
            failures = tuple(item for item in results if isinstance(item, PipelineFailure))
            counts: dict[str, int] = {}
            for failure in failures:
                for error in failure.failure.errors:
                    counts[error.code] = counts.get(error.code, 0) + 1
            failure_codes = ", ".join(
                f"{code}:{count}" for code, count in sorted(counts.items())
            )
            return Err(
                (
                    ErrorInfo(
                        "SCAN_NO_PROCESSABLE_RESULT",
                        "error.scan_no_processable_result",
                        "source",
                        {
                            "failed_pages": len(failures),
                            "failure_codes": failure_codes or "none",
                        },
                    ),
                )
            )
        try:
            profile, roster, thresholds, expected_refs = prepared
        except (TypeError, ValueError):
            return _error("SCAN_PREPARATION_INVALID", "operation_id")
        pages = tuple(item.page for item in results)
        if (
            not isinstance(profile, Profile)
            or not isinstance(roster, RosterSnapshot)
            or not isinstance(thresholds, RecognitionThresholds)
            or len(expected_refs) != len(pages)
            or tuple(page.page_ref for page in pages) != expected_refs
            or any(page.page_ref.session_id != command.session_id for page in pages)
            or any(
                isinstance(item, PipelineFailure) and item.failure.page_ref != item.page.page_ref
                for item in results
            )
        ):
            return _error("SCAN_RESULT_IDENTITY_MISMATCH", "source")
        try:
            created = _utc()
            generated = self._artifacts(command, profile.sha256, roster, results, created)
            if isinstance(generated, Err):
                return generated
            artifacts, canonical_responses, automatic_pages = generated.value
            generation_id = uuid.uuid4().hex
            record = SessionRecord(
                1,
                command.session_id,
                1,
                SessionState.RECOGNIZED,
                command.exam_name,
                command.exam_year,
                command.exam_term,
                created,
                None,
                created,
            )
            answer_key = _unset_answer_key()
            combined = {
                "session": record.to_dict(),
                "roster": roster.to_dict(),
                "responses": [item.to_dict() for item in canonical_responses],
                "scores": None,
                "answer_key": answer_key.to_dict(),
                "failures": [
                    item.failure.to_dict() for item in results if isinstance(item, PipelineFailure)
                ],
            }
            snapshot = {
                "session_id": command.session_id,
                "revision": 1,
                "generation_id": generation_id,
            }
            names = {row.student_id: row.name for row in roster.rows if row.student_id is not None}
            detail_index: list[dict[str, object]] = []
            pages_by_id = {page.page_ref.work_item_id: page for page in automatic_pages}
            for response in canonical_responses:
                page = pages_by_id[response.work_item_id]
                detail_path = f"details/{response.work_item_id}.json"
                image_path = f"images/{response.work_item_id}.png"
                artifacts[f"evidence/{response.work_item_id}.json"] = _json_bytes(page.to_dict())
                artifacts[detail_path] = _json_bytes(
                    {
                        "schema_version": 1,
                        "snapshot": snapshot,
                        "work_item_id": response.work_item_id,
                        "payload": {
                            "response": response.to_dict(),
                            "recognition": page.to_dict(),
                            "evidence_path": f"evidence/{response.work_item_id}.json",
                        },
                    }
                )
                detail_index.append(
                    {
                        "work_item_id": response.work_item_id,
                        "detail_path": detail_path,
                        "image_path": image_path,
                        "student_id": response.student_id,
                        "student_name": names.get(response.student_id)
                        if response.student_id is not None
                        else None,
                        "score": None,
                        "rank": None,
                    }
                )
            artifacts["semantic_inputs.json"] = _json_bytes({"combined": combined})
            artifacts["projection_request.json"] = _json_bytes(
                {
                    "schema_version": 1,
                    "base_snapshot": None,
                    "base_response_ids": [item.work_item_id for item in canonical_responses],
                    "projection": {
                        "automatic_pages": [item.to_dict() for item in automatic_pages],
                        "imported_responses": [],
                        "corrections": [],
                    },
                }
            )
            artifacts["detail_index.json"] = _json_bytes(
                {"schema_version": 1, "snapshot": snapshot, "work_items": detail_index}
            )
            base_ids = tuple(
                sorted((row.work_item_id for row in canonical_responses), key=str.encode)
            )
            files = tuple(
                ManifestFile(path, len(payload), hashlib.sha256(payload).hexdigest(), _mime(path))
                for path, payload in sorted(artifacts.items(), key=lambda pair: pair[0].encode())
            )
            manifest = SessionManifest(
                1,
                command.session_id,
                1,
                generation_id,
                None,
                None,
                None,
                command.operation_id,
                OperationKind.RECOGNIZE,
                self._app_version,
                created,
                SessionState.RECOGNIZED,
                base_ids,
                profile.sha256,
                _semantic_sha(roster),
                _semantic_sha(answer_key),
                str(thresholds.version),
                _semantic_sha(asdict(thresholds)),
                files,
                ManifestSummary(
                    len(canonical_responses),
                    sum(page.processing_status.value == "processed" for page in automatic_pages),
                    sum(
                        page.processing_status.value == "needs_manual_review"
                        for page in automatic_pages
                    ),
                    None,
                ),
            )
            identity = IdentityRecord(1, command.session_id, created, CreationKind.SCAN)
            display_name = result_base_name(command.exam_name, created)
        except (TypeError, ValueError, OSError, json.JSONDecodeError) as exc:
            return _error("SCAN_COMMIT_SERIALIZATION_FAILED", str(exc))
        return self._sessions._create_initial_generation(
            identity=identity,
            manifest=manifest,
            session=record,
            display_name=display_name,
            artifacts=artifacts,
        )

    def _roster(self, command: ScanCommand) -> Result[RosterSnapshot]:
        if command.roster_path is None:
            return Ok(
                RosterSnapshot(
                    1, RosterSnapshotKind.NONE, None, None, None, NORMALIZATION_VERSION, (), ()
                )
            )
        # The scan DTO freezes only a path; the roster import contract uses its first worksheet.
        from openpyxl import load_workbook

        try:
            workbook = load_workbook(
                command.roster_path, read_only=True, data_only=False, keep_links=False
            )
            try:
                if not workbook.sheetnames:
                    return _error("XLSX_SHEET_NOT_FOUND", "roster_path")
                sheet_name = workbook.sheetnames[0]
            finally:
                workbook.close()
        except (BadZipFile, InvalidFileException, OSError, ValueError):
            return _error("XLSX_INVALID_WORKBOOK", "roster_path")
        return import_roster(command.roster_path, sheet_name)

    def _artifacts(
        self,
        command: ScanCommand,
        profile_sha256: str,
        roster: RosterSnapshot,
        results: tuple[PipelineResult, ...],
        created_at: str,
    ) -> Result[tuple[dict[str, bytes], tuple[EffectiveResponse, ...], tuple[AutomaticPage, ...]]]:
        output: dict[str, bytes] = {"recognition/roster.json": _json_bytes(roster.to_dict())}
        automatic_pages = tuple(item.page for item in results if isinstance(item, PipelineSuccess))
        projected = project_effective_responses(
            EffectiveResponseProjection(automatic_pages, (), ()),
            session_id=command.session_id,
            expected_base_revision=1,
        )
        if isinstance(projected, Err):
            return projected
        rows: list[ImportedResponseRef] = []
        for ordinal, result in enumerate(results):
            page = result.page
            output[f"automatic/{page.page_ref.work_item_id}.json"] = _json_bytes(page.to_dict())
            if isinstance(result, PipelineFailure):
                output[f"recognition/failures/{page.page_ref.work_item_id}.json"] = _json_bytes(
                    result.failure.to_dict()
                )
                output[f"{REVIEW_DIR}/{page.page_ref.work_item_id}.json"] = _json_bytes(
                    result.failure.to_dict()
                )
                for source in command.source.paths:
                    candidate = Path(source)
                    if candidate.is_file() and candidate.name == page.page_ref.source_display_name:
                        output[f"{REVIEW_DIR}/{safe_exam_name(candidate.name)}"] = (
                            candidate.read_bytes()
                        )
                        break
                continue
            key = page.page_ref.work_item_id
            output[f"recognition/{key}/normalized.png"] = result.artifacts.normalized_png
            output[f"recognition/{key}/coordinates.json"] = result.artifacts.coordinates_json
            output[f"recognition/{key}/overlay.png"] = result.artifacts.overlay_png
            output[f"images/{key}.png"] = result.artifacts.normalized_png
            display_stem = f"{ordinal + 1:03}_{key}"
            output[f"{OCR_IMAGE_DIR}/{display_stem}.png"] = result.artifacts.overlay_png
            output[f"{COORDINATE_DIR}/{display_stem}.json"] = (
                result.artifacts.coordinates_json
            )
            student_id = page.student_id.value or ""
            rows.append(
                ImportedResponseRef(
                    1,
                    key,
                    page.page_ref.source_sha256,
                    "recognition",
                    ordinal + 1,
                    ordinal,
                    ordinal + 1,
                    page.page_ref.source_display_name,
                    student_id,
                    "미등록",
                    tuple(answer.value for answer in page.answers),
                    "",
                )
            )
        try:
            with tempfile.TemporaryDirectory() as directory:
                book = Path(directory) / "responses.xlsx"
                write_response_projection(
                    book,
                    tuple(rows),
                    session_id=command.session_id,
                    revision=1,
                    manifest_sha256="0" * 64,
                )
                output["responses.xlsx"] = book.read_bytes()
                output[ocr_filename(command.exam_name, created_at)] = output["responses.xlsx"]
        except (OSError, TypeError, ValueError) as exc:
            return _error("SCAN_RESPONSE_PROJECTION_FAILED", str(exc))
        output["recognition/projection.json"] = _json_bytes(
            {
                "profile_sha256": profile_sha256,
                "roster": roster.to_dict(),
                "pages": [item.page.to_dict() for item in results],
            }
        )
        output.setdefault(
            f"{REVIEW_DIR}/현재_수동확인_항목_없음.txt",
            "현재 수동 확인이 필요한 항목이 없습니다.\n".encode(),
        )
        return Ok((output, projected.value, automatic_pages))


class ScanControllerAdapter:
    """Adapt the orchestration's explicit commit authority to the UI's unary port."""

    def __init__(self, scan: ScanUseCase, coordinator: ScanRuntime) -> None:
        self._scan = scan
        self._coordinator = coordinator

    def run_scan(self, command: ScanCommand) -> Result[SessionCreateResult]:
        return self._scan.run_scan(command, self._coordinator)

    def cancel_scan(self, command: CancelOperationCommand) -> Result[None]:
        return self._scan.cancel_scan(command)


def bind_scan_runtime(runtime: ScanRuntime) -> ScanUseCasePort:
    """Return a controller-facing scan port that cannot bypass ``commit_scan``."""
    return ScanControllerAdapter(ScanUseCase(runtime), runtime)


ScanTaskSource = ScanRuntime
ScanCommitAdapter = ScanRuntime
ConcreteScanTaskSource = ScanRuntime
ConcreteScanCommitAdapter = ScanRuntime
ConcreteScanControllerAdapter = ScanControllerAdapter
